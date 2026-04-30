from django.shortcuts import get_object_or_404, redirect, render
import json
import secrets
import smtplib
import time
import ftplib
import socket
import os
import re
import shutil
import subprocess
import fnmatch
import tempfile
import uuid
import threading
import sys
import zipfile
from email.mime.text import MIMEText
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from django.http import JsonResponse, HttpRequest, HttpResponse
from django.conf import settings
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_exempt
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from django.views.decorators.http import require_POST
from .firebase_service import ContactConfigService, RobotService, ProcessService, QueueService, ReportService, ScheduleService, SAPTemplateService
from .forms import FTPAccountForm, MailAccountForm, TelegramBotForm, TelegramGroupForm
from .models import (
	FTPAccount,
	MailAccount,
	RobotAgent,
	RobotAgentEvent,
	RobotAgentRelease,
	RobotJob,
	SapProcess,
	SapProcessStep,
	TelegramBot,
	TelegramBotButton,
	TelegramBotMenu,
	TelegramGroup,
)
from .sap_service import SAPScanService
from .sap_keyboard_utils import build_sendkeys_from_config
from .sap_popup_utils import collect_popup_controls, select_popup_radio_by_id, fill_popup_input_value
from .windows_dialog_utils import scan_visible_dialogs

try:
	import paramiko
except Exception:
	paramiko = None

try:
	import tkinter as tk
except Exception:
	tk = None

try:
	from tkinter import messagebox as tk_messagebox
except Exception:
	tk_messagebox = None

try:
	from openpyxl import load_workbook
except Exception:
	load_workbook = None


_PROCESS_RUNTIME = {}
_PROCESS_RUNTIME_LOCK = threading.Lock()


def _runtime_init(process_id, process_name, total_steps):
	pid = int(process_id)
	with _PROCESS_RUNTIME_LOCK:
		_PROCESS_RUNTIME[pid] = {
			'process_id': pid,
			'process_name': str(process_name or ''),
			'total_steps': int(total_steps or 0),
			'current_step': 0,
			'step_name': '',
			'paused': False,
			'stop_requested': False,
			'running': True,
			'updated_at': datetime.now().isoformat(),
			'logs': [],
		}


def _runtime_get(process_id):
	pid = int(process_id)
	with _PROCESS_RUNTIME_LOCK:
		state = _PROCESS_RUNTIME.get(pid)
		if not state:
			return None
		return {
			'process_id': state.get('process_id'),
			'process_name': state.get('process_name', ''),
			'total_steps': int(state.get('total_steps', 0) or 0),
			'current_step': int(state.get('current_step', 0) or 0),
			'step_name': state.get('step_name', ''),
			'paused': bool(state.get('paused')),
			'stop_requested': bool(state.get('stop_requested')),
			'running': bool(state.get('running')),
			'updated_at': state.get('updated_at'),
			'logs': list(state.get('logs') or []),
		}


def _runtime_touch(process_id):
	pid = int(process_id)
	with _PROCESS_RUNTIME_LOCK:
		state = _PROCESS_RUNTIME.get(pid)
		if not state:
			return
		state['updated_at'] = datetime.now().isoformat()


def _runtime_set_controls(process_id, *, paused=None, stop_requested=None):
	pid = int(process_id)
	with _PROCESS_RUNTIME_LOCK:
		state = _PROCESS_RUNTIME.get(pid)
		if not state:
			return
		if paused is not None:
			state['paused'] = bool(paused)
		if stop_requested is not None:
			state['stop_requested'] = bool(stop_requested)
		state['updated_at'] = datetime.now().isoformat()


def _runtime_set_step(process_id, step_no, total_steps, step_name):
	pid = int(process_id)
	with _PROCESS_RUNTIME_LOCK:
		state = _PROCESS_RUNTIME.get(pid)
		if not state:
			return
		state['current_step'] = int(step_no or 0)
		state['total_steps'] = int(total_steps or state.get('total_steps', 0) or 0)
		state['step_name'] = str(step_name or '')
		state['updated_at'] = datetime.now().isoformat()


def _runtime_push_log(process_id, text):
	msg = str(text or '').strip()
	if not msg:
		return
	pid = int(process_id)
	with _PROCESS_RUNTIME_LOCK:
		state = _PROCESS_RUNTIME.get(pid)
		if not state:
			return
		logs = state.setdefault('logs', [])
		logs.append({
			'step': int(state.get('current_step', 0) or 0),
			'msg': msg,
		})
		if len(logs) > 120:
			del logs[:-120]
		state['updated_at'] = datetime.now().isoformat()


def _runtime_finish(process_id):
	pid = int(process_id)
	with _PROCESS_RUNTIME_LOCK:
		state = _PROCESS_RUNTIME.get(pid)
		if not state:
			return
		state['running'] = False
		state['updated_at'] = datetime.now().isoformat()


def get_dashboard_stats():
	"""Fetch dashboard statistics from Firebase"""
	robots = RobotService.get_all_robots() or {}
	processes = ProcessService.get_all_processes() or {}
	
	total_robots = len(robots)
	online_robots = sum(1 for r in robots.values() if isinstance(r, dict) and r.get('status') == 'online')
	total_processes = len(processes)
	
	return {
		'total_robots': total_robots or 5,
		'online_robots': online_robots or 5,
		'total_processes': total_processes or 12,
		'success_rate': '98.2%',
	}


def dashboard(request):
	stats = get_dashboard_stats()
	robots_data = RobotService.get_all_robots() or {}
	processes_data = ProcessService.get_all_processes() or {}
	
	return render(
		request,
		'core/dashboard.html',
		{
			'page_title': 'Dashboard',
			'page_subtitle': 'Robot operasyon merkezi',
			'stats': stats,
			'robots': robots_data,
			'processes': processes_data,
		},
	)


def robots(request):
	robots_data = RobotService.get_all_robots() or {}
	
	return render(
		request,
		'core/section.html',
		{
			'page_title': 'Robotlar',
			'page_subtitle': 'Robot envanteri ve canli durum takibi',
			'section_type': 'robots',
			'data': robots_data,
		},
	)


def processes(request):
	processes_data = ProcessService.get_all_processes() or {}
	
	return render(
		request,
		'core/section.html',
		{
			'page_title': 'Surecler',
			'page_subtitle': 'Süreç performansı ve hata analizi',
			'section_type': 'processes',
			'data': processes_data,
		},
	)


def queues(request):
	queues_data = QueueService.get_all_queues() or {}
	
	return render(
		request,
		'core/section.html',
		{
			'page_title': 'Kuyruk Yonetimi',
			'page_subtitle': 'Is kuyruklari, onceliklendirme ve SLA takibi',
			'section_type': 'queues',
			'data': queues_data,
		},
	)


def scheduler(request):
	schedules_data = ScheduleService.get_all_schedules() or {}
	
	return render(
		request,
		'core/section.html',
		{
			'page_title': 'Scheduler',
			'page_subtitle': 'Planli gorevler ve zamanlama panosu',
			'section_type': 'scheduler',
			'data': schedules_data,
		},
	)


def reports(request):
	reports_data = ReportService.get_all_reports() or {}
	
	return render(
		request,
		'core/section.html',
		{
			'page_title': 'Raporlar',
			'page_subtitle': 'Operasyon raporlari ve KPI ozetleri',
			'section_type': 'reports',
			'data': reports_data,
		},
	)


def settings_page(request):
	return render(
		request,
		'core/section.html',
		{
			'page_title': 'Ayarlar',
			'page_subtitle': 'Sistem konfigrasyonlari ve entegrasyonlar',
			'section_type': 'settings',
		},
	)


def robot_control_center(request):
	"""Robot ajanları ve iş kuyruğunu yönetmek için operasyon ekranı."""
	agents = list(RobotAgent.objects.all().order_by('code').values('code', 'name'))
	processes = list(SapProcess.objects.all().order_by('name').values('id', 'name'))
	releases = list(
		RobotAgentRelease.objects.all().order_by('-created_at').values(
			'version', 'is_active', 'is_mandatory', 'download_url', 'setup_file', 'install_command', 'created_at'
		)
	)
	return render(
		request,
		'core/robot_control_center.html',
		{
			'current': 'robot_control_center',
			'page_title': 'Robot Operasyon Merkezi',
			'page_subtitle': 'Ajan durumlarını izle, kuyruk işlerini yönet ve yeni iş ata',
			'agents_json': json.dumps(agents, ensure_ascii=False),
			'processes_json': json.dumps(processes, ensure_ascii=False),
			'releases_json': json.dumps(releases, ensure_ascii=False, default=str),
		},
	)


def _extract_telegram_http_error(err):
	"""Telegram HTTPError gövdesindeki açıklamayı çıkarıp okunur bir mesaj döndürür."""
	status = getattr(err, 'code', 'unknown')
	description = ''

	try:
		raw = err.read()
		if raw:
			payload = json.loads(raw.decode('utf-8', errors='replace') or '{}')
			description = str(payload.get('description') or '').strip()
	except Exception:
		pass

	if description:
		return f'HTTP {status}: {description}'

	reason = str(getattr(err, 'reason', '') or '').strip()
	if reason:
		return f'HTTP {status}: {reason}'

	return f'HTTP {status}'


def _send_telegram_group_test(group, payload=None):
	payload = payload or {}
	bot = group.default_bot
	if not bot or not bot.is_active:
		return False, 'Bu grup icin aktif varsayilan bot yok.'

	token = bot.get_bot_token()
	if not token:
		return False, 'Bot token cozulmedi veya bos.'

	custom_text = str(payload.get('test_message') or '').strip()
	text = custom_text or f"Saggio RPA test mesaji\nGrup: {group.name}\nBot: {bot.name}"
	ok, detail = _send_telegram_message(bot, group.chat_id, text)
	if ok:
		return True, 'Test mesaji gonderildi.'
	return False, f'Telegram hatasi: {detail}'


def _send_mail_test(account, payload=None):
	payload = payload or {}
	password = account.get_smtp_password()
	if not password:
		return False, 'SMTP şifresi çözülemedi veya boş.'

	from_display = account.from_name or account.name
	from_value = f'{from_display} <{account.email}>'
	to_value = str(payload.get('test_to') or '').strip() or account.email
	subject = str(payload.get('test_subject') or '').strip() or 'Saggio RPA - Test Maili'
	body = str(payload.get('test_body') or '').strip() or 'Bu e-posta Saggio RPA tarafindan test amacli gonderildi.'

	msg = MIMEText(body, 'plain', 'utf-8')
	msg['Subject'] = subject
	msg['From'] = from_value
	msg['To'] = to_value

	try:
		if account.use_ssl:
			server = smtplib.SMTP_SSL(account.smtp_host, account.smtp_port, timeout=15)
		else:
			server = smtplib.SMTP(account.smtp_host, account.smtp_port, timeout=15)
			if account.use_tls:
				server.starttls()

		with server:
			server.login(account.smtp_username, password)
			server.sendmail(account.email, [to_value], msg.as_string())
		return True, f'Test maili gonderildi: {to_value}'
	except Exception as e:
		return False, f'SMTP test hatası: {e}'


def _send_ftp_test(account, payload=None):
	payload = payload or {}
	password = account.get_password()
	if not password:
		return False, 'FTP sifresi çözülemedi veya boş.'

	remote_path = str(payload.get('test_path') or '').strip() or account.remote_base_path or '.'
	protocol = str(account.protocol or 'sftp').lower()

	try:
		if protocol == 'sftp':
			if paramiko is None:
				return False, 'SFTP testi için paramiko kurulu değil.'
			transport = paramiko.Transport((account.host, int(account.port or 22)))
			transport.connect(username=account.username, password=password)
			sftp = paramiko.SFTPClient.from_transport(transport)
			sftp.listdir(remote_path)
			sftp.close()
			transport.close()
			return True, f'SFTP bağlantısı başarılı. Yol erişimi: {remote_path}'

		if protocol == 'ftps':
			ftp = ftplib.FTP_TLS()
			ftp.connect(account.host, int(account.port or 21), timeout=15)
			ftp.login(account.username, password)
			ftp.prot_p()
			ftp.cwd(remote_path)
			ftp.quit()
			return True, f'FTPS bağlantısı başarılı. Yol erişimi: {remote_path}'

		ftp = ftplib.FTP()
		ftp.connect(account.host, int(account.port or 21), timeout=15)
		ftp.login(account.username, password)
		ftp.cwd(remote_path)
		ftp.quit()
		return True, f'FTP bağlantısı başarılı. Yol erişimi: {remote_path}'
	except (socket.timeout, ConnectionRefusedError) as e:
		return False, f'FTP bağlantı hatası: {e}'
	except Exception as e:
		return False, f'FTP test hatası: {e}'


def _send_telegram_message(bot, chat_id, text):
	if not bot or not bot.is_active:
		return False, 'Telegram bot aktif degil veya secilmedi.'
	if not chat_id:
		return False, 'Telegram chat id secilmedi.'
	token = bot.get_bot_token()
	if not token:
		return False, 'Telegram token cozulmedi.'

	msg_text = str(text or '')
	base_data = {'chat_id': str(chat_id), 'text': msg_text}
	data = dict(base_data)
	if bot.default_parse_mode:
		data['parse_mode'] = bot.default_parse_mode

	def _call_send_message(payload):
		req = Request(
			f'https://api.telegram.org/bot{token}/sendMessage',
			data=urlencode(payload).encode('utf-8'),
			method='POST',
		)
		with urlopen(req, timeout=10) as resp:
			return json.loads(resp.read().decode('utf-8') or '{}')

	try:
		payload = _call_send_message(data)
		if payload.get('ok'):
			return True, 'telegram_ok'
		desc = str(payload.get('description', 'telegram_api_error') or '')
		# parse_mode kaynaklı hata olursa sade metin ile yeniden dene.
		if data.get('parse_mode') and ('parse entities' in desc.casefold() or 'can\'t parse' in desc.casefold()):
			payload2 = _call_send_message(base_data)
			if payload2.get('ok'):
				return True, 'telegram_ok_plain'
			desc = str(payload2.get('description', 'telegram_api_error') or '')

		# Uzun mesaj veya genel API hatalarında güvenli fallback: düz metin + kısaltılmış içerik.
		if len(msg_text) > 3900:
			truncated_data = {'chat_id': str(chat_id), 'text': (msg_text[:3900] + '...')}
			payload3 = _call_send_message(truncated_data)
			if payload3.get('ok'):
				return True, 'telegram_ok_truncated'
			desc = str(payload3.get('description', desc or 'telegram_api_error') or '')
		return False, desc or 'telegram_api_error'
	except HTTPError as e:
		return False, _extract_telegram_http_error(e)
	except URLError as e:
		return False, f'Baglanti hatasi: {e.reason}'
	except Exception as e:
		return False, str(e)


def _send_telegram_voice_message(bot, chat_id, text):
	"""Windows SAPI ile ses üretip Telegram'a gerçek voice note olarak gönderir."""
	if not bot or not bot.is_active:
		return False, 'Telegram bot aktif degil veya secilmedi.'
	if not chat_id:
		return False, 'Telegram chat id secilmedi.'
	token = bot.get_bot_token()
	if not token:
		return False, 'Telegram token cozulmedi.'

	voice_text = str(text or '').strip()
	if not voice_text:
		return False, 'Sesli mesaj metni boş.'

	tmp_wav_path = None
	tmp_ogg_path = None
	tmp_mp3_path = None
	co_initialized = False
	errors = []

	def _send_voice_file(path_value, filename_value, content_type_value):
		with open(path_value, 'rb') as fp:
			file_data = fp.read()

		boundary = f'----SaggioBoundary{uuid.uuid4().hex}'
		parts = []

		def _add_text(name, value):
			parts.append(f'--{boundary}'.encode('utf-8'))
			parts.append(f'Content-Disposition: form-data; name="{name}"'.encode('utf-8'))
			parts.append(b'')
			parts.append(str(value).encode('utf-8'))

		_add_text('chat_id', str(chat_id))
		_add_text('caption', 'Saggio RPA sesli bildirim')

		parts.append(f'--{boundary}'.encode('utf-8'))
		parts.append(f'Content-Disposition: form-data; name="voice"; filename="{filename_value}"'.encode('utf-8'))
		parts.append(f'Content-Type: {content_type_value}'.encode('utf-8'))
		parts.append(b'')
		parts.append(file_data)

		parts.append(f'--{boundary}--'.encode('utf-8'))
		parts.append(b'')
		body = b'\r\n'.join(parts)

		req = Request(
			f'https://api.telegram.org/bot{token}/sendVoice',
			data=body,
			method='POST',
			headers={'Content-Type': f'multipart/form-data; boundary={boundary}'},
		)
		with urlopen(req, timeout=20) as resp:
			payload = json.loads(resp.read().decode('utf-8') or '{}')
			if payload.get('ok'):
				return True, 'telegram_voice_ok'
			return False, payload.get('description', 'telegram_voice_api_error')

	def _send_audio_file(path_value, filename_value, content_type_value):
		with open(path_value, 'rb') as fp:
			file_data = fp.read()

		boundary = f'----SaggioBoundary{uuid.uuid4().hex}'
		parts = []

		def _add_text(name, value):
			parts.append(f'--{boundary}'.encode('utf-8'))
			parts.append(f'Content-Disposition: form-data; name="{name}"'.encode('utf-8'))
			parts.append(b'')
			parts.append(str(value).encode('utf-8'))

		_add_text('chat_id', str(chat_id))
		_add_text('caption', 'Saggio RPA sesli bildirim')

		parts.append(f'--{boundary}'.encode('utf-8'))
		parts.append(f'Content-Disposition: form-data; name="audio"; filename="{filename_value}"'.encode('utf-8'))
		parts.append(f'Content-Type: {content_type_value}'.encode('utf-8'))
		parts.append(b'')
		parts.append(file_data)

		parts.append(f'--{boundary}--'.encode('utf-8'))
		parts.append(b'')
		body = b'\r\n'.join(parts)

		req = Request(
			f'https://api.telegram.org/bot{token}/sendAudio',
			data=body,
			method='POST',
			headers={'Content-Type': f'multipart/form-data; boundary={boundary}'},
		)
		with urlopen(req, timeout=20) as resp:
			payload = json.loads(resp.read().decode('utf-8') or '{}')
			if payload.get('ok'):
				return True, 'telegram_audio_ok'
			return False, payload.get('description', 'telegram_audio_api_error')
	try:
		try:
			import pythoncom
			import win32com.client

			pythoncom.CoInitialize()
			co_initialized = True
			with tempfile.NamedTemporaryFile(suffix='.wav', delete=False) as tmpf:
				tmp_wav_path = tmpf.name
			with tempfile.NamedTemporaryFile(suffix='.ogg', delete=False) as tmpf:
				tmp_ogg_path = tmpf.name

			voice = win32com.client.Dispatch('SAPI.SpVoice')
			# Türkçe sese öncelik ver (kuruluysa)
			try:
				voices = voice.GetVoices()
				selected_voice = None
				count = int(getattr(voices, 'Count', 0) or 0)
				for i in range(count):
					try:
						candidate = voices.Item(i)
						desc = str(candidate.GetDescription() or '').casefold()
						lang = str(candidate.GetAttribute('Language') or '').casefold()
						if 'turkish' in desc or 'türk' in desc or '041f' in lang:
							selected_voice = candidate
							break
					except Exception:
						continue
				if selected_voice is not None:
					voice.Voice = selected_voice
			except Exception:
				pass
			try:
				voice.Rate = -1
			except Exception:
				pass
			stream = win32com.client.Dispatch('SAPI.SpFileStream')
			# 3 = SSFMCreateForWrite
			stream.Open(tmp_wav_path, 3, False)
			voice.AudioOutputStream = stream
			voice.Speak(voice_text)
			stream.Close()

			ffmpeg_path = shutil.which('ffmpeg')
			if ffmpeg_path:
				convert_cmd = [
					ffmpeg_path,
					'-y',
					'-i',
					tmp_wav_path,
					'-c:a',
					'libopus',
					'-b:a',
					'24k',
					tmp_ogg_path,
				]
				convert_result = subprocess.run(convert_cmd, capture_output=True, text=True)
				if convert_result.returncode == 0 and os.path.exists(tmp_ogg_path) and os.path.getsize(tmp_ogg_path) > 0:
					ok, detail = _send_voice_file(tmp_ogg_path, 'saggio_notification.ogg', 'audio/ogg')
					if ok:
						return True, detail
					errors.append(f'ogg_send_failed: {detail}')
				else:
					err = (convert_result.stderr or convert_result.stdout or 'ffmpeg convert hatasi').strip()
					errors.append(f'ffmpeg_convert_failed: {err}')
			else:
				errors.append('ffmpeg_not_found')
		except Exception as sapi_ex:
			errors.append(f'sapi_flow_failed: {sapi_ex}')

		# SAPI/ffmpeg başarısız olursa gTTS ile MP3 üretip sendVoice dene.
		try:
			from gtts import gTTS
			with tempfile.NamedTemporaryFile(suffix='.mp3', delete=False) as tmpf:
				tmp_mp3_path = tmpf.name
			clean_text = voice_text.replace('*', '').replace('_', '')
			gTTS(text=clean_text, lang='tr', slow=False).save(tmp_mp3_path)
			ok, detail = _send_voice_file(tmp_mp3_path, 'saggio_notification.mp3', 'audio/mpeg')
			if ok:
				return True, detail
			errors.append(f'gtts_send_failed: {detail}')
			# Voice endpoint reddederse normal audio olarak düşür.
			a_ok, a_msg = _send_audio_file(tmp_mp3_path, 'saggio_notification.mp3', 'audio/mpeg')
			if a_ok:
				return True, a_msg
			errors.append(f'gtts_audio_fallback_failed: {a_msg}')
		except Exception as ge:
			errors.append(f'gtts_fallback_failed: {ge}')

		return False, 'Voice note gonderilemedi: ' + ' | '.join(errors)
	except Exception as e:
		return False, str(e)
	finally:
		if co_initialized:
			try:
				pythoncom.CoUninitialize()
			except Exception:
				pass
		if tmp_wav_path and os.path.exists(tmp_wav_path):
			try:
				os.remove(tmp_wav_path)
			except Exception:
				pass
		if tmp_ogg_path and os.path.exists(tmp_ogg_path):
			try:
				os.remove(tmp_ogg_path)
			except Exception:
				pass
		if tmp_mp3_path and os.path.exists(tmp_mp3_path):
			try:
				os.remove(tmp_mp3_path)
			except Exception:
				pass


@require_POST
def _sap_process_scan_popups_impl(request, process_id):
	"""Açık SAP popup pencerelerini ve içeriklerindeki buton/radio alanlarını tarar."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	steps = _extract_runtime_steps(body, proc)
	if not steps:
		return JsonResponse({'ok': False, 'error': 'Önce en az bir adım tanımlayın.'}, status=400)

	conn, conn_err = _resolve_connection_from_steps(steps)
	if conn_err:
		return JsonResponse({'ok': False, 'error': conn_err}, status=400)

	service = SAPScanService()
	ok, payload = service.apply_to_screen(
		sys_id=conn.get('sys_id', ''),
		client=conn.get('client', ''),
		user=conn.get('user', ''),
		pwd=conn.get('pwd', ''),
		lang=conn.get('lang', 'TR'),
		t_code='',
		root_id=conn.get('root_id', 'wnd[0]'),
		extra_wait=conn.get('extra_wait', 0),
		actions=[],
		execute_f8=False,
	)
	if not ok:
		return JsonResponse({'ok': False, 'error': payload}, status=500)

	session = getattr(service, 'session', None)
	if session is None:
		return JsonResponse({'ok': False, 'error': 'Aktif SAP session bulunamadı.'}, status=500)

	popups = []
	seen = set()
	try:
		children = getattr(session, 'Children', None)
		count = int(getattr(children, 'Count', 0) or 0) if children is not None else 0
		for idx in range(1, count):
			wnd = None
			try:
				wnd = children(idx)
			except Exception:
				try:
					wnd = children.Item(idx)
				except Exception:
					wnd = None
			if wnd is None:
				continue

			popup_id = _normalize_session_element_id(str(getattr(wnd, 'Id', '') or '').strip())
			title = str(getattr(wnd, 'Text', '') or '').strip()
			legacy_text = _collect_popup_text_legacy(session, popup_id or 'wnd[1]', limit=220)
			message_text = _collect_popup_message_text(session, popup_id or 'wnd[1]', limit=120)
			deep_text = _collect_node_text(wnd, limit=220)
			text = ' | '.join([p for p in [legacy_text, message_text, deep_text] if p])
			key = f'{popup_id}|{title}|{text}'.casefold()
			if key in seen:
				continue
			seen.add(key)

			buttons, radios, inputs = collect_popup_controls(wnd, _normalize_session_element_id, _iter_children)

			popups.append({
				'id': popup_id,
				'title': title,
				'text': text,
				'buttons': buttons,
				'radios': radios,
				'inputs': inputs,
				'label': f'{title or popup_id or "Popup"} [{popup_id}]',
			})
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Popup tarama hatası: {ex}'}, status=500)

	return JsonResponse({'ok': True, 'popups': popups, 'count': len(popups)})


@require_POST
def sap_process_scan_popups(request, process_id):
	"""Açık SAP popup pencerelerini ve içeriklerindeki butonları tarar (hata güvenli JSON)."""
	try:
		return _sap_process_scan_popups_impl(request, process_id)
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Popup tarama beklenmeyen hata: {ex}'}, status=500)


def _send_mail_message(account, to_value, subject, body, attachment_path=None):
	if not account or not account.is_active:
		return False, 'Mail hesabi aktif degil veya secilmedi.'
	password = account.get_smtp_password()
	if not password:
		return False, 'SMTP şifresi çözülemedi.'

	from_display = account.from_name or account.name
	from_value = f'{from_display} <{account.email}>'

	if attachment_path:
		from email.mime.multipart import MIMEMultipart
		from email.mime.base import MIMEBase
		from email import encoders
		msg = MIMEMultipart()
		msg.attach(MIMEText(body, 'plain', 'utf-8'))
		try:
			import os as _os
			with open(attachment_path, 'rb') as f:
				part = MIMEBase('application', 'octet-stream')
				part.set_payload(f.read())
			encoders.encode_base64(part)
			part.add_header('Content-Disposition', f'attachment; filename="{_os.path.basename(attachment_path)}"')
			msg.attach(part)
		except Exception as e:
			return False, f'Ek dosya okunamadı: {e}'
	else:
		msg = MIMEText(body, 'plain', 'utf-8')

	msg['Subject'] = subject
	msg['From'] = from_value
	msg['To'] = to_value

	try:
		if account.use_ssl:
			server = smtplib.SMTP_SSL(account.smtp_host, account.smtp_port, timeout=15)
		else:
			server = smtplib.SMTP(account.smtp_host, account.smtp_port, timeout=15)
			if account.use_tls:
				server.starttls()

		with server:
			server.login(account.smtp_username, password)
			server.sendmail(account.email, [to_value], msg.as_string())
		return True, 'mail_ok'
	except Exception as e:
		return False, str(e)


def _generate_row_report_xlsx(row_results, output_path=None):
	"""row_results listesinden xlsx rapor dosyası üretir. Dosya yolu döndürür."""
	import tempfile
	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill, Alignment

	wb = Workbook()
	ws = wb.active
	ws.title = 'Rapor'

	headers = ['#', 'Cari / Değer', 'Excel Satır No', 'Durum', 'Neden / Notlar', 'Zaman']
	header_font = Font(bold=True, color='FFFFFF')
	header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
	for col_idx, h in enumerate(headers, 1):
		cell = ws.cell(row=1, column=col_idx, value=h)
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

	status_colors = {'ok': 'C8E6C9', 'skip': 'FFF9C4', 'error': 'FFCDD2'}
	for row_idx, r in enumerate(row_results or [], start=2):
		status = str(r.get('status', '') or '').lower()
		fill_color = status_colors.get(status, 'FFFFFF')
		row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
		values = [
			row_idx - 1,
			str(r.get('cari', '') or ''),
			r.get('excel_row', ''),
			str(r.get('status', '') or '').upper(),
			str(r.get('reason', '') or ''),
			str(r.get('timestamp', '') or ''),
		]
		for col_idx, val in enumerate(values, 1):
			cell = ws.cell(row=row_idx, column=col_idx, value=val)
			cell.fill = row_fill

	col_widths = [6, 30, 14, 10, 50, 20]
	for col_idx, w in enumerate(col_widths, 1):
		ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

	# ── 2. Sheet: Doldurma Detayları ─────────────────────────────────────
	detail_rows = []
	for r in (row_results or []):
		for fd in (r.get('filled_data') or []):
			detail_rows.append({
				'excel_row': r.get('excel_row', ''),
				'cari': str(r.get('cari', '') or ''),
				'alan': str(fd.get('alan', '') or ''),
				'kaynak': str(fd.get('kaynak', '') or ''),
				'deger': str(fd.get('deger', '') or ''),
			})

	if detail_rows:
		ws2 = wb.create_sheet(title='Doldurma Detayları')
		d_headers = ['Excel Satır No', 'Cari / Değer', 'SAP Alan ID', 'Kaynak', 'Doldurma Değeri']
		for col_idx, h in enumerate(d_headers, 1):
			cell = ws2.cell(row=1, column=col_idx, value=h)
			cell.font = header_font
			cell.fill = header_fill
			cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
		for row_idx, d in enumerate(detail_rows, start=2):
			ws2.cell(row=row_idx, column=1, value=d['excel_row'])
			ws2.cell(row=row_idx, column=2, value=d['cari'])
			ws2.cell(row=row_idx, column=3, value=d['alan'])
			ws2.cell(row=row_idx, column=4, value=d['kaynak'])
			ws2.cell(row=row_idx, column=5, value=d['deger'])
		for col_idx, w in enumerate([14, 28, 55, 18, 40], 1):
			ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = w

	if not output_path:
		tmp = tempfile.NamedTemporaryFile(delete=False, suffix='_rapor.xlsx', prefix='saggio_rpa_')
		output_path = tmp.name
		tmp.close()

	wb.save(output_path)
	return output_path


def _resolve_placeholders(text, runtime_state):
	"""runtime_state değerlerini metindeki {cari}, {excel_satir} vb. yer tutuculara yazar."""
	if not text:
		return text
	text = text.replace('{cari}', str(runtime_state.get('loop_value', '') or ''))
	text = text.replace('{excel_satir}', str((runtime_state.get('excel_loop_index', 0) or 0) + 1))
	text = text.replace('{excel_index}', str(runtime_state.get('excel_loop_index', 0) or 0))
	text = text.replace('{loop_index}', str(runtime_state.get('loop_index', 0) or 0))
	text = text.replace('{popup_text}', str(runtime_state.get('last_popup_text', '') or ''))
	text = text.replace('{status_text}', str(runtime_state.get('last_status_text', '') or ''))
	text = text.replace('{status_type}', str(runtime_state.get('last_status_type', '') or ''))
	return text


def _notify_sap_event(notification, phase, result_payload=None):
	notification = notification or {}
	notes = []
	stamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

	tg_bot_id = notification.get('telegram_bot_id')
	tg_group_id = notification.get('telegram_group_id')
	tg_voice_enabled = bool(notification.get('telegram_voice_enabled'))
	mail_id = notification.get('mail_account_id')

	bot = TelegramBot.objects.filter(pk=tg_bot_id, is_active=True).first() if tg_bot_id else None
	group = TelegramGroup.objects.filter(pk=tg_group_id, is_active=True).first() if tg_group_id else None
	mail_account = MailAccount.objects.filter(pk=mail_id, is_active=True).first() if mail_id else None

	if bot and group:
		default_tg = f'SAP sureci {phase} ({stamp})'
		if phase == 'start':
			text = str(notification.get('telegram_start_message') or '').strip() or default_tg
		else:
			total = len(result_payload or [])
			ok_count = sum(1 for r in (result_payload or []) if r.get('ok'))
			err_count = max(0, total - ok_count)
			end_default = f'SAP süreci tamamlandı ({stamp}) | Başarılı: {ok_count}, Hatalı: {err_count}'
			text = str(notification.get('telegram_end_message') or '').strip() or end_default
		ok, msg = _send_telegram_message(bot, group.chat_id, text)
		notes.append({'channel': 'telegram', 'ok': ok, 'msg': msg})
		if tg_voice_enabled:
			v_ok, v_msg = _send_telegram_voice_message(bot, group.chat_id, text)
			if (not v_ok) and ok:
				notes.append({'channel': 'telegram_voice', 'ok': True, 'msg': f'Voice gonderilemedi ama metin gonderildi: {v_msg}'})
			else:
				notes.append({'channel': 'telegram_voice', 'ok': v_ok, 'msg': v_msg})

	if mail_account:
		to_value = str(notification.get('mail_to') or '').strip() or mail_account.email
		subject = str(notification.get('mail_subject') or '').strip() or 'Saggio RPA SAP Bildirimi'
		if phase == 'start':
			body = str(notification.get('mail_start_message') or '').strip() or f'SAP sureci basladi. Zaman: {stamp}'
			ok, msg = _send_mail_message(mail_account, to_value, subject, body)
		else:
			total = len(result_payload or [])
			ok_count = sum(1 for r in (result_payload or []) if r.get('ok'))
			err_count = max(0, total - ok_count)
			body = str(notification.get('mail_end_message') or '').strip() or f'SAP süreci tamamlandı. Başarılı: {ok_count}, Hatalı: {err_count}. Zaman: {stamp}'
			report_path = None
			if result_payload:
				try:
					report_path = _generate_row_report_xlsx(result_payload)
				except Exception:
					report_path = None
			ok, msg = _send_mail_message(mail_account, to_value, subject, body, attachment_path=report_path)
			if report_path:
				try:
					import os as _os2
					_os2.unlink(report_path)
				except Exception:
					pass
		notes.append({'channel': 'mail', 'ok': ok, 'msg': msg})

	return notes


def _manage_contact_entity(request, *, model, form_class, page_title, page_subtitle, success_message, test_handler=None, firebase_sync_handler=None, firebase_entity_name=None, current_key=''):
	edit_id = request.GET.get('edit')
	edit_instance = model.objects.filter(pk=edit_id).first() if edit_id else None
	create_form = form_class(prefix='create')
	edit_form = form_class(instance=edit_instance, prefix='edit') if edit_instance else None
	error_message = ''

	if request.method == 'POST':
		action = request.POST.get('action', '')
		if action == 'create':
			create_form = form_class(request.POST, prefix='create')
			if create_form.is_valid():
				obj = create_form.save()
				if firebase_sync_handler is not None:
					firebase_sync_handler(obj)
				return redirect(f'{request.path}?ok=create')
			error_message = 'Kayit olusturulamadi. Alanlari kontrol edin.'
		elif action == 'update':
			object_id = request.POST.get('object_id', '')
			instance = get_object_or_404(model, pk=object_id)
			edit_form = form_class(request.POST, instance=instance, prefix='edit')
			if edit_form.is_valid():
				obj = edit_form.save()
				if firebase_sync_handler is not None:
					firebase_sync_handler(obj)
				return redirect(f'{request.path}?ok=update')
			error_message = 'Kayıt güncellenemedi. Alanları kontrol edin.'
		elif action == 'delete':
			object_id = request.POST.get('object_id', '')
			instance = get_object_or_404(model, pk=object_id)
			if firebase_entity_name:
				ContactConfigService.delete_entity(firebase_entity_name, instance.id)
			instance.delete()
			return redirect(f'{request.path}?ok=delete')
		elif action == 'test' and test_handler is not None:
			object_id = request.POST.get('object_id', '')
			instance = get_object_or_404(model, pk=object_id)
			ok, msg = test_handler(instance, request.POST)
			state = 'ok' if ok else 'err'
			query = urlencode({'test': state, 'msg': msg or ''})
			return redirect(f'{request.path}?{query}')

	ok_action = request.GET.get('ok', '')
	status_message = ''
	test_state = request.GET.get('test', '')
	test_msg = request.GET.get('msg', '')
	if ok_action == 'create':
		status_message = f'{success_message} olusturuldu.'
	elif ok_action == 'update':
		status_message = f'{success_message} güncellendi.'
	elif ok_action == 'delete':
		status_message = f'{success_message} silindi.'
	elif test_state == 'ok':
		status_message = test_msg or 'Test islemi basarili.'
	elif test_state == 'err':
		error_message = test_msg or 'Test islemi basarisiz.'

	return render(
		request,
		'core/contact_crud.html',
		{
			'current': current_key,
			'page_title': page_title,
			'page_subtitle': page_subtitle,
			'entity_name': success_message,
			'entity_type': model._meta.model_name,
			'items': model.objects.all(),
			'create_form': create_form,
			'edit_form': edit_form,
			'edit_instance': edit_instance,
			'status_message': status_message,
			'error_message': error_message,
		},
	)


def telegram_bots_manage(request):
	return _manage_contact_entity(
		request,
		model=TelegramBot,
		form_class=TelegramBotForm,
		page_title='Telegram Botlari',
		page_subtitle='Bot token, parse mode ve aktiflik durumunu yonetin',
		success_message='Telegram botu',
		current_key='telegram_bots_manage',
		firebase_sync_handler=ContactConfigService.sync_telegram_bot,
		firebase_entity_name='telegram_bots',
	)


def telegram_groups_manage(request):
	return _manage_contact_entity(
		request,
		model=TelegramGroup,
		form_class=TelegramGroupForm,
		page_title='Telegram Gruplari',
		page_subtitle='Sahip ekipler, chat id ve varsayilan bot baglantisini yonetin',
		success_message='Telegram grubu',
		current_key='telegram_groups_manage',
		test_handler=_send_telegram_group_test,
		firebase_sync_handler=ContactConfigService.sync_telegram_group,
		firebase_entity_name='telegram_groups',
	)


def mail_accounts_manage(request):
	return _manage_contact_entity(
		request,
		model=MailAccount,
		form_class=MailAccountForm,
		page_title='Mail Hesaplari',
		page_subtitle='SMTP hesaplari ve gonderim ayarlarini yonetin',
		success_message='Mail hesabi',
		current_key='mail_accounts_manage',
		test_handler=_send_mail_test,
		firebase_sync_handler=ContactConfigService.sync_mail_account,
		firebase_entity_name='mail_accounts',
	)


def ftp_accounts_manage(request):
	return _manage_contact_entity(
		request,
		model=FTPAccount,
		form_class=FTPAccountForm,
		page_title='FTP Hesaplari',
		page_subtitle='FTP/SFTP baglanti profillerini yonetin',
		success_message='FTP hesabi',
		current_key='ftp_accounts_manage',
		test_handler=_send_ftp_test,
		firebase_sync_handler=ContactConfigService.sync_ftp_account,
		firebase_entity_name='ftp_accounts',
	)


def sap_scan(request):
	service = SAPScanService()
	sys_options = service.get_system_list()

	form_data = {
		'sys_id': request.POST.get('sys_id', (sys_options[0] if sys_options else '02-QA')),
		'client': request.POST.get('client', '300'),
		'lang': request.POST.get('lang', 'TR'),
		'user': request.POST.get('user', ''),
		'pwd': request.POST.get('pwd', ''),
		't_code': request.POST.get('t_code', 'ZFI0501N'),
		'root_id': request.POST.get('root_id', 'wnd[0]'),
		'extra_wait': request.POST.get('extra_wait', '0'),
		'loop_values': request.POST.get('loop_values', ''),
	}

	results = []
	error = ''
	total_count = 0
	with_id_count = 0
	editable_count = 0

	if request.method == 'POST':
		missing = [k for k in ('sys_id', 'client', 'user', 'pwd') if not form_data.get(k)]
		if missing:
			error = 'Sistem ID, Client, Kullanıcı ve Şifre alanları zorunludur.'
		else:
			try:
				extra_wait = float(form_data.get('extra_wait') or 0)
				extra_wait = max(0.0, min(5.0, extra_wait))
			except (TypeError, ValueError):
				extra_wait = 0.0

			ok, payload = service.scan_screen(
				sys_id=form_data['sys_id'],
				client=form_data['client'],
				lang=form_data['lang'],
				user=form_data['user'],
				pwd=form_data['pwd'],
				t_code=form_data['t_code'],
				root_id=form_data['root_id'],
				extra_wait=extra_wait,
			)

			if ok:
				results = payload
				for row in results:
					row['entries_json'] = json.dumps(row.get('entries', []), ensure_ascii=False)
				results.sort(key=lambda x: (x.get('level', 0), x.get('type', ''), x.get('id', '')))
				total_count = len(results)
				with_id_count = sum(1 for x in results if x.get('id'))
				editable_count = sum(1 for x in results if x.get('changeable'))
			else:
				error = payload

	return render(
		request,
		'core/sap_scan.html',
		{
			'page_title': 'SAP Derin Tarama',
			'page_subtitle': 'Baglanti bilgileri ile ekrani ac ve teknik ID listesi al',
			'form_data': form_data,
			'results': results,
			'error': error,
			'total_count': total_count,
			'with_id_count': with_id_count,
			'editable_count': editable_count,
			'sys_options': sys_options,
			'telegram_bots': TelegramBot.objects.filter(is_active=True).order_by('name'),
			'telegram_groups': TelegramGroup.objects.filter(is_active=True).order_by('name'),
			'mail_accounts': MailAccount.objects.filter(is_active=True).order_by('name'),
		},
	)

@require_POST
def sap_apply(request):
	"""Secili satirlari SAP ekranina uygula."""
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({"ok": False, "error": "Geçersiz istek verisi."}, status=400)

	conn    = body.get("connection", {})
	actions = body.get("actions", [])
	notification = body.get('notification', {})

	if not actions:
		return JsonResponse({"ok": False, "error": "Uygulanacak aksiyon yok."})

	service = SAPScanService()
	notify_start = _notify_sap_event(notification, 'start')
	ok, payload = service.apply_to_screen(
		sys_id=conn.get("sys_id", ""),
		client=conn.get("client", ""),
		user=conn.get("user", ""),
		pwd=conn.get("pwd", ""),
		lang=conn.get("lang", "TR"),
		t_code=conn.get("t_code", ""),
		root_id=conn.get("root_id", "wnd[0]"),
		extra_wait=conn.get("extra_wait", 0),
		actions=actions,
	)
	notify_end = _notify_sap_event(notification, 'end', payload if ok else [])

	if ok:
		return JsonResponse({"ok": True, "results": payload, 'notifications': {'start': notify_start, 'end': notify_end}})
	return JsonResponse({"ok": False, "error": payload})


@require_POST
def sap_run(request):
	"""Seçili satırları SAP ekranına uygula ve F8 ile çalıştır."""
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({"ok": False, "error": "Geçersiz istek verisi."}, status=400)

	conn = body.get("connection", {})
	actions = body.get("actions", [])
	notification = body.get('notification', {})

	service = SAPScanService()
	notify_start = _notify_sap_event(notification, 'start')
	ok, payload = service.apply_to_screen(
		sys_id=conn.get("sys_id", ""),
		client=conn.get("client", ""),
		user=conn.get("user", ""),
		pwd=conn.get("pwd", ""),
		lang=conn.get("lang", "TR"),
		t_code=conn.get("t_code", ""),
		root_id=conn.get("root_id", "wnd[0]"),
		extra_wait=conn.get("extra_wait", 0),
		actions=actions,
		execute_f8=True,
	)
	notify_end = _notify_sap_event(notification, 'end', payload if ok else [])

	if ok:
		return JsonResponse({"ok": True, "results": payload, 'notifications': {'start': notify_start, 'end': notify_end}})
	return JsonResponse({"ok": False, "error": payload})


def sap_template_list(request):
	"""Firebase'deki SAP şablon adlarını döndür."""
	names = SAPTemplateService.list_template_names()
	return JsonResponse({"ok": True, "names": names})


def sap_template_get(request):
	"""Seçili şablonu Firebase'den getir."""
	name = str(request.GET.get("name", "") or "").strip()
	if not name:
		return JsonResponse({"ok": False, "error": "Şablon adı gerekli."}, status=400)

	tpl = SAPTemplateService.get_template(name)
	if not tpl:
		return JsonResponse({"ok": False, "error": "Şablon bulunamadı."}, status=404)

	state = tpl.get("state", {}) if isinstance(tpl, dict) else {}
	return JsonResponse({"ok": True, "name": name, "state": state})


@require_POST
def sap_template_save(request):
	"""SAP şablonunu Firebase'e kaydet/güncelle."""
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({"ok": False, "error": "Geçersiz istek verisi."}, status=400)

	name = str(body.get("name", "") or "").strip()
	state = body.get("state", {})

	if not name:
		return JsonResponse({"ok": False, "error": "Şablon adı gerekli."}, status=400)

	result = SAPTemplateService.save_template(name, state)
	if not result.get('ok'):
		return JsonResponse({"ok": False, "error": result.get('error', 'Şablon kaydetme hatası.')}, status=500)

	return JsonResponse({
		"ok": True,
		"name": name,
		"storage": result.get('storage', 'unknown'),
		"reason": result.get('reason', ''),
	})


@require_POST
def sap_template_delete(request):
	"""SAP şablonunu Firebase/yerel depodan sil."""
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({"ok": False, "error": "Geçersiz istek verisi."}, status=400)

	name = str(body.get("name", "") or "").strip()
	if not name:
		return JsonResponse({"ok": False, "error": "Şablon adı gerekli."}, status=400)

	result = SAPTemplateService.delete_template(name)
	if not result.get('ok'):
		return JsonResponse({"ok": False, "error": result.get('error', 'Şablon silme hatası.')}, status=500)

	return JsonResponse({"ok": True, "name": name, "storage": result.get('storage', 'unknown'), "reason": result.get('reason', '')})


# ---------------------------------------------------------------------------
# SAP Süreç Builder
# ---------------------------------------------------------------------------

def sap_process_list(request):
	"""SAP süreçlerini listele; yeni süreç oluştur (POST)."""
	if request.method == 'POST':
		name = str(request.POST.get('name', '') or '').strip()
		desc = str(request.POST.get('description', '') or '').strip()
		if not name:
			return JsonResponse({'ok': False, 'error': 'Süreç adı boş olamaz.'}, status=400)
		if SapProcess.objects.filter(name=name).exists():
			return JsonResponse({'ok': False, 'error': 'Bu isimde bir süreç zaten var.'}, status=400)
		proc = SapProcess.objects.create(name=name, description=desc)
		return JsonResponse({'ok': True, 'id': proc.pk, 'name': proc.name})
	# GET: hem JSON hem HTML desteği
	if request.headers.get('Accept') == 'application/json' or request.GET.get('fmt') == 'json':
		processes = list(SapProcess.objects.values('id', 'name', 'description', 'updated_at').order_by('name'))
		return JsonResponse({'ok': True, 'processes': processes})
	return render(request, 'core/sap_process_list.html', {
		'current': 'sap_process',
		'page_title': 'SAP Süreç Tanımları',
		'page_subtitle': 'Otomasyon akışı tanımlayın',
	})


@ensure_csrf_cookie
def sap_process_builder(request, process_id):
	"""Belirli bir sürecin adım builder sayfasını göster."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	steps = list(proc.steps.values('id', 'order', 'step_type', 'label', 'config').order_by('order'))
	template_names = SAPTemplateService.list_template_names()
	ftp_accounts = list(FTPAccount.objects.filter(is_active=True).values('id', 'name', 'protocol', 'host', 'port').order_by('name'))
	other_processes = list(SapProcess.objects.exclude(pk=proc.pk).values('id', 'name').order_by('name'))
	mail_accounts = list(MailAccount.objects.filter(is_active=True).values('id', 'name', 'email').order_by('name'))
	return render(request, 'core/sap_process_builder.html', {
		'current': 'sap_process',
		'page_title': f'Süreç: {proc.name}',
		'page_subtitle': 'Adım adım otomasyon akışı',
		'process': proc,
		'steps_json': json.dumps(steps, default=str),
		'template_names': template_names,
		'ftp_accounts_json': json.dumps(ftp_accounts, default=str),
		'processes_json': json.dumps(other_processes, default=str),
		'mail_accounts_json': json.dumps(mail_accounts, default=str),
	})


def sap_process_backup(request, process_id):
	"""Süreç tanımını ve adımlarını JSON dosyası olarak indir."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	steps = list(proc.steps.values('id', 'order', 'step_type', 'label', 'config').order_by('order'))
	now = datetime.now()
	payload = {
		'version': 1,
		'exported_at': now.isoformat(),
		'process': {
			'id': proc.pk,
			'name': proc.name,
			'description': proc.description,
			'ghost_overlay_enabled': bool(proc.ghost_overlay_enabled),
			'office_express_auto_close': bool(proc.office_express_auto_close),
			'telegram_notifications_enabled': bool(proc.telegram_notifications_enabled),
			'telegram_voice_enabled': bool(proc.telegram_voice_enabled),
			'mail_notifications_enabled': bool(proc.mail_notifications_enabled),
		},
		'steps': steps,
	}
	safe_name = re.sub(r'[^0-9A-Za-z_-]+', '_', str(proc.name or f'process_{proc.pk}')).strip('_') or f'process_{proc.pk}'
	filename = f'sap_process_{proc.pk}_{safe_name}_{now.strftime("%Y%m%d_%H%M%S")}.json'
	body = json.dumps(payload, ensure_ascii=False, indent=2, default=str)
	resp = HttpResponse(body, content_type='application/json; charset=utf-8')
	resp['Content-Disposition'] = f'attachment; filename="{filename}"'
	return resp


def sap_process_delete(request, process_id):
	"""Süreci sil."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	proc.delete()
	return JsonResponse({'ok': True})


@require_POST
def sap_process_step_save(request, process_id):
	"""Adımları toplu kaydet (tam liste — mevcut adımları sil, yenileri ekle)."""
	try:
		proc = get_object_or_404(SapProcess, pk=process_id)
		try:
			body = json.loads(request.body)
		except (json.JSONDecodeError, TypeError):
			return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

		steps_data = body.get('steps', [])
		if not isinstance(steps_data, list):
			return JsonResponse({'ok': False, 'error': 'steps bir liste olmalı.'}, status=400)

		valid_types = {c[0] for c in SapProcessStep.STEP_TYPE_CHOICES}
		proc.steps.all().delete()
		saved = []
		for i, s in enumerate(steps_data):
			step_type = str(s.get('step_type', '') or '').strip()
			if step_type not in valid_types:
				continue
			step = SapProcessStep.objects.create(
				process=proc,
				order=i,
				step_type=step_type,
				label=str(s.get('label', '') or '').strip()[:300],
				config=s.get('config', {}) if isinstance(s.get('config'), dict) else {},
			)
			saved.append({'id': step.pk, 'order': step.order, 'step_type': step.step_type})

		return JsonResponse({'ok': True, 'saved': len(saved), 'steps': saved})
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Adımlar kaydedilemedi: {ex}'}, status=500)


@require_POST
def sap_process_rename(request, process_id):
	"""Süreç adını / açıklamasını güncelle."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)
	name = str(body.get('name', '') or '').strip()
	desc = str(body.get('description', '') or '').strip()
	if not name:
		return JsonResponse({'ok': False, 'error': 'Süreç adı boş olamaz.'}, status=400)
	if SapProcess.objects.exclude(pk=process_id).filter(name=name).exists():
		return JsonResponse({'ok': False, 'error': 'Bu isimde başka bir süreç var.'}, status=400)
	proc.name = name
	proc.description = desc
	proc.save(update_fields=['name', 'description', 'updated_at'])
	return JsonResponse({'ok': True, 'name': proc.name})


@require_POST
def sap_process_runtime_settings_save(request, process_id):
	"""Süreç çalışma ayarlarını (overlay + popup + bildirim) güncelle."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	proc.ghost_overlay_enabled = _as_bool(body.get('ghost_overlay_enabled', proc.ghost_overlay_enabled), proc.ghost_overlay_enabled)
	proc.office_express_auto_close = _as_bool(body.get('office_express_auto_close', proc.office_express_auto_close), proc.office_express_auto_close)
	proc.telegram_notifications_enabled = _as_bool(body.get('telegram_notifications_enabled', proc.telegram_notifications_enabled), proc.telegram_notifications_enabled)
	proc.telegram_voice_enabled = _as_bool(body.get('telegram_voice_enabled', proc.telegram_voice_enabled), proc.telegram_voice_enabled)
	proc.mail_notifications_enabled = _as_bool(body.get('mail_notifications_enabled', proc.mail_notifications_enabled), proc.mail_notifications_enabled)
	proc.sap_retry_enabled = _as_bool(body.get('sap_retry_enabled', proc.sap_retry_enabled), proc.sap_retry_enabled)
	try:
		_ri = int(body.get('sap_retry_interval_minutes', proc.sap_retry_interval_minutes))
		if _ri >= 1:
			proc.sap_retry_interval_minutes = _ri
	except (TypeError, ValueError):
		pass
	try:
		_md = int(body.get('sap_retry_max_duration_minutes', proc.sap_retry_max_duration_minutes))
		if _md >= 0:
			proc.sap_retry_max_duration_minutes = _md
	except (TypeError, ValueError):
		pass

	proc.save(update_fields=[
		'ghost_overlay_enabled',
		'office_express_auto_close',
		'telegram_notifications_enabled',
		'telegram_voice_enabled',
		'mail_notifications_enabled',
		'sap_retry_enabled',
		'sap_retry_interval_minutes',
		'sap_retry_max_duration_minutes',
		'updated_at',
	])

	return JsonResponse({
		'ok': True,
		'ghost_overlay_enabled': proc.ghost_overlay_enabled,
		'office_express_auto_close': proc.office_express_auto_close,
		'telegram_notifications_enabled': proc.telegram_notifications_enabled,
		'telegram_voice_enabled': proc.telegram_voice_enabled,
		'mail_notifications_enabled': proc.mail_notifications_enabled,
		'sap_retry_enabled': proc.sap_retry_enabled,
		'sap_retry_interval_minutes': proc.sap_retry_interval_minutes,
		'sap_retry_max_duration_minutes': proc.sap_retry_max_duration_minutes,
	})


@require_POST
def sap_process_runtime_control(request, process_id):
	"""Canlı süreç kontrolü: duraklat/devam et/durdur."""
	get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	action = str(body.get('action', '') or '').strip().casefold()
	state = _runtime_get(process_id)
	if state is None:
		if action == 'force_stop':
			# Kayıt yoksa zaten temiz; başarı dön.
			return JsonResponse({'ok': True, 'state': {'running': False, 'paused': False, 'stop_requested': False}})
		return JsonResponse({'ok': False, 'error': 'Bu süreç için aktif runtime bulunamadı.'}, status=404)

	if action == 'pause_toggle':
		_runtime_set_controls(process_id, paused=not bool(state.get('paused')), stop_requested=bool(state.get('stop_requested')))
	elif action == 'pause':
		_runtime_set_controls(process_id, paused=True)
	elif action == 'resume':
		_runtime_set_controls(process_id, paused=False)
	elif action == 'stop':
		_runtime_set_controls(process_id, paused=False, stop_requested=True)
	elif action == 'force_stop':
		# Sıkışmış (orphan) runtime kaydını anında temizle.
		_runtime_set_controls(process_id, paused=False, stop_requested=True)
		_runtime_push_log(process_id, 'Kullanıcı zorla sıfırlama yaptı; runtime durumu temizlendi.')
		_runtime_finish(process_id)
	else:
		return JsonResponse({'ok': False, 'error': f'Desteklenmeyen aksiyon: {action}'}, status=400)

	return JsonResponse({'ok': True, 'state': _runtime_get(process_id)})


def sap_process_runtime_status(request, process_id):
	"""Canlı süreç durumunu ve hayalet log akışını döndürür."""
	get_object_or_404(SapProcess, pk=process_id)
	state = _runtime_get(process_id)
	if state is None:
		return JsonResponse({'ok': True, 'state': {'running': False, 'paused': False, 'stop_requested': False, 'current_step': 0, 'total_steps': 0, 'step_name': '', 'logs': []}})
	return JsonResponse({'ok': True, 'state': state})


@require_POST
def sap_process_scan_buttons(request, process_id):
	"""Açık SAP ekranındaki butonları tarayıp process builder için döndürür."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	steps = _extract_runtime_steps(body, proc)
	if not steps:
		return JsonResponse({'ok': False, 'error': 'Önce en az bir adım tanımlayın.'}, status=400)

	conn, conn_err = _resolve_connection_from_steps(steps)
	if conn_err:
		return JsonResponse({'ok': False, 'error': conn_err}, status=400)

	service = SAPScanService()
	use_current_screen = bool(body.get('use_current_screen', True))
	t_code = '' if use_current_screen else str(conn.get('t_code', '') or '').strip()

	# apply_to_screen ile bağlan — session'a erişerek tüm pencereleri (wnd[0]+wnd[1]+...) tarayabilelim
	ok, payload = service.apply_to_screen(
		sys_id=conn.get('sys_id', ''),
		client=conn.get('client', ''),
		user=conn.get('user', ''),
		pwd=conn.get('pwd', ''),
		lang=conn.get('lang', 'TR'),
		t_code=t_code,
		root_id=conn.get('root_id', 'wnd[0]'),
		extra_wait=conn.get('extra_wait', 0),
		actions=[],
		execute_f8=False,
	)
	if not ok:
		return JsonResponse({'ok': False, 'error': payload}, status=500)

	session = getattr(service, 'session', None)
	if session is None:
		return JsonResponse({'ok': False, 'error': 'Aktif SAP session bulunamadı.'}, status=500)

	buttons = []
	seen_ids = set()
	seen_ctx_specs = set()

	def _collect_buttons(node):
		stack = [node] if node is not None else []
		while stack:
			current = stack.pop(0)
			try:
				node_id = str(getattr(current, 'Id', '') or '').strip()
				node_type = str(getattr(current, 'Type', '') or '').strip()
				norm_id = _normalize_session_element_id(node_id)
				lower_type = node_type.casefold()
				lower_id = norm_id.casefold()
				is_button = ('button' in lower_type) or ('/btn[' in lower_id)
				if is_button and norm_id and norm_id not in seen_ids:
					name = str(getattr(current, 'Name', '') or '').strip()
					text = str(getattr(current, 'Text', '') or '').strip()
					seen_ids.add(norm_id)
					buttons.append({
						'id': norm_id,
						'type': node_type,
						'name': name,
						'text': text,
						'label': f'{text or name or node_type or "Buton"} [{norm_id}]',
					})

				# SAP menü öğeleri (GuiMenu) — wnd[0]/mbar/menu[..] yolları
				is_menu_node = (
					('/mbar/' in lower_id or '/menu[' in lower_id)
					or (('menu' in lower_type) and ('menubar' not in lower_type))
				)
				if is_menu_node and norm_id and norm_id not in seen_ids:
					name = str(getattr(current, 'Name', '') or '').strip()
					text = str(getattr(current, 'Text', '') or '').strip()
					try:
						_kids = getattr(current, 'Children', None)
						_kid_count = int(getattr(_kids, 'Count', 0) or 0) if _kids is not None else 0
					except Exception:
						_kid_count = 0
					is_leaf = (_kid_count == 0)
					label_prefix = 'Menü Öğe' if is_leaf else 'Menü Grubu'
					display = text or name or 'Menü'
					seen_ids.add(norm_id)
					buttons.append({
						'id': norm_id,
						'type': node_type or 'GuiMenu',
						'name': name,
						'text': text,
						'label': f'{label_prefix}: {display} [{norm_id}]',
					})

				# GuiShell üzerindeki context-toolbar komutları (örn: &MB_EXPORT)
				is_shell = ('shell' in lower_type) or norm_id.casefold().endswith('/shell')
				if is_shell and norm_id:
					ctx_spec = f'ctxbtn:{norm_id}|&MB_EXPORT'
					if ctx_spec not in seen_ctx_specs:
						seen_ctx_specs.add(ctx_spec)
						buttons.append({
							'id': ctx_spec,
							'type': f'{node_type}:ContextToolbar',
							'name': '&MB_EXPORT',
							'text': 'Toolbar Context: Disa Aktar',
							'label': f'Toolbar Context: Disa Aktar (&MB_EXPORT) [{norm_id}]',
						})
					ctx_spec_pc = f'ctxbtn:{norm_id}|&MB_EXPORT|&PC'
					if ctx_spec_pc not in seen_ctx_specs:
						seen_ctx_specs.add(ctx_spec_pc)
						buttons.append({
							'id': ctx_spec_pc,
							'type': f'{node_type}:ContextToolbar',
							'name': '&MB_EXPORT -> &PC',
							'text': 'Toolbar Context: Disa Aktar > PC',
							'label': f'Toolbar Context: Disa Aktar -> PC (&MB_EXPORT -> &PC) [{norm_id}]',
						})
			except Exception:
				pass
			stack.extend(_iter_children(current))

	try:
		children = getattr(session, 'Children', None)
		wnd_count = int(getattr(children, 'Count', 0) or 0) if children is not None else 0
		if wnd_count == 0:
			_collect_buttons(session)
		else:
			for idx in range(wnd_count):
				wnd = None
				try:
					wnd = children(idx)
				except Exception:
					try:
						wnd = children.Item(idx)
					except Exception:
						wnd = None
				if wnd is not None:
					_collect_buttons(wnd)
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Buton tarama hatası: {ex}'}, status=500)

	buttons.sort(key=lambda x: x['id'])
	return JsonResponse({'ok': True, 'buttons': buttons, 'count': len(buttons)})


@require_POST
def sap_process_scan_selectables(request, process_id):
	"""Açık SAP ekranındaki radio/checkbox elementlerini tarar (wnd[0]+popup)."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	steps = _extract_runtime_steps(body, proc)
	if not steps:
		return JsonResponse({'ok': False, 'error': 'Önce en az bir adım tanımlayın.'}, status=400)

	conn, conn_err = _resolve_connection_from_steps(steps)
	if conn_err:
		return JsonResponse({'ok': False, 'error': conn_err}, status=400)

	service = SAPScanService()
	ok, payload = service.apply_to_screen(
		sys_id=conn.get('sys_id', ''),
		client=conn.get('client', ''),
		user=conn.get('user', ''),
		pwd=conn.get('pwd', ''),
		lang=conn.get('lang', 'TR'),
		t_code='',
		root_id=conn.get('root_id', 'wnd[0]'),
		extra_wait=conn.get('extra_wait', 0),
		actions=[],
		execute_f8=False,
	)
	if not ok:
		return JsonResponse({'ok': False, 'error': payload}, status=500)

	session = getattr(service, 'session', None)
	if session is None:
		return JsonResponse({'ok': False, 'error': 'Aktif SAP session bulunamadı.'}, status=500)

	items = []
	seen_ids = set()

	def _collect(node):
		stack = [node] if node is not None else []
		while stack:
			current = stack.pop(0)
			try:
				node_id = str(getattr(current, 'Id', '') or '').strip()
				node_type = str(getattr(current, 'Type', '') or '').strip()
				norm_id = _normalize_session_element_id(node_id)
				lower_type = node_type.casefold()
				lower_id = norm_id.casefold()
				is_radio = ('radiobutton' in lower_type) or ('/rad' in lower_id)
				is_checkbox = ('checkbox' in lower_type) or ('/chk' in lower_id)
				if (is_radio or is_checkbox) and norm_id and norm_id not in seen_ids:
					seen_ids.add(norm_id)
					text = str(getattr(current, 'Text', '') or '').strip()
					name = str(getattr(current, 'Name', '') or '').strip()
					kind = 'radio' if is_radio else 'checkbox'
					window_hint = 'wnd[1]' if 'wnd[1]' in norm_id else 'wnd[0]'
					label_text = text or name or node_type or 'Secim Alani'
					items.append({
						'id': norm_id,
						'type': node_type,
						'kind': kind,
						'window': window_hint,
						'text': text,
						'name': name,
						'label': f'{label_text} ({kind}) [{norm_id}]',
					})
			except Exception:
				pass
			stack.extend(_iter_children(current))

	try:
		children = getattr(session, 'Children', None)
		wnd_count = int(getattr(children, 'Count', 0) or 0) if children is not None else 0
		if wnd_count == 0:
			_collect(session)
		else:
			for idx in range(wnd_count):
				wnd = None
				try:
					wnd = children(idx)
				except Exception:
					try:
						wnd = children.Item(idx)
					except Exception:
						wnd = None
				if wnd is not None:
					_collect(wnd)
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Secilebilir alan tarama hatası: {ex}'}, status=500)

	items.sort(key=lambda x: x['id'])
	return JsonResponse({'ok': True, 'controls': items, 'count': len(items)})


@require_POST
def sap_process_scan_inputs(request, process_id):
	"""Açık SAP ekranındaki input alanlarını tarar (wnd[0]+popup)."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	steps = _extract_runtime_steps(body, proc)
	if not steps:
		return JsonResponse({'ok': False, 'error': 'Önce en az bir adım tanımlayın.'}, status=400)

	conn, conn_err = _resolve_connection_from_steps(steps)
	if conn_err:
		return JsonResponse({'ok': False, 'error': conn_err}, status=400)

	service = SAPScanService()
	ok, payload = service.apply_to_screen(
		sys_id=conn.get('sys_id', ''),
		client=conn.get('client', ''),
		user=conn.get('user', ''),
		pwd=conn.get('pwd', ''),
		lang=conn.get('lang', 'TR'),
		t_code='',
		root_id=conn.get('root_id', 'wnd[0]'),
		extra_wait=conn.get('extra_wait', 0),
		actions=[],
		execute_f8=False,
	)
	if not ok:
		return JsonResponse({'ok': False, 'error': payload}, status=500)

	session = getattr(service, 'session', None)
	if session is None:
		return JsonResponse({'ok': False, 'error': 'Aktif SAP session bulunamadı.'}, status=500)

	items = []
	seen_ids = set()

	def _collect(node):
		stack = [node] if node is not None else []
		while stack:
			current = stack.pop(0)
			try:
				node_id = str(getattr(current, 'Id', '') or '').strip()
				node_type = str(getattr(current, 'Type', '') or '').strip()
				norm_id = _normalize_session_element_id(node_id)
				lower_type = node_type.casefold()
				lower_id = norm_id.casefold()
				is_input = (
					'/txt' in lower_id
					or '/ctxt' in lower_id
					or '/pwd' in lower_id
					or '/okcd' in lower_id
					or 'textfield' in lower_type
					or 'textbox' in lower_type
					or 'inputfield' in lower_type
					or 'okcodefield' in lower_type
					or 'passwordfield' in lower_type
					or 'combobox' in lower_type
				)
				if is_input and norm_id and norm_id not in seen_ids:
					seen_ids.add(norm_id)
					text = str(getattr(current, 'Text', '') or '').strip()
					name = str(getattr(current, 'Name', '') or '').strip()
					window_hint = 'wnd[1]' if 'wnd[1]' in norm_id else 'wnd[0]'
					label_text = name or text or node_type or 'Input Alani'
					items.append({
						'id': norm_id,
						'type': node_type,
						'window': window_hint,
						'text': text,
						'name': name,
						'label': f'{label_text} [{norm_id}]',
					})
			except Exception:
				pass
			stack.extend(_iter_children(current))

	try:
		children = getattr(session, 'Children', None)
		wnd_count = int(getattr(children, 'Count', 0) or 0) if children is not None else 0
		if wnd_count == 0:
			_collect(session)
		else:
			for idx in range(wnd_count):
				wnd = None
				try:
					wnd = children(idx)
				except Exception:
					try:
						wnd = children.Item(idx)
					except Exception:
						wnd = None
				if wnd is not None:
					_collect(wnd)
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Input alan tarama hatası: {ex}'}, status=500)

	items.sort(key=lambda x: x['id'])
	return JsonResponse({'ok': True, 'controls': items, 'count': len(items)})


@require_POST
def sap_process_scan_windows_dialogs(request, process_id):
	"""Açık Windows diyaloglarını ve buton/checkbox kontrollerini tarar."""
	get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body) if request.body else {}
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	title_filter = str((body or {}).get('title_filter', '') or '').strip()
	try:
		dialogs = scan_visible_dialogs(title_filter=title_filter)
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Windows popup tarama hatası: {ex}'}, status=500)

	return JsonResponse({'ok': True, 'dialogs': dialogs, 'count': len(dialogs)})


@require_POST
def sap_process_excel_browse(request, process_id):
	"""Sunucu makinesinde Excel dosyası seçtirir."""
	get_object_or_404(SapProcess, pk=process_id)
	if tk is None:
		return JsonResponse({'ok': False, 'error': 'Dosya seçici için tkinter kullanılamıyor.'}, status=500)

	try:
		body = json.loads(request.body) if request.body else {}
	except (json.JSONDecodeError, TypeError):
		body = {}

	initial_path = str((body or {}).get('initial_path', '') or '').strip()
	initial_dir = ''
	if initial_path:
		initial_path = os.path.expandvars(os.path.expanduser(initial_path))
		if os.path.isdir(initial_path):
			initial_dir = initial_path
		elif os.path.isfile(initial_path):
			initial_dir = os.path.dirname(initial_path)

	root = None
	try:
		from tkinter import filedialog
		root = tk.Tk()
		root.withdraw()
		try:
			root.attributes('-topmost', True)
		except Exception:
			pass
		picked = filedialog.askopenfilename(
			title='Excel dosyası seç',
			filetypes=[('Excel Dosyalari', '*.xlsx *.xlsm *.xls'), ('Tum Dosyalar', '*.*')],
			initialdir=initial_dir or None,
		)
	finally:
		if root is not None:
			try:
				root.destroy()
			except Exception:
				pass

	if not picked:
		return JsonResponse({'ok': True, 'cancelled': True, 'path': ''})
	return JsonResponse({'ok': True, 'cancelled': False, 'path': str(picked)})


@require_POST
def sap_process_excel_sheets(request, process_id):
	"""Verilen Excel dosyasındaki sayfa adlarını döndürür."""
	get_object_or_404(SapProcess, pk=process_id)
	if load_workbook is None:
		return JsonResponse({'ok': False, 'error': 'Excel okuma için openpyxl kurulu değil.'}, status=500)

	try:
		body = json.loads(request.body) if request.body else {}
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	excel_path = str((body or {}).get('excel_file_path', '') or '').strip()
	if not excel_path:
		return JsonResponse({'ok': False, 'error': 'Excel dosya yolu boş.'}, status=400)
	excel_path = os.path.expandvars(os.path.expanduser(excel_path))
	if not os.path.isfile(excel_path):
		return JsonResponse({'ok': False, 'error': f'Excel dosyası bulunamadı: {excel_path}'}, status=404)

	try:
		wb = load_workbook(excel_path, data_only=True, read_only=True)
		sheets = list(wb.sheetnames or [])
		wb.close()
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Excel okunamadı: {ex}'}, status=500)

	return JsonResponse({'ok': True, 'sheets': sheets, 'count': len(sheets)})


@require_POST
def sap_process_excel_columns(request, process_id):
	"""Verilen Excel dosyasında seçili sayfanın başlık satırına göre sütun listesini döndürür."""
	get_object_or_404(SapProcess, pk=process_id)
	if load_workbook is None:
		return JsonResponse({'ok': False, 'error': 'Excel okuma için openpyxl kurulu değil.'}, status=500)

	try:
		body = json.loads(request.body) if request.body else {}
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	excel_path = str((body or {}).get('excel_file_path', '') or '').strip()
	if not excel_path:
		return JsonResponse({'ok': False, 'error': 'Excel dosya yolu boş.'}, status=400)
	excel_path = os.path.expandvars(os.path.expanduser(excel_path))
	if not os.path.isfile(excel_path):
		return JsonResponse({'ok': False, 'error': f'Excel dosyası bulunamadı: {excel_path}'}, status=404)

	sheet_name = str((body or {}).get('excel_sheet_name', '') or '').strip()
	try:
		header_row = max(1, int((body or {}).get('excel_header_row') or 1))
	except (TypeError, ValueError):
		header_row = 1

	def _column_letter(col_index):
		s = ''
		n = int(col_index or 0)
		while n > 0:
			n, r = divmod(n - 1, 26)
			s = chr(65 + r) + s
		return s or 'A'

	try:
		wb = load_workbook(excel_path, data_only=True, read_only=True)
		if sheet_name:
			if sheet_name not in wb.sheetnames:
				wb.close()
				return JsonResponse({'ok': False, 'error': f'Sayfa bulunamadı: {sheet_name}'}, status=404)
			ws = wb[sheet_name]
		else:
			ws = wb[wb.sheetnames[0]]

		max_col = int(getattr(ws, 'max_column', 0) or 0)
		if max_col <= 0:
			try:
				header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), ())
				max_col = len(tuple(header_values or ()))
			except Exception:
				max_col = 0
		max_col = max(1, min(max_col, 300))
		columns = []
		for idx in range(1, max_col + 1):
			letter = _column_letter(idx)
			head_val = ws.cell(row=header_row, column=idx).value
			head = str(head_val or '').strip()
			label = f'{letter} - {head}' if head else letter
			columns.append({'value': letter, 'label': label, 'header': head})
		wb.close()
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Sütunlar okunamadı: {ex}'}, status=500)

	return JsonResponse({'ok': True, 'columns': columns, 'count': len(columns)})


@require_POST
def sap_process_scan_screens(request, process_id):
	"""Açık SAP oturumundaki ekran başlıklarını (wnd[*]) tarar."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	steps = _extract_runtime_steps(body, proc)
	if not steps:
		return JsonResponse({'ok': False, 'error': 'Önce en az bir adım tanımlayın.'}, status=400)

	conn, conn_err = _resolve_connection_from_steps(steps)
	if conn_err:
		return JsonResponse({'ok': False, 'error': conn_err}, status=400)

	service = SAPScanService()
	ok, payload = service.apply_to_screen(
		sys_id=conn.get('sys_id', ''),
		client=conn.get('client', ''),
		user=conn.get('user', ''),
		pwd=conn.get('pwd', ''),
		lang=conn.get('lang', 'TR'),
		t_code='',
		root_id=conn.get('root_id', 'wnd[0]'),
		extra_wait=conn.get('extra_wait', 0),
		actions=[],
		execute_f8=False,
	)
	if not ok:
		return JsonResponse({'ok': False, 'error': payload}, status=500)

	session = getattr(service, 'session', None)
	if session is None:
		return JsonResponse({'ok': False, 'error': 'Aktif SAP session bulunamadı.'}, status=500)

	windows = []
	seen = set()
	try:
		children = getattr(session, 'Children', None)
		count = int(getattr(children, 'Count', 0) or 0) if children is not None else 0
		for idx in range(count):
			wnd = None
			try:
				wnd = children(idx)
			except Exception:
				try:
					wnd = children.Item(idx)
				except Exception:
					wnd = None
			if wnd is None:
				continue
			wid = str(getattr(wnd, 'Id', '') or '').strip()
			title = str(getattr(wnd, 'Text', '') or '').strip()
			if not title:
				title = wid or f'wnd[{idx}]'
			key = f'{wid}|{title}'.casefold()
			if key in seen:
				continue
			seen.add(key)
			windows.append({'id': _normalize_session_element_id(wid), 'title': title, 'label': f'{title} [{_normalize_session_element_id(wid)}]'})
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Ekran tarama hatası: {ex}'}, status=500)

	if not windows:
		try:
			current_title = str(service._get_window_title(session) or '').strip()
		except Exception:
			current_title = ''
		if current_title:
			windows.append({'id': 'wnd[0]', 'title': current_title, 'label': f'{current_title} [wnd[0]]'})

	return JsonResponse({'ok': True, 'screens': windows, 'count': len(windows)})


@require_POST
def sap_process_scan_grids(request, process_id):
	"""Açık SAP oturumundaki grid/tablo elementlerini tarar ve döndürür."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	steps = _extract_runtime_steps(body, proc)
	if not steps:
		return JsonResponse({'ok': False, 'error': 'Önce en az bir adım tanımlayın.'}, status=400)

	conn, conn_err = _resolve_connection_from_steps(steps)
	if conn_err:
		return JsonResponse({'ok': False, 'error': conn_err}, status=400)

	service = SAPScanService()
	# apply_to_screen ile bağlan (scan_screens ile aynı yöntem) — session nesnesine erişmek için
	ok, payload = service.apply_to_screen(
		sys_id=conn.get('sys_id', ''),
		client=conn.get('client', ''),
		user=conn.get('user', ''),
		pwd=conn.get('pwd', ''),
		lang=conn.get('lang', 'TR'),
		t_code='',
		root_id=conn.get('root_id', 'wnd[0]'),
		extra_wait=conn.get('extra_wait', 0),
		actions=[],
		execute_f8=False,
	)
	if not ok:
		return JsonResponse({'ok': False, 'error': payload}, status=500)

	session = getattr(service, 'session', None)
	if session is None:
		return JsonResponse({'ok': False, 'error': 'Aktif SAP session bulunamadı.'}, status=500)

	grids = []
	seen_ids = set()

	def _collect_grids(node):
		"""node altındaki tüm GuiTableControl ve GuiShell grid elementlerini recursive toplar."""
		stack = [node] if node is not None else []
		while stack:
			current = stack.pop(0)
			try:
				node_id = str(getattr(current, 'Id', '') or '').strip()
				node_type = str(getattr(current, 'Type', '') or '').strip()
				norm_id = _normalize_session_element_id(node_id)
				lower_type = node_type.casefold()
				lower_id = norm_id.casefold()
				last_segment = lower_id.split('/')[-1] if lower_id else ''
				# Yalnızca gerçek grid kontrolü dönsün; hücre (txt..., lbl...) ID'lerini dışarıda bırak.
				is_table_control = ('tablecontrol' in lower_type) or last_segment.startswith('tbl')
				is_shell_grid = ('shell' in lower_type) and (
					last_segment.startswith('shell') or 'grid' in lower_id or 'shellcont' in lower_id
				)
				is_grid = is_table_control or is_shell_grid
				if is_grid and norm_id and norm_id not in seen_ids:
					name = str(getattr(current, 'Name', '') or '').strip()
					text = str(getattr(current, 'Text', '') or '').strip()
					window_hint = 'wnd[1]' if 'wnd[1]' in norm_id else 'wnd[0]'
					label_text = text or name or node_type or 'Grid'
					seen_ids.add(norm_id)
					grids.append({
						'id': norm_id,
						'type': node_type,
						'name': name,
						'text': text,
						'window': window_hint,
						'label': f'{label_text} [{norm_id}]',
					})
			except Exception:
				pass
			stack.extend(_iter_children(current))

	# Tüm SAP pencerelerini (wnd[0], wnd[1], ...) tara
	try:
		children = getattr(session, 'Children', None)
		wnd_count = int(getattr(children, 'Count', 0) or 0) if children is not None else 0
		if wnd_count == 0:
			_collect_grids(session)
		else:
			for idx in range(wnd_count):
				wnd = None
				try:
					wnd = children(idx)
				except Exception:
					try:
						wnd = children.Item(idx)
					except Exception:
						wnd = None
				if wnd is not None:
					_collect_grids(wnd)
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Grid tarama hatası: {ex}'}, status=500)

	grids.sort(key=lambda x: x['id'])
	return JsonResponse({'ok': True, 'grids': grids, 'count': len(grids)})


import re as _re

def _normalize_session_element_id(element_id):
	"""'/app/con[0]/ses[0]/wnd[0]/...' → 'wnd[0]/...'  (veya zaten kısa ise olduğu gibi)."""
	eid = str(element_id or '').strip()
	# /app/con[N]/ses[N]/ önekini at
	eid = _re.sub(r'^/?app/con\[\d+\]/ses\[\d+\]/', '', eid)
	eid = eid.lstrip('/')
	return eid


def _parse_toolbar_context_command(raw_value):
	"""
	SAP buton alanından context-toolbar komutu ayrıştırır.
	Desteklenen formatlar:
	- ctxbtn:wnd[0]/.../shell|&MB_EXPORT
	- ctxbtn:wnd[0]/.../shell|&MB_EXPORT|&PC
	- ctxitem:wnd[0]/.../shell|&PC
	- session.findById("wnd[0]/.../shell").pressToolbarContextButton "&MB_EXPORT"
	- session.findById("wnd[0]/.../shell").pressToolbarContextButton "&MB_EXPORT".selectContextMenuItem "&PC"
	- session.findById("wnd[0]/.../shell").selectContextMenuItem "&PC"
	"""
	raw = str(raw_value or '').strip()
	if not raw:
		return None

	low = raw.casefold()
	if low.startswith('ctxbtn:') and '|' in raw:
		payload = raw.split(':', 1)[1]
		parts = payload.split('|')
		if len(parts) < 2:
			return None
		shell_raw = parts[0]
		cmd_raw = parts[1]
		menu_raw = parts[2] if len(parts) > 2 else ''
		shell_id = _normalize_session_element_id(shell_raw)
		command = str(cmd_raw or '').strip()
		menu_item = str(menu_raw or '').strip()
		if shell_id and command:
			return {
				'shell_id': shell_id,
				'command': command,
				'menu_item': menu_item,
				'raw': f'ctxbtn:{shell_id}|{command}' + (f'|{menu_item}' if menu_item else ''),
			}
		return None

	if low.startswith('ctxitem:') and '|' in raw:
		payload = raw.split(':', 1)[1]
		parts = payload.split('|')
		if len(parts) < 2:
			return None
		shell_raw = parts[0]
		menu_raw = parts[1]
		shell_id = _normalize_session_element_id(shell_raw)
		menu_item = str(menu_raw or '').strip()
		if shell_id and menu_item:
			return {
				'shell_id': shell_id,
				'command': '',
				'menu_item': menu_item,
				'raw': f'ctxitem:{shell_id}|{menu_item}',
			}
		return None

	m = _re.search(
		r'session\.findById\(\s*["\']([^"\']+)["\']\s*\)\s*\.\s*pressToolbarContextButton\s*["\']([^"\']+)["\'](?:\s*\.\s*selectContextMenuItem\s*["\']([^"\']+)["\'])?',
		raw,
		flags=_re.IGNORECASE,
	)
	if m:
		shell_id = _normalize_session_element_id(m.group(1))
		command = str(m.group(2) or '').strip()
		menu_item = str(m.group(3) or '').strip()
		if shell_id and command:
			return {
				'shell_id': shell_id,
				'command': command,
				'menu_item': menu_item,
				'raw': f'ctxbtn:{shell_id}|{command}' + (f'|{menu_item}' if menu_item else ''),
			}

	m2 = _re.search(
		r'session\.findById\(\s*["\']([^"\']+)["\']\s*\)\s*\.\s*selectContextMenuItem\s*["\']([^"\']+)["\']',
		raw,
		flags=_re.IGNORECASE,
	)
	if m2:
		shell_id = _normalize_session_element_id(m2.group(1))
		menu_item = str(m2.group(2) or '').strip()
		if shell_id and menu_item:
			return {
				'shell_id': shell_id,
				'command': '',
				'menu_item': menu_item,
				'raw': f'ctxitem:{shell_id}|{menu_item}',
			}

	return None


def _is_menu_target(node, raw_id=''):
	"""SAP menü öğesi (GuiMenu / mbar/menu[..]) mi? press() yerine select() ile çalışır."""
	nid = str(raw_id or '').casefold()
	if '/mbar/' in nid or '/menu[' in nid:
		return True
	try:
		ntype = str(getattr(node, 'Type', '') or '').casefold()
	except Exception:
		ntype = ''
	if not ntype:
		return False
	return ('menu' in ntype) and ('menubar' not in ntype)


def _invoke_button_or_menu(node, raw_id=''):
	"""Hedef GuiMenu ise select(), aksi halde press() çağırır."""
	if _is_menu_target(node, raw_id):
		node.select()
	else:
		node.press()


def _calc_dynamic_date(key):
	"""JS calcDynamicDate'in Python karşılığı — SAP DD.MM.YYYY formatında tarih döner."""
	from datetime import date, timedelta
	import calendar
	today = date.today()
	y, m = today.year, today.month

	if key == 'today':
		d = today
	elif key == 'yesterday':
		d = today - timedelta(days=1)
	elif key == 'year_start':
		d = date(y, 1, 1)
	elif key == 'month_start':
		d = date(y, m, 1)
	elif key == 'year_end':
		d = date(y, 12, 31)
	elif key == 'month_end':
		d = date(y, m, calendar.monthrange(y, m)[1])
	elif key == 'prev_month_start':
		pm = m - 1 if m > 1 else 12
		py = y if m > 1 else y - 1
		d = date(py, pm, 1)
	elif key == 'prev_month_end':
		pm = m - 1 if m > 1 else 12
		py = y if m > 1 else y - 1
		d = date(py, pm, calendar.monthrange(py, pm)[1])
	elif key == 'month_5':
		d = date(y, m, 5)
	elif key == 'prev_year_start':
		d = date(y - 1, 1, 1)
	elif key == 'days_15':
		d = today - timedelta(days=15)
	elif key == 'days_30':
		d = today - timedelta(days=30)
	elif key == 'days_45':
		d = today - timedelta(days=45)
	elif key == 'days_60':
		d = today - timedelta(days=60)
	elif key == 'days_365':
		d = today - timedelta(days=365)
	else:
		return ''

	return f"{d.day:02d}.{d.month:02d}.{d.year}"


def _parse_loop_values(raw):
	"""Virgül / noktalı virgül / satır sonu ile verilen döngü değerlerini normalize eder."""
	text = str(raw or '').strip()
	if not text:
		return []
	parts = _re.split(r'[\n,;]+', text)
	return [p.strip() for p in parts if str(p).strip()]


def _as_bool(value, default=False):
	if value is None:
		return bool(default)
	if isinstance(value, bool):
		return value
	text = str(value).strip().casefold()
	if not text:
		return bool(default)
	return text in ('1', 'true', 'evet', 'yes', 'on')


def _parse_decimal_text(text):
	raw = str(text or '').strip()
	if not raw:
		raise ValueError('Bos deger')

	clean = raw.replace('\u00a0', '').replace(' ', '')
	if not _re.search(r'\d', clean):
		raise ValueError(f'Sayisal deger bulunamadi: {raw}')

	filtered = ''.join(ch for ch in clean if (ch.isdigit() or ch in ',.-+'))
	if not filtered:
		raise ValueError(f'Sayisal deger ayristirilamadi: {raw}')

	# SAP trailing-minus desteği: "1.234,56-" → negatif
	trailing_negative = False
	if filtered.endswith('-'):
		trailing_negative = True
		filtered = filtered[:-1]

	sign = ''
	if filtered.startswith(('+', '-')):
		sign = filtered[0]
		filtered = filtered[1:]
	elif trailing_negative:
		sign = '-'

	if not filtered or not _re.search(r'\d', filtered):
		raise ValueError(f'Sayisal deger ayristirilamadi: {raw}')

	last_dot = filtered.rfind('.')
	last_comma = filtered.rfind(',')
	if last_dot >= 0 and last_comma >= 0:
		dec_sep = '.' if last_dot > last_comma else ','
	elif last_comma >= 0:
		dec_sep = ','
	elif last_dot >= 0:
		dec_sep = '.'
	else:
		dec_sep = None

	if dec_sep == ',':
		canon = filtered.replace('.', '').replace(',', '.')
	elif dec_sep == '.':
		canon = filtered.replace(',', '')
	else:
		canon = _re.sub(r'\D+', '', filtered)

	canon = f'{sign}{canon}'
	if not _re.fullmatch(r'[+-]?\d+(?:\.\d+)?', canon):
		raise ValueError(f'Sayisal deger parse edilemedi: {raw}')
	return Decimal(canon)


def _resolve_step_no_to_index(raw_step_no, steps_len):
	try:
		step_no = int(raw_step_no)
	except (TypeError, ValueError):
		return None
	if step_no < 1 or step_no > steps_len:
		return None
	return step_no - 1


def _resolve_step_target_index(steps, cfg, *, step_no_key, step_id_key=None):
	"""Hedef adımı önce step_id ile, yoksa step_no ile çözer."""
	if not isinstance(steps, list) or not isinstance(cfg, dict):
		return None

	if step_id_key:
		raw_id = cfg.get(step_id_key)
		step_id = str(raw_id or '').strip()
		if step_id:
			for idx, st in enumerate(steps):
				if not isinstance(st, dict):
					continue
				st_id = str(st.get('id') or '').strip()
				if st_id and st_id == step_id:
					return idx
			# Builder kaydında adımlar yeniden üretilebildiği için ID stale kalabilir.
			# Bu durumda step_no varsa ona düşmek çalışma akışını korur.
			return _resolve_step_no_to_index(cfg.get(step_no_key), len(steps))

	return _resolve_step_no_to_index(cfg.get(step_no_key), len(steps))


def _resolve_rule_target_index(steps, rule):
	"""if_else popup/status kurallarındaki hedefi çözer (step_id öncelikli)."""
	if not isinstance(steps, list) or not isinstance(rule, dict):
		return None
	raw_id = rule.get('step_id')
	step_id = str(raw_id or '').strip()
	if step_id:
		for idx, st in enumerate(steps):
			if not isinstance(st, dict):
				continue
			st_id = str(st.get('id') or '').strip()
			if st_id and st_id == step_id:
				return idx
		# Restore/clone sonrası ID stale kalabiliyor; mevcut step_no varsa ona düş.
		return _resolve_step_no_to_index(rule.get('step_no'), len(steps))
	return _resolve_step_no_to_index(rule.get('step_no'), len(steps))


def _read_sap_element_text(session, element_id):
	raw_id = str(element_id or '').strip()
	norm_id = _normalize_session_element_id(raw_id)
	if not norm_id:
		return False, '', 'Element ID bos.'

	elem = None
	for candidate in (norm_id, raw_id):
		if not candidate:
			continue
		try:
			elem = session.findById(candidate)
			if elem is not None:
				break
		except Exception:
			continue

	if elem is None:
		return False, '', f'Element bulunamadi: {norm_id}'

	for attr in ('Text', 'Value', 'Key'):
		try:
			val = getattr(elem, attr, None)
			if val is not None:
				text = str(val).strip()
				if text != '':
					return True, text, ''
		except Exception:
			pass

	try:
		selected = getattr(elem, 'Selected', None)
		if selected is not None:
			return True, ('1' if bool(selected) else '0'), ''
	except Exception:
		pass

	return False, '', f'Element degeri okunamadi: {norm_id}'


def _read_sap_statusbar(session):
	"""SAP alt durum çubuğundaki mesajı ve tipini okur (S/E/W/I/A)."""
	if session is None:
		return False, '', '', 'Session yok.'
	obj = None
	try:
		obj = session.findById('wnd[0]/sbar')
	except Exception:
		obj = None
	if obj is None:
		return False, '', '', 'Status bar bulunamadı (wnd[0]/sbar).'

	msg = ''
	for attr in ('Text', 'text', 'MessageText', 'Value'):
		try:
			v = getattr(obj, attr, None)
			if v is not None:
				t = str(v).strip()
				if t:
					msg = t
					break
		except Exception:
			pass

	msg_type = ''
	for attr in ('MessageType', 'messageType', 'Type'):
		try:
			v = getattr(obj, attr, None)
			if v is not None:
				t = str(v).strip().upper()
				if t:
					msg_type = t[0]
					break
		except Exception:
			pass

	return True, msg, msg_type, ''


def _ensure_runtime_loop_state(runtime_state, state, cfg):
	"""runtime_state içine loop_values ve aktif loop_value bilgisini ilk kez yükler."""
	if not isinstance(runtime_state, dict):
		return
	if runtime_state.get('loop_values'):
		return

	form = state.get('form', {}) if isinstance(state, dict) and isinstance(state.get('form'), dict) else {}
	raw = (
		cfg.get('loop_values')
		or cfg.get('loop_values_override')
		or form.get('loop_values')
		or ''
	)
	values = _parse_loop_values(raw)
	if not values:
		return

	runtime_state['loop_values'] = values
	runtime_state['loop_index'] = 0
	runtime_state['loop_value'] = values[0]


def _build_excel_cursor_key(cfg):
	if not isinstance(cfg, dict):
		return ''
	excel_path_raw = str(cfg.get('excel_file_path', '') or '').strip()
	# Windows'ta / ve \\ farkı aynı dosyaya işaret edebilir; cursor key'de tekilleştir.
	if excel_path_raw:
		try:
			excel_path = os.path.normpath(excel_path_raw).strip().lower()
		except Exception:
			excel_path = excel_path_raw.lower()
	else:
		excel_path = ''
	sheet_name = str(cfg.get('excel_sheet_name', '') or '').strip().lower()
	try:
		header_row = max(1, int(cfg.get('excel_header_row') or 1))
	except (TypeError, ValueError):
		header_row = 1
	if not excel_path:
		return ''
	return f'{excel_path}|{sheet_name}|{header_row}'


def _get_excel_cursor_index(runtime_state, cfg):
	if not isinstance(runtime_state, dict):
		return 0
	key = _build_excel_cursor_key(cfg)
	if not key:
		return 0
	cursors = runtime_state.get('excel_cursors')
	if not isinstance(cursors, dict):
		return 0
	try:
		return max(0, int(cursors.get(key, 0) or 0))
	except Exception:
		return 0


def _set_excel_cursor_index(runtime_state, cfg, index_value):
	if not isinstance(runtime_state, dict):
		return
	key = _build_excel_cursor_key(cfg)
	if not key:
		return
	cursors = runtime_state.get('excel_cursors')
	if not isinstance(cursors, dict):
		cursors = {}
		runtime_state['excel_cursors'] = cursors
	try:
		cursors[key] = max(0, int(index_value or 0))
	except Exception:
		cursors[key] = 0


def _reset_excel_cursors(runtime_state):
	if not isinstance(runtime_state, dict):
		return
	runtime_state['excel_cursors'] = {}


def _get_excel_row_index(cfg, runtime_state):
	row_source = str(cfg.get('excel_row_index_source', '') or '').strip().casefold()
	if not row_source:
		# Excel Satır Sonraki adımı kullanıldıysa (excel_cursors'ta bu dosya için kayıt varsa)
		# otomatik olarak excel_loop moduna geç
		_cursor_key = _build_excel_cursor_key(cfg)
		_cursors = runtime_state.get('excel_cursors') if isinstance(runtime_state, dict) else None
		if _cursor_key and isinstance(_cursors, dict) and _cursor_key in _cursors:
			row_source = 'excel_loop'
		else:
			row_source = 'outer_loop' if _as_bool(cfg.get('excel_use_loop_index', True), True) else 'fixed'
	if row_source == 'excel_loop':
		return _get_excel_cursor_index(runtime_state, cfg)
	if row_source == 'fixed':
		return 0
	return int(runtime_state.get('loop_index', 0) or 0) if isinstance(runtime_state, dict) else 0


def _get_excel_total_data_rows(excel_path, sheet_name, header_row, runtime_state=None, cfg=None):
	cache = runtime_state.get('excel_row_count_cache') if isinstance(runtime_state, dict) else None
	cache_key = None
	if isinstance(runtime_state, dict):
		cache_key = _build_excel_cursor_key(cfg or {
			'excel_file_path': excel_path,
			'excel_sheet_name': sheet_name,
			'excel_header_row': header_row,
		})
		if not isinstance(cache, dict):
			cache = {}
			runtime_state['excel_row_count_cache'] = cache
		if cache_key and cache_key in cache:
			return True, int(cache.get(cache_key, 0) or 0), ''

	if load_workbook is None:
		return False, 0, 'Excel okuma için openpyxl kurulu değil.'
	try:
		wb = load_workbook(excel_path, data_only=True, read_only=True)
	except Exception as ex:
		return False, 0, f'Excel açılamadı: {ex}'

	try:
		if sheet_name:
			if sheet_name not in wb.sheetnames:
				return False, 0, f'Sayfa bulunamadı: {sheet_name}'
			ws = wb[sheet_name]
		else:
			ws = wb[wb.sheetnames[0]]
		max_row = int(getattr(ws, 'max_row', 0) or 0)
		count = max(0, max_row - int(header_row))
		if count <= 0:
			# read_only çalışma modunda max_row güvenilmez olabiliyor; veri satırlarını fiilen say.
			count = 0
			for row in ws.iter_rows(min_row=int(header_row) + 1, values_only=True):
				if any((cell is not None and str(cell).strip() != '') for cell in (row or ())):
					count += 1
		if cache_key and isinstance(cache, dict):
			cache[cache_key] = count
		return True, count, ''
	finally:
		try:
			wb.close()
		except Exception:
			pass


def _excel_column_to_index(column_ref):
	text = str(column_ref or '').strip().upper()
	if not text:
		return None
	if text.isdigit():
		idx = int(text)
		return idx if idx > 0 else None
	if not _re.fullmatch(r'[A-Z]+', text):
		return None
	idx = 0
	for ch in text:
		idx = (idx * 26) + (ord(ch) - 64)
	return idx if idx > 0 else None


def _normalize_excel_value(value):
	if value is None:
		return ''

	def _format_decimal_tr(num, *, force_group_for_decimal=True):
		try:
			dec = Decimal(str(num))
		except (InvalidOperation, ValueError, TypeError):
			return str(num or '').strip()
		sign = '-' if dec < 0 else ''
		dec_abs = abs(dec)
		txt = format(dec_abs, 'f')
		if '.' in txt:
			txt = txt.rstrip('0').rstrip('.')
		if not txt:
			return '0'
		if '.' in txt:
			int_part, frac_part = txt.split('.', 1)
		else:
			int_part, frac_part = txt, ''
		if frac_part:
			if force_group_for_decimal:
				int_grouped = f"{int(int_part):,}".replace(',', '.')
			else:
				int_grouped = int_part
			return f"{sign}{int_grouped},{frac_part}"
		return f"{sign}{int_part}"

	# Tarih/datetime değeri TR formatına çevir.
	if isinstance(value, datetime):
		return value.strftime('%d.%m.%Y')
	if isinstance(value, date):
		return value.strftime('%d.%m.%Y')

	# Sayısal değerlerde ondalık ayracı korunur; Türk biçimine çevrilir.
	if isinstance(value, (int, float)):
		if isinstance(value, float) and value.is_integer():
			return str(int(value))
		return _format_decimal_tr(value, force_group_for_decimal=True)

	text = str(value or '').strip()
	if not text:
		return ''

	date_formats = [
		'%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y',
		'%Y-%m-%d', '%Y/%m/%d',
		'%d.%m.%Y %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S',
	]
	for fmt in date_formats:
		try:
			return datetime.strptime(text, fmt).strftime('%d.%m.%Y')
		except Exception:
			pass

	# Harf içermeyen değerlerde Türk sayı formatını koru (ondalık ayracı: ,).
	if _re.search(r'\d', text) and not _re.search(r'[A-Za-zÇĞİÖŞÜçğıöşü]', text):
		num_raw = text.replace('\u00a0', '').replace(' ', '')
		sign = ''
		if num_raw.startswith(('+', '-')):
			sign = num_raw[0]
			num_raw = num_raw[1:]
		if num_raw.isdigit():
			return f'{sign}{num_raw}'
		if ('.' in num_raw) or (',' in num_raw):
			last_dot = num_raw.rfind('.')
			last_comma = num_raw.rfind(',')
			if last_dot >= 0 and last_comma >= 0:
				dec_sep = '.' if last_dot > last_comma else ','
			elif last_comma >= 0:
				dec_sep = ','
			else:
				dec_sep = '.'
			if dec_sep == ',':
				canon = num_raw.replace('.', '').replace(',', '.')
			else:
				canon = num_raw.replace(',', '')
			if _re.fullmatch(r'\d+(?:\.\d+)?', canon):
				return _format_decimal_tr(f'{sign}{canon}', force_group_for_decimal=True)
		return _re.sub(r'\D+', '', f'{sign}{num_raw}')

	return text


def _resolve_fill_input_excel_value(cfg, runtime_state):
	if load_workbook is None:
		return False, '', 'Excel okuma için openpyxl kurulu değil.'

	excel_path = str(cfg.get('excel_file_path', '') or '').strip()
	if not excel_path:
		return False, '', 'Excel dosya yolu boş.'
	if not os.path.isfile(excel_path):
		return False, '', f'Excel dosyası bulunamadı: {excel_path}'

	column_ref = str(cfg.get('excel_column', '') or '').strip()
	if not column_ref:
		return False, '', 'Excel sütunu boş.'

	try:
		header_row = max(1, int(cfg.get('excel_header_row') or 1))
	except (TypeError, ValueError):
		header_row = 1
	try:
		row_offset = int(cfg.get('excel_row_offset') or 0)
	except (TypeError, ValueError):
		row_offset = 0

	try:
		wb = load_workbook(excel_path, data_only=True, read_only=True)
	except Exception as ex:
		return False, '', f'Excel açılamadı: {ex}'

	try:
		sheet_name = str(cfg.get('excel_sheet_name', '') or '').strip()
		if sheet_name:
			if sheet_name not in wb.sheetnames:
				return False, '', f'Sayfa bulunamadı: {sheet_name}'
			ws = wb[sheet_name]
		else:
			ws = wb[wb.sheetnames[0]]

		col_idx = None
		headers = [str(c.value or '').strip() for c in ws[header_row]]
		needle = column_ref.casefold()
		for idx, head in enumerate(headers, start=1):
			if head.casefold() == needle:
				col_idx = idx
				break
		if col_idx is None:
			col_idx = _excel_column_to_index(column_ref)
		if col_idx is None:
			return False, '', f'Sütun bulunamadı: {column_ref}'

		loop_idx = _get_excel_row_index(cfg, runtime_state)
		data_row_start = header_row + 1
		target_row = data_row_start + loop_idx + row_offset
		if target_row < 1:
			target_row = 1

		cell_value = ws.cell(row=target_row, column=col_idx).value
		resolved = _normalize_excel_value(cell_value)
		if resolved == '':
			return False, '', f'Excel hücresi boş: {ws.title}!R{target_row}C{col_idx}'
		msg = f'Excel değer alındı: {ws.title}!R{target_row}C{col_idx} -> {resolved}'
		return True, resolved, msg
	finally:
		try:
			wb.close()
		except Exception:
			pass


def _close_office_express_popups(service):
	"""Açık SAP popup'larında Ofis Ekspres mesajı varsa otomatik kapatır."""
	closed = 0
	try:
		session = getattr(service, 'session', None)
		if session is None:
			return 0
		deadline = time.time() + 2.0
		max_pass = 3
		pass_no = 0
		while pass_no < max_pass and time.time() < deadline:
			pass_no += 1
			children = getattr(session, 'Children', None)
			count = int(getattr(children, 'Count', 0) or 0) if children is not None else 0
			if count <= 1:
				break
			popup = session.findById('wnd[1]', False)
			if not popup:
				break
			title = str(getattr(popup, 'Text', '') or '').strip().casefold()
			if 'ofis ekspres' in title or 'office express' in title:
				popup.sendVKey(0)
				time.sleep(0.25)
				closed += 1
				continue
			break
	except Exception:
		return closed
	return closed


class _GhostOverlayWindow:
	"""Süreç çalışırken masaüstünde üstte kalan basit durum penceresi."""
	def __init__(self, enabled, process_name, process_id=None):
		self.enabled = bool(enabled)
		self.process_name = str(process_name or '').strip() or 'SAP Süreci'
		self.process_id = int(process_id) if process_id is not None else None
		self.pc_name = os.environ.get('COMPUTERNAME') or socket.gethostname() or 'Bilinmeyen'
		self.root = None
		self.header_title = None
		self.header_subtitle = None
		self.status_badge = None
		self.step_label = None
		self.log_text = None
		self.log_scroll = None
		self.pause_btn = None
		self.stop_btn = None
		self.minimize_btn = None
		self.settings_btn = None
		self._settings_win = None
		self._alpha_val = 0.97
		self.logs = []
		self.current_step = ''
		self.paused = False
		self.stop_requested = False
		self.logs_collapsed = False
		self._expanded_geometry = ''
		if not self.enabled or tk is None:
			self.enabled = False
			return
		try:
			self.root = tk.Tk()
			self.root.title('Saggio Hayalet Ekran')
			self.root.attributes('-topmost', True)
			self.root.attributes('-alpha', self._alpha_val)
			self.root.config(bg='#0f172a')
			# Pencere X butonuna tıklanırsa süreç durdurma talebi olarak ele al;
			# pencereyi yok etme — runtime cleanup ana akışta yapılacak.
			try:
				self.root.protocol('WM_DELETE_WINDOW', self._on_window_close)
			except Exception:
				pass
			screen_w = self.root.winfo_screenwidth()
			x = max(20, screen_w - 580)
			self.root.geometry(f'550x360+{x}+24')
			self.root.minsize(520, 320)
			self.root.maxsize(720, 720)
			self.root.resizable(True, True)

			container = tk.Frame(self.root, bg='#0f172a', bd=0, highlightthickness=1, highlightbackground='#334155')
			container.pack(expand=True, fill='both', padx=8, pady=8)

			header = tk.Frame(container, bg='#0f172a')
			header.pack(fill='x', padx=14, pady=(14, 8))

			self.header_title = tk.Label(
				header,
				text='SAGGIO HAYALET EKRAN',
				font=('Consolas', 12, 'bold'),
				fg='#e2e8f0',
				bg='#0f172a',
				anchor='w',
			)
			self.header_title.pack(fill='x')

			self.header_subtitle = tk.Label(
				header,
				text='',
				font=('Consolas', 9),
				fg='#94a3b8',
				bg='#0f172a',
				anchor='w',
			)
			self.header_subtitle.pack(fill='x', pady=(4, 0))

			status_wrap = tk.Frame(container, bg='#0f172a')
			status_wrap.pack(fill='x', padx=14, pady=(0, 10))

			self.status_badge = tk.Label(
				status_wrap,
				text='',
				font=('Consolas', 9, 'bold'),
				fg='white',
				bg='#1d4ed8',
				padx=10,
				pady=4,
			)
			self.status_badge.pack(side='left')

			self.step_label = tk.Label(
				status_wrap,
				text='',
				font=('Consolas', 9),
				fg='#cbd5e1',
				bg='#0f172a',
				anchor='w',
				justify='left',
			)
			self.step_label.pack(side='left', fill='x', expand=True, padx=(10, 0))

			btn_wrap = tk.Frame(container, bg='#0f172a')
			btn_wrap.pack(fill='x', padx=14, pady=(0, 10))
			self.minimize_btn = tk.Button(
				btn_wrap,
				text='Logu Gizle',
				font=('Consolas', 10, 'bold'),
				bg='#334155',
				fg='white',
				activebackground='#334155',
				activeforeground='white',
				relief='flat',
				bd=0,
				padx=10,
				pady=6,
				command=self.toggle_logs,
			)
			self.minimize_btn.pack(side='left', padx=(0, 6))

			self.pause_btn = tk.Button(
				btn_wrap,
				text='Duraklat',
				font=('Consolas', 10, 'bold'),
				bg='#1d4ed8',
				fg='white',
				activebackground='#1d4ed8',
				activeforeground='white',
				relief='flat',
				bd=0,
				padx=12,
				pady=6,
				command=self.toggle_pause,
			)
			self.pause_btn.pack(side='left', fill='x', expand=True, padx=(0, 6))
			self.stop_btn = tk.Button(
				btn_wrap,
				text='Durdur',
				font=('Consolas', 10, 'bold'),
				bg='#da3633',
				fg='white',
				activebackground='#da3633',
				activeforeground='white',
				relief='flat',
				bd=0,
				padx=12,
				pady=6,
				command=self.request_stop,
			)
			self.stop_btn.pack(side='left', fill='x', expand=True)

			self.settings_btn = tk.Button(
				btn_wrap,
				text='⚙',
				font=('Consolas', 11, 'bold'),
				bg='#475569',
				fg='white',
				activebackground='#475569',
				activeforeground='white',
				relief='flat',
				bd=0,
				padx=10,
				pady=6,
				command=self.open_settings,
			)
			self.settings_btn.pack(side='left', padx=(6, 0))

			log_card = tk.Frame(container, bg='#111827', bd=0, highlightthickness=1, highlightbackground='#334155')
			log_card.pack(expand=True, fill='both', padx=14, pady=(0, 14))

			log_head = tk.Label(
				log_card,
				text='Akış Logları',
				font=('Consolas', 10, 'bold'),
				fg='#e2e8f0',
				bg='#111827',
				anchor='w',
				padx=12,
				pady=10,
			)
			log_head.pack(fill='x')

			log_wrap = tk.Frame(log_card, bg='#111827')
			log_wrap.pack(expand=True, fill='both', padx=10, pady=(0, 10))
			self.log_text = tk.Text(
				log_wrap,
				font=('Consolas', 9),
				fg='#cbd5e1',
				bg='#111827',
				insertbackground='#cbd5e1',
				relief='flat',
				borderwidth=0,
				highlightthickness=0,
				padx=2,
				pady=2,
				wrap='word',
			)
			self.log_scroll = tk.Scrollbar(log_wrap, orient='vertical', command=self.log_text.yview)
			self.log_text.configure(yscrollcommand=self.log_scroll.set)
			self.log_text.pack(side='left', expand=True, fill='both')
			self.log_scroll.pack(side='right', fill='y')
			self.log_text.tag_configure('ok', foreground='#22c55e')
			self.log_text.tag_configure('warn', foreground='#f59e0b')
			self.log_text.tag_configure('err', foreground='#ef4444')
			self.log_text.tag_configure('info', foreground='#cbd5e1')
			self.log_text.configure(state='disabled')
			self._render()
		except Exception:
			self.enabled = False
			self.root = None
			self.header_title = None
			self.header_subtitle = None
			self.status_badge = None
			self.step_label = None

	def _render(self):
		if not self.enabled or self.root is None or self.header_title is None:
			return
		if not self._root_alive():
			# Tk root harici olarak yok edildi (kullanıcı pencereyi kapattı vb.).
			# Stop talebi olarak işaretle ve sessizce çık.
			self.enabled = False
			self.stop_requested = True
			if self.process_id is not None:
				try:
					_runtime_set_controls(self.process_id, paused=False, stop_requested=True)
				except Exception:
					pass
			return
		try:
			stamp = datetime.now().strftime('%H:%M:%S')
			status_text = 'Durduruldu' if self.stop_requested else ('Duraklatıldı' if self.paused else 'Çalışıyor')
			status_colors = {
				'Çalışıyor': '#1d4ed8',
				'Duraklatıldı': '#c2410c',
				'Durduruldu': '#b91c1c',
			}
			self.header_subtitle.config(text=f'Süreç: {self.process_name}   |   PC: {self.pc_name}   |   Güncelleme: {stamp}')
			if self.status_badge is not None:
				self.status_badge.config(text=status_text, bg=status_colors.get(status_text, '#1d4ed8'))
			if self.step_label is not None:
				self.step_label.config(text=f'Adım: {self.current_step or "-"}')
			if self.log_text is not None:
				self.log_text.configure(state='normal')
				self.log_text.delete('1.0', 'end')
				log_lines = self.logs[-200:] if self.logs else ['Hazır']
				for line in log_lines:
					log_tag = self._log_tag_for_line(line)
					self.log_text.insert('end', f'• {line}\n', log_tag)
				self.log_text.see('end')
				self.log_text.configure(state='disabled')
			if self.pause_btn is not None:
				self.pause_btn.config(text='Devam Et' if self.paused else 'Duraklat')
			if self.minimize_btn is not None:
				self.minimize_btn.config(text='Logu Göster' if self.logs_collapsed else 'Logu Gizle')
			self.root.update_idletasks()
			self.root.update()
		except Exception:
			pass

	def _log_tag_for_line(self, line):
		text = str(line or '').casefold()
		if any(k in text for k in ('hata', 'başarısız', 'basarisiz', 'error', 'failed', 'bulunamadı', 'bulunamadi')):
			return 'err'
		if any(k in text for k in ('uyarı', 'uyari', 'atlandı', 'atlandi', 'timeout', 'geçersiz', 'gecersiz')):
			return 'warn'
		if any(k in text for k in ('tamamlandı', 'tamamlandi', 'gönderildi', 'gonderildi', 'uygulandı', 'uygulandi', 'basıldı', 'basildi')):
			return 'ok'
		return 'info'

	def toggle_logs(self):
		if not self.enabled or self.root is None:
			return
		self.logs_collapsed = not self.logs_collapsed
		try:
			if self.logs_collapsed:
				self._expanded_geometry = self.root.winfo_geometry()
				self.root.geometry('550x170')
			else:
				if self._expanded_geometry:
					self.root.geometry(self._expanded_geometry)
				else:
					self.root.geometry('550x360')
		except Exception:
			pass
		self._render()

	def toggle_pause(self):
		if not self.enabled or self.stop_requested:
			return
		self.paused = not self.paused
		if self.process_id is not None:
			_runtime_set_controls(self.process_id, paused=self.paused, stop_requested=self.stop_requested)
		self._render()

	def request_stop(self):
		if not self.enabled:
			return
		self.stop_requested = True
		self.paused = False
		if self.process_id is not None:
			_runtime_set_controls(self.process_id, paused=False, stop_requested=True)
		self._render()

	def poll_controls(self):
		if not self.enabled:
			return bool(self.stop_requested)
		if not self._root_alive():
			# Pencere harici olarak kapatıldıysa stop sayılır.
			self.enabled = False
			self.stop_requested = True
			return True
		self._render()
		return bool(self.stop_requested)

	def wait_if_paused(self):
		if not self.enabled:
			return bool(self.stop_requested)
		while self.paused and not self.stop_requested:
			if not self._root_alive():
				self.enabled = False
				self.stop_requested = True
				return True
			try:
				self._render()
				time.sleep(0.15)
			except Exception:
				break
		return bool(self.stop_requested)

	def _root_alive(self):
		"""Tk root hâlâ geçerli mi (kullanıcı X ile kapatmadı mı)?"""
		if self.root is None:
			return False
		try:
			return bool(self.root.winfo_exists())
		except Exception:
			return False

	def _on_window_close(self):
		"""Kullanıcı pencerenin X butonuna bastığında: durdurma talebi tetikle.
		Pencereyi burada yok etme; ana akış cleanup'ı sağlasın."""
		try:
			self.stop_requested = True
			self.paused = False
			if self.process_id is not None:
				try:
					_runtime_set_controls(self.process_id, paused=False, stop_requested=True)
				except Exception:
					pass
				try:
					_runtime_push_log(self.process_id, 'Hayalet ekran kullanıcı tarafından kapatıldı; durdurma istendi.')
				except Exception:
					pass
		except Exception:
			pass

	def set_step(self, step_no, total_steps, step_name):
		if self.process_id is not None:
			_runtime_set_step(self.process_id, step_no, total_steps, step_name)
		if not self.enabled:
			return
		self.current_step = f'{step_no}/{total_steps} - {step_name}'
		self._render()

	def _stamp_log_message(self, msg):
		"""Log mesajını üretildiği anın saatiyle etiketle."""
		text = str(msg or '').strip()
		if not text:
			return ''
		stamp = datetime.now().strftime('%H:%M:%S')
		return f'[{stamp}] {text}'

	def push_log(self, text):
		msg = str(text or '').strip()
		if msg:
			msg = self._stamp_log_message(msg)
		if self.process_id is not None and msg:
			_runtime_push_log(self.process_id, msg)
		if not self.enabled:
			return
		if msg:
			self.logs.append(msg)
		self._render()

	def close(self):
		if self.root is None:
			return
		try:
			self.root.destroy()
		except Exception:
			pass
		self.root = None
		self.header_title = None
		self.header_subtitle = None
		self.status_badge = None
		self.step_label = None
		self.log_text = None
		self.log_scroll = None
		self.pause_btn = None
		self.stop_btn = None
		self.minimize_btn = None
		self.settings_btn = None
		try:
			if self._settings_win is not None:
				self._settings_win.destroy()
		except Exception:
			pass
		self._settings_win = None

	def open_settings(self):
		if not self.enabled or self.root is None:
			return
		try:
			if self._settings_win is not None and self._settings_win.winfo_exists():
				self._settings_win.focus_force()
				return
		except Exception:
			pass

		win = tk.Toplevel(self.root)
		self._settings_win = win
		win.title('Ayarlar')
		win.attributes('-topmost', True)
		win.config(bg='#0f172a')
		win.geometry('320x160')
		win.resizable(False, False)

		tk.Label(win, text='Şeffaflık Ayarları', font=('Consolas', 11, 'bold'),
			fg='#e2e8f0', bg='#0f172a').pack(anchor='w', padx=16, pady=(14, 6))

		slider_frame = tk.Frame(win, bg='#0f172a')
		slider_frame.pack(fill='x', padx=16)

		tk.Label(slider_frame, text='Opaklık:', font=('Consolas', 9),
			fg='#94a3b8', bg='#0f172a').pack(side='left')

		alpha_var = tk.DoubleVar(value=self._alpha_val)

		pct_lbl = tk.Label(slider_frame, text=f'{int(self._alpha_val * 100)}%',
			font=('Consolas', 9, 'bold'), fg='#e2e8f0', bg='#0f172a', width=5)
		pct_lbl.pack(side='right')

		def _on_alpha(val):
			v = round(float(val), 2)
			self._alpha_val = v
			try:
				self.root.attributes('-alpha', v)
			except Exception:
				pass
			pct_lbl.config(text=f'{int(v * 100)}%')

		scale = tk.Scale(win, from_=0.20, to=1.0, resolution=0.01, orient='horizontal',
			variable=alpha_var, command=_on_alpha,
			bg='#0f172a', fg='#e2e8f0', troughcolor='#334155',
			highlightthickness=0, bd=0, sliderlength=18, length=260)
		scale.pack(padx=16, pady=(4, 0))

		tk.Button(win, text='Kapat', font=('Consolas', 10, 'bold'),
			bg='#334155', fg='white', activebackground='#334155', activeforeground='white',
			relief='flat', bd=0, padx=12, pady=5,
			command=win.destroy).pack(pady=(10, 0))

	def __del__(self):
		self.close()

	def show_message(self, title, message):
		msg_title = str(title or '').strip() or 'Bilgi'
		msg_body = str(message or '').strip() or 'Devam etmek için Tamam butonuna basın.'
		if not self.enabled or self.root is None or tk is None:
			return _show_blocking_message_dialog(msg_title, msg_body)

		closed = {'done': False}
		win = None

		def _finish():
			closed['done'] = True
			try:
				if win is not None:
					win.grab_release()
			except Exception:
				pass
			try:
				if win is not None:
					win.destroy()
			except Exception:
				pass

		try:
			win = tk.Toplevel(self.root)
			win.title(msg_title)
			win.attributes('-topmost', True)
			win.transient(self.root)
			win.configure(bg='#0f172a')
			win.resizable(False, False)
			win.protocol('WM_DELETE_WINDOW', _finish)

			title_lbl = tk.Label(win, text=msg_title, font=('Consolas', 11, 'bold'), fg='#e2e8f0', bg='#0f172a', anchor='w', justify='left')
			title_lbl.pack(fill='x', padx=16, pady=(14, 8))

			body_lbl = tk.Label(win, text=msg_body, font=('Consolas', 10), fg='#cbd5e1', bg='#0f172a', justify='left', anchor='w', wraplength=420)
			body_lbl.pack(fill='both', expand=True, padx=16, pady=(0, 12))

			ok_btn = tk.Button(win, text='Tamam', font=('Consolas', 10, 'bold'), bg='#1d4ed8', fg='white', activebackground='#1d4ed8', activeforeground='white', relief='flat', command=_finish)
			ok_btn.pack(pady=(0, 14), ipadx=18, ipady=2)

			win.update_idletasks()
			width = max(360, win.winfo_reqwidth())
			height = max(170, win.winfo_reqheight())
			screen_w = win.winfo_screenwidth()
			screen_h = win.winfo_screenheight()
			x = max(40, int((screen_w - width) / 2))
			y = max(40, int((screen_h - height) / 2))
			win.geometry(f'{width}x{height}+{x}+{y}')
			win.grab_set()
			ok_btn.focus_set()

			while not closed['done']:
				try:
					self._render()
					win.update_idletasks()
					win.update()
				except Exception:
					break
				time.sleep(0.05)
			return True, ''
		except Exception as ex:
			try:
				if win is not None:
					win.destroy()
			except Exception:
				pass
			return _show_blocking_message_dialog(msg_title, msg_body, ex)


def _show_blocking_message_dialog(title, message, first_error=None):
	msg_title = str(title or '').strip() or 'Bilgi'
	msg_body = str(message or '').strip() or 'Devam etmek için Tamam butonuna basın.'
	if tk is None:
		if first_error is not None:
			return False, f'Mesaj penceresi açılamadı: {first_error}'
		return False, 'Mesaj penceresi için tkinter kullanılamıyor.'

	root = None
	try:
		if tk_messagebox is not None:
			root = tk.Tk()
			root.withdraw()
			try:
				root.attributes('-topmost', True)
			except Exception:
				pass
			tk_messagebox.showinfo(msg_title, msg_body, parent=root)
			return True, ''
		return False, 'Mesaj kutusu aracı kullanılamıyor.'
	except Exception as ex:
		base = first_error or ex
		return False, f'Mesaj penceresi açılamadı: {base}'
	finally:
		if root is not None:
			try:
				root.destroy()
			except Exception:
				pass


def _build_actions_from_template_state_with_runtime(state, runtime_state=None):
	"""
	Şablon state'indeki rows'dan SAP apply_to_screen actions listesi oluşturur.
	runtime_state: dongu tipi alanlar için {"loop_value": "..."} gibi değerler içerebilir.
	"""
	if not isinstance(state, dict):
		return []
	rows = state.get('rows', {})
	if not isinstance(rows, dict):
		return []

	rt = runtime_state or {}
	actions = []

	for raw_id, row in rows.items():
		if not isinstance(row, dict):
			continue
		if not row.get('checked'):
			continue

		element_id  = _normalize_session_element_id(raw_id)
		if not element_id or not element_id.startswith('wnd['):
			continue

		action_type = str(row.get('action_type', '') or '').strip()
		if not action_type:
			continue

		if action_type == 'sabit':
			value = str(row.get('value_text', '') or '')
		elif action_type == 'dinamik':
			value = _calc_dynamic_date(str(row.get('value_date', '') or ''))
		elif action_type == 'selectbox':
			value = str(row.get('value_select', '') or '')
		elif action_type == 'dongu':
			value = str(rt.get('loop_value', '') or '')
		else:
			# radio, chk, secilecek, vb.
			value = str(row.get('value_text', '') or '')

		# {sutun_N} ve {loop_value} placeholder'larını runtime değerleriyle değiştir
		if '{' in value and '}' in value:
			for rt_key, rt_val in rt.items():
				value = value.replace(f'{{{rt_key}}}', str(rt_val or ''))

		actions.append({'element_id': element_id, 'action_type': action_type, 'value': value})

	return actions


def _iter_children(node):
	try:
		children = getattr(node, 'Children', None)
		if children is None:
			return []
		try:
			count = int(children.Count)
		except Exception:
			count = 0
		result = []
		for idx in range(count):
			try:
				result.append(children(idx))
			except Exception:
				try:
					result.append(children.Item(idx))
				except Exception:
					continue
		return result
	except Exception:
		return []


def find_alv_grid(root, candidate_id=None, grid_type='main'):
	search_id = _normalize_session_element_id(candidate_id) if candidate_id else ''
	preferred = None
	stack = [root] if root is not None else []
	while stack:
		node = stack.pop(0)
		try:
			node_id = str(getattr(node, 'Id', '') or getattr(node, 'ID', '') or '')
			node_type = str(getattr(node, 'Type', '') or '')
			lower_id = node_id.lower()
			lower_type = node_type.lower()
			if search_id and node_id == search_id:
				return node
			is_grid_like = (
				'grid' in lower_id
				or 'grid' in lower_type
				or 'shell' in lower_type
				or 'tablecontrol' in lower_type
				or lower_id.split('/')[-1].startswith('tbl')
			)
			if is_grid_like:
				if grid_type == 'detail':
					if 'wnd[1]' in node_id or 'alv_ht' in lower_id:
						return node
				else:
					if 'wnd[0]' in node_id:
						return node
					if preferred is None:
						preferred = node
		except Exception:
			pass
		stack.extend(_iter_children(node))
	if search_id:
		return None
	return preferred


def _has_popup_window(session):
	try:
		children = getattr(session, 'Children', None)
		count = int(getattr(children, 'Count', 0) or 0) if children is not None else 0
		return count > 1
	except Exception:
		return False


def _get_grid_row_count(grid):
	for attr in ('rowCount', 'RowCount', 'visibleRowCount', 'VisibleRowCount'):
		try:
			value = int(getattr(grid, attr, 0) or 0)
			if value >= 0:
				return value
		except Exception:
			continue
	return 0


def _normalize_match_text(value):
	return ' '.join(str(value or '').strip().casefold().split())


def _resolve_grid_row_by_text(grid, text_contains):
	needle = _normalize_match_text(text_contains)
	if not needle:
		return None

	row_count = _get_grid_row_count(grid)
	if row_count <= 0:
		return None

	columns = []
	try:
		order = list(getattr(grid, 'ColumnOrder', []) or [])
		columns = [c for c in order if c is not None and str(c).strip()]
	except Exception:
		columns = []

	if not columns:
		try:
			cols_obj = getattr(grid, 'columns', None)
			cnt = int(getattr(cols_obj, 'Count', 0) or 0) if cols_obj is not None else 0
			for ci in range(cnt):
				try:
					col = cols_obj.elementAt(ci)
				except Exception:
					try:
						col = cols_obj.Item(ci)
					except Exception:
						col = None
				if col is None:
					continue
				name = str(getattr(col, 'Name', '') or getattr(col, 'name', '') or '').strip()
				columns.append(name if name else ci)
		except Exception:
			columns = []

	for ridx in range(row_count):
		cell_values = []
		seen_values = set()
		for col in columns:
			try:
				v = str(grid.getCellValue(ridx, col) or '').strip()
				if v and v not in seen_values:
					cell_values.append(v)
					seen_values.add(v)
				continue
			except Exception:
				pass
			try:
				cell = grid.GetCell(ridx, col)
				v = str(getattr(cell, 'Text', '') or getattr(cell, 'text', '') or '').strip()
				if v and v not in seen_values:
					cell_values.append(v)
					seen_values.add(v)
			except Exception:
				continue

		# GuiTableControl için index bazlı hücre okumayı her durumda dene.
		for ci in range(30):
			try:
				cell = grid.GetCell(ridx, ci)
				v = str(getattr(cell, 'Text', '') or getattr(cell, 'text', '') or '').strip()
				if v and v not in seen_values:
					cell_values.append(v)
					seen_values.add(v)
			except Exception:
				if ci > 8:
					break

		row_text = _normalize_match_text(' | '.join(v for v in cell_values if v))
		if row_text and needle in row_text:
			return ridx

	return None


def _read_grid_row_data(grid, row_index):
	"""
	Seçilen grid satırındaki tüm hücre değerlerini sutun_1, sutun_2, ... olarak döndürür.
	Boş hücreler dahil edilmez; her sutun_N için hem sütun adı hem değer saklanır.
	Dönüş: {'sutun_1': 'değer', 'sutun_2': 'değer', ...}  (en fazla 50 sütun)
	"""
	result = {}
	ridx = max(0, int(row_index or 0))
	columns = []

	try:
		order = list(getattr(grid, 'ColumnOrder', []) or [])
		columns = [c for c in order if c is not None and str(c).strip()]
	except Exception:
		columns = []

	if not columns:
		try:
			cols_obj = getattr(grid, 'columns', None)
			cnt = int(getattr(cols_obj, 'Count', 0) or 0) if cols_obj is not None else 0
			for ci in range(min(cnt, 50)):
				try:
					col = cols_obj.elementAt(ci)
				except Exception:
					try:
						col = cols_obj.Item(ci)
					except Exception:
						col = None
				if col is None:
					continue
				name = str(getattr(col, 'Name', '') or getattr(col, 'name', '') or '').strip()
				columns.append(name if name else ci)
		except Exception:
			columns = []

	slot = 1
	seen_values = set()
	for col in columns[:50]:
		v = ''
		try:
			v = str(grid.getCellValue(ridx, col) or '').strip()
		except Exception:
			pass
		if not v:
			try:
				cell = grid.GetCell(ridx, col)
				v = str(getattr(cell, 'Text', '') or getattr(cell, 'text', '') or '').strip()
			except Exception:
				pass
		if v and v not in seen_values:
			result[f'sutun_{slot}'] = v
			seen_values.add(v)
			slot += 1

	# GuiTableControl index bazlı fallback
	if not result:
		for ci in range(50):
			try:
				cell = grid.GetCell(ridx, ci)
				v = str(getattr(cell, 'Text', '') or getattr(cell, 'text', '') or '').strip()
				if v and v not in seen_values:
					result[f'sutun_{slot}'] = v
					seen_values.add(v)
					slot += 1
			except Exception:
				if ci > 8:
					break

	return result


def _select_row_on_grid(grid, row_index):
	"""Hem ALV grid hem GuiTableControl için satır seçmeyi dener."""
	try:
		idx = max(0, int(row_index or 0))
	except Exception:
		idx = 0

	row_count = _get_grid_row_count(grid)
	if row_count > 0:
		idx = min(idx, row_count - 1)

	errors = []

	# GuiTableControl için en güvenilir yöntem
	try:
		abs_row = grid.getAbsoluteRow(idx)
		abs_row.selected = True
		try:
			grid.currentCellRow = idx
		except Exception:
			pass
		# Eski scriptteki gibi ilk hücreye focus vermeyi dene
		try:
			first_cell = None
			for child in _iter_children(grid):
				try:
					cid = str(getattr(child, 'Id', '') or '')
					if cid.endswith('[0,0]') or cid.endswith(f'[0,{idx}]'):
						first_cell = child
						break
				except Exception:
					pass
			if first_cell is not None:
				first_cell.setFocus()
				try:
					first_cell.caretPosition = 0
				except Exception:
					pass
		except Exception:
			pass
		return True, idx, None
	except Exception as ex:
		errors.append(f'getAbsoluteRow: {ex}')

	# ALV GridView için yaygın yöntem
	try:
		grid.currentCellRow = idx
		try:
			grid.selectedRows = str(idx)
		except Exception:
			pass
		return True, idx, None
	except Exception as ex:
		errors.append(f'currentCellRow/selectedRows: {ex}')

	return False, idx, ' | '.join(errors)


def _find_grid(service, grid_id='', timeout_sec=5, grid_type='main'):
	normalized_id = _normalize_session_element_id(grid_id)
	deadline = time.time() + max(1, timeout_sec)
	while time.time() < deadline:
		service._wait_until_idle(service.session, timeout_sec=3, stable_checks=1)
		grid = None
		if normalized_id:
			grid = service._safe_find(service.session, normalized_id)
			# Bazı kayıtlarda grid yerine hücre id gelebilir; üst segmentlerden gerçek gridi bul.
			if grid is None and '/' in normalized_id:
				parts = normalized_id.split('/')
				for end in range(len(parts) - 1, 0, -1):
					candidate = '/'.join(parts[:end])
					last = candidate.split('/')[-1].lower()
					if not (last.startswith('tbl') or last.startswith('shell') or 'grid' in last):
						continue
					grid = service._safe_find(service.session, candidate)
					if grid is not None:
						break
		if grid is None:
			grid = find_alv_grid(service.session, normalized_id or None, grid_type=grid_type)
		# İstenen id verilmişse ve o bulunamadıysa yanlış grid'e düşmeyelim.
		if normalized_id and grid is not None:
			try:
				gid = _normalize_session_element_id(str(getattr(grid, 'Id', '') or ''))
				if gid != normalized_id and (f'{gid}/' not in f'{normalized_id}/'):
					grid = None
			except Exception:
				pass
		if grid is not None:
			return grid
		time.sleep(0.2)
	return None


def _collect_node_text(node, limit=40):
	parts = []
	stack = [node] if node is not None else []

	def _push_value(value):
		v = str(value or '').replace('\r', ' ').replace('\n', ' ').strip()
		if not v:
			return
		v = ' '.join(v.split())
		if v and v not in parts:
			parts.append(v)

	def _extract_html_text(obj):
		for doc_attr in ('Document', 'document', 'HtmlDocument', 'BrowserDocument'):
			try:
				doc = getattr(obj, doc_attr, None)
			except Exception:
				doc = None
			if doc is None:
				continue
			candidates = [doc]
			for sub_attr in ('body', 'documentElement'):
				try:
					sub = getattr(doc, sub_attr, None)
				except Exception:
					sub = None
				if sub is not None:
					candidates.append(sub)
			for cand in candidates:
				for txt_attr in ('innerText', 'Text', 'text', 'innerHTML'):
					try:
						raw = getattr(cand, txt_attr, None)
					except Exception:
						raw = None
					if not raw:
						continue
					value = str(raw)
					if txt_attr == 'innerHTML':
						value = re.sub(r'<[^>]+>', ' ', value)
					_push_value(value)

	while stack and len(parts) < limit:
		current = stack.pop(0)
		for attr in ('Text', 'text', 'Tooltip', 'DefaultTooltip', 'Name', 'Caption', 'Title', 'Value', 'MessageText'):
			try:
				value = str(getattr(current, attr, '') or '').strip()
			except Exception:
				value = ''
			_push_value(value)
		_extract_html_text(current)
		stack.extend(_iter_children(current))
	return ' | '.join(parts)


def _collect_popup_message_text(session, popup_root_id='wnd[1]', limit=80):
	"""SAP popup'larda soru metnini doğrudan txt/lbl alanlarından toplamayı dener."""
	if session is None:
		return ''
	root = _normalize_session_element_id(popup_root_id or 'wnd[1]')
	parts = []
	seen = set()

	def _push(v):
		try:
			txt = str(v or '').replace('\r', ' ').replace('\n', ' ').strip()
		except Exception:
			txt = ''
		if not txt:
			return
		txt = ' '.join(txt.split())
		if txt and txt not in seen:
			seen.add(txt)
			parts.append(txt)

	attrs = (
		'Text', 'text', 'Value', 'DisplayedText', 'PromptText', 'Caption', 'Title',
		'Tooltip', 'DefaultTooltip', 'MessageText', 'Name'
	)

	# Bilinen popup mesaj alanlarını doğrudan dene.
	candidates = []
	for i in range(1, 10):
		candidates.extend([
			f'{root}/usr/txtMESSTXT{i}',
			f'{root}/usr/txtSPOP-TEXTLINE{i}',
			f'{root}/usr/subSUBSCREEN:SAPLSPO1:0502/txtSPOP-TEXTLINE{i}',
		])
	for cid in candidates:
		try:
			obj = session.findById(cid)
		except Exception:
			obj = None
		if obj is None:
			continue
		for attr in attrs:
			try:
				_push(getattr(obj, attr, ''))
			except Exception:
				continue

	# Hala yoksa popup altındaki txt/lbl tiplerini gez.
	try:
		root_obj = session.findById(root)
	except Exception:
		root_obj = None

	stack = [root_obj] if root_obj is not None else []
	while stack and len(parts) < limit:
		node = stack.pop(0)
		try:
			node_id = str(getattr(node, 'Id', '') or getattr(node, 'ID', '') or '')
			node_type = str(getattr(node, 'Type', '') or '')
		except Exception:
			node_id = ''
			node_type = ''
		low_id = node_id.casefold()
		low_type = node_type.casefold()
		if '/txt' in low_id or '/lbl' in low_id or 'label' in low_type or 'text' in low_type:
			for attr in attrs:
				try:
					_push(getattr(node, attr, ''))
				except Exception:
					continue
		stack.extend(_iter_children(node))

	# Sık görülen teknik etiketleri eleyerek daha temiz mesaj döndür.
	noise_tokens = {'wnd[1]', 'usr', 'tbar[0]', 'shellcont', 'shell', 'button_1', 'button_2', 'button_3'}
	clean = []
	for p in parts:
		if p.casefold() in noise_tokens:
			continue
		clean.append(p)
	return ' | '.join(clean[:limit])


def _collect_popup_text_legacy(session, popup_root_id='wnd[1]', limit=200):
	"""Eski süreçte çalışan yöntemi birebir uygular: recursive Children.Item(i) + Text/text."""
	if session is None:
		return ''
	root = _normalize_session_element_id(popup_root_id or 'wnd[1]')
	try:
		popup = session.findById(root)
	except Exception:
		return ''

	parts = []
	seen = set()

	def _push(v):
		try:
			txt = str(v or '').replace('\r', ' ').replace('\n', ' ').strip()
		except Exception:
			txt = ''
		if not txt:
			return
		txt = ' '.join(txt.split())
		if txt and txt not in seen:
			seen.add(txt)
			parts.append(txt)

	def _extract_all_text(obj):
		if obj is None or len(parts) >= limit:
			return
		try:
			t = getattr(obj, 'text', getattr(obj, 'Text', ''))
		except Exception:
			t = ''
		_push(t)

		try:
			children = getattr(obj, 'Children', None)
			count = int(getattr(children, 'Count', 0) or 0) if children is not None else 0
		except Exception:
			count = 0

		for i in range(count):
			if len(parts) >= limit:
				break
			child = None
			try:
				child = children.Item(i)
			except Exception:
				try:
					child = children(i)
				except Exception:
					child = None
			if child is not None:
				_extract_all_text(child)

	_extract_all_text(popup)
	return ' | '.join(parts[:limit])


def _press_popup_button_by_text(popup, keyword_list):
	"""Popup içinde metnine göre uygun butonu bulup basar."""
	keywords = [str(k or '').strip().casefold() for k in (keyword_list or []) if str(k or '').strip()]
	if not keywords:
		return False, 'anahtar kelime listesi boş.'

	stack = [popup] if popup is not None else []
	while stack:
		node = stack.pop(0)
		try:
			node_type = str(getattr(node, 'Type', '') or '').casefold()
			node_id = str(getattr(node, 'Id', '') or '').strip()
			if 'button' in node_type:
				text = str(getattr(node, 'Text', '') or '').strip()
				name = str(getattr(node, 'Name', '') or '').strip()
				tip = str(getattr(node, 'Tooltip', '') or '').strip()
				haystack = f'{text} {name} {tip}'.casefold()
				if any(k in haystack for k in keywords):
					node.press()
					return True, node_id or text or name
		except Exception:
			pass
		stack.extend(_iter_children(node))
	return False, 'metne uyan popup butonu bulunamadı.'


def _popup_has_button_by_text(popup, keyword_list):
	"""Popup içinde verilen anahtar kelimelerden birini içeren buton var mı?"""
	keywords = [str(k or '').strip().casefold() for k in (keyword_list or []) if str(k or '').strip()]
	if not popup or not keywords:
		return False

	stack = [popup]
	while stack:
		node = stack.pop(0)
		try:
			node_type = str(getattr(node, 'Type', '') or '').casefold()
			node_id = str(getattr(node, 'Id', '') or '').casefold()
			if 'button' in node_type or '/btn' in node_id:
				text = str(getattr(node, 'Text', '') or '').strip()
				name = str(getattr(node, 'Name', '') or '').strip()
				tip = str(getattr(node, 'Tooltip', '') or '').strip()
				haystack = f'{text} {name} {tip}'.casefold()
				if any(k in haystack for k in keywords):
					return True
		except Exception:
			pass
		stack.extend(_iter_children(node))
	return False


def _safe_send_popup_mail(cfg, mail_enabled=True, runtime_state=None, notification_cfg=None, popup_title='', popup_text=''):
	if not mail_enabled:
		return 'Mail gönderimi süreç ayarında kapalı.'
	if not bool(cfg.get('send_mail_on_match')):
		return None
	rt = runtime_state if isinstance(runtime_state, dict) else {}
	notify = notification_cfg if isinstance(notification_cfg, dict) else {}

	mail_to = str(cfg.get('mail_to', '') or '').strip() or str(notify.get('mail_to', '') or '').strip()
	account_id = cfg.get('mail_account_id') or notify.get('mail_account_id')
	if not account_id:
		return 'Mail atlandi: mail hesabı seçili değil.'
	account = MailAccount.objects.filter(pk=account_id, is_active=True).first()
	if not account:
		return f'Mail atlandi: hesap bulunamadi (id={account_id}).'
	if not mail_to:
		mail_to = str(account.email or '').strip()
	if not mail_to:
		return 'Mail atlandi: alıcı bulunamadı.'

	subject = str(cfg.get('mail_subject', '') or '').strip() or f'SAP popup uyarısı: {popup_title or "Popup"}'
	body_lines = []
	custom_body = str(cfg.get('mail_body', '') or '').strip()
	if custom_body:
		body_lines.append(custom_body)
	body_lines.append(f'Popup Başlığı: {popup_title or "(boş)"}')
	body_lines.append(f'Popup Metni: {popup_text or "(boş)"}')

	memory_items = []
	for k in sorted(rt.keys(), key=lambda x: str(x)):
		ks = str(k)
		if ks.startswith('sutun_') or ks.startswith('loop_'):
			memory_items.append((ks, rt.get(k)))
	if memory_items:
		body_lines.append('')
		body_lines.append('Hafıza Değerleri:')
		for mk, mv in memory_items:
			body_lines.append(f'- {mk}: {mv}')
	body = '\n'.join(body_lines)

	try:
		msg = MIMEText(body, _charset='utf-8')
		msg['Subject'] = subject
		msg['From'] = account.email
		msg['To'] = mail_to
		if account.use_ssl:
			server = smtplib.SMTP_SSL(account.smtp_host, int(account.smtp_port or 465), timeout=20)
		else:
			server = smtplib.SMTP(account.smtp_host, int(account.smtp_port or 587), timeout=20)
		try:
			server.ehlo()
			if account.use_tls and not account.use_ssl:
				server.starttls()
				server.ehlo()
			server.login(account.smtp_username, account.get_smtp_password())
			server.sendmail(account.email, [mail_to], msg.as_string())
		finally:
			server.quit()
		return f'Mail gonderildi: {mail_to}'
	except Exception as ex:
		return f'Mail gonderim hatasi: {ex}'

def _send_sap_hotkey(service, combo_text='', key='', use_ctrl=False, use_alt=False, use_shift=False, use_win=False):
	"""SAP aktif pencereye klavye kombinasyonu gönderir.
	ESC/ENTER gibi saf VKey'ler için önce SAP sendVKey kullanılır; diğerleri WScript.Shell ile gönderilir."""
	if service is None or getattr(service, 'session', None) is None:
		return False, 'SAP session hazir degil.'

	# Sadece modifier'sız tek tuş ise VKey ile gönder (daha güvenilir)
	_no_modifier = not use_ctrl and not use_alt and not use_shift and not use_win
	_combo_clean = str(combo_text or '').strip().upper()
	_key_clean = str(key or '').strip().upper()
	_vkey_map = {'ENTER': 0, 'ESC': 12, 'ESCAPE': 12,
	             'F1': 112, 'F2': 113, 'F3': 114, 'F4': 115, 'F5': 5, 'F6': 6, 'F7': 7, 'F8': 8,
	             'F9': 9, 'F10': 10, 'F11': 11, 'F12': 99,
	             'F13': 14, 'F14': 15, 'F15': 16, 'F16': 17, 'F17': 18, 'F18': 19,
	             'F19': 20, 'F20': 21, 'F21': 22, 'F22': 23, 'F23': 24, 'F24': 25}

	effective_key = _key_clean
	if not _combo_clean and _no_modifier:
		vk = _vkey_map.get(effective_key)
		if vk is not None:
			try:
				wnd = service.session.findById('wnd[0]')
				try:
					wnd.setFocus()
				except Exception:
					pass
				service._wait_until_idle(service.session, timeout_sec=3, stable_checks=1)
				wnd.sendVKey(vk)
				service._wait_until_idle(service.session, timeout_sec=5, stable_checks=1)
				return True, f'VKey({vk}) [{effective_key}]'
			except Exception as ex:
				return False, str(ex)

	send_keys = build_sendkeys_from_config(
		combo_text=combo_text,
		key=key,
		use_ctrl=use_ctrl,
		use_alt=use_alt,
		use_shift=use_shift,
		use_win=use_win,
	)
	if not send_keys:
		return False, 'Gonderilecek tus/kombinasyon bos.'

	try:
		wnd = service.session.findById('wnd[0]')
		try:
			wnd.setFocus()
		except Exception:
			pass
		service._wait_until_idle(service.session, timeout_sec=3, stable_checks=1)
		import win32com.client
		shell = win32com.client.Dispatch('WScript.Shell')
		shell.SendKeys(send_keys)
		service._wait_until_idle(service.session, timeout_sec=5, stable_checks=1)
		return True, send_keys
	except Exception as ex:
		return False, str(ex)


def _extract_runtime_steps(body, proc):
	"""İstekte gelen steps varsa onu, yoksa DB steps'i kullan."""
	steps_data = body.get('steps')
	if isinstance(steps_data, list):
		clean = []
		for i, s in enumerate(steps_data):
			if not isinstance(s, dict):
				continue
			clean.append({
				'id': s.get('id'),
				'order': i,
				'step_type': str(s.get('step_type', '') or '').strip(),
				'label': str(s.get('label', '') or '').strip(),
				'config': s.get('config', {}) if isinstance(s.get('config'), dict) else {},
			})
		return clean

	return list(proc.steps.values('id', 'order', 'step_type', 'label', 'config').order_by('order'))


def _resolve_connection_from_steps(steps):
	"""İlk uygun sap_fill adımındaki şablondan connection/form bilgisini al."""
	for step in steps:
		step_type = str(step.get('step_type', '') or '').strip()
		if step_type != SapProcessStep.TYPE_SAP_FILL:
			continue
		cfg = step.get('config', {}) if isinstance(step.get('config'), dict) else {}
		tpl_name = str(cfg.get('template_name', '') or '').strip()
		if not tpl_name:
			continue
		tpl = SAPTemplateService.get_template(tpl_name)
		if not isinstance(tpl, dict):
			continue
		state = tpl.get('state', {}) if isinstance(tpl.get('state'), dict) else {}
		form = state.get('form', {}) if isinstance(state.get('form'), dict) else {}
		conn = {
			'sys_id': str(form.get('sys_id', '') or '').strip(),
			'client': str(form.get('client', '') or '').strip(),
			'user': str(form.get('user', '') or '').strip(),
			'pwd': str(form.get('pwd', '') or '').strip(),
			'lang': str(form.get('lang', 'TR') or 'TR').strip(),
			't_code': str(form.get('t_code', '') or '').strip(),
			'root_id': str(form.get('root_id', 'wnd[0]') or 'wnd[0]').strip(),
			'extra_wait': form.get('extra_wait', 0),
			'template_name': tpl_name,
		}
		if conn['sys_id'] and conn['client'] and conn['user'] and conn['pwd']:
			return conn, None
		return None, f'Şablon bağlantı bilgileri eksik: {tpl_name}'

	return None, 'Bağlantı bilgisi çözümlemek için en az bir geçerli Şablon adımı gerekli.'


def _find_loop_next_step_index(steps, current_index):
	"""loop_next adımını bulur; önce ileri, yoksa baştan mevcut adıma kadar arar."""
	try:
		start_idx = int(current_index)
	except Exception:
		start_idx = -1

	for k in range(start_idx + 1, len(steps)):
		st = str(steps[k].get('step_type', '') or '').strip()
		if st == SapProcessStep.TYPE_LOOP_NEXT:
			return k

	for k in range(0, max(0, start_idx + 1)):
		st = str(steps[k].get('step_type', '') or '').strip()
		if st == SapProcessStep.TYPE_LOOP_NEXT:
			return k

	return None


def _advance_loop_runtime(steps, runtime_state, current_index):
	"""loop değerini bir sonraki elemana taşır ve hedef adıma dönüş indexini hesaplar."""
	loop_values = runtime_state.get('loop_values') or []
	if not loop_values:
		return False, 'Döngü değeri tanımlı değil', None
	current_idx = int(runtime_state.get('loop_index', 0) or 0)
	next_loop_idx = current_idx + 1
	if next_loop_idx >= len(loop_values):
		return False, 'Döngü değerleri tamamlandı', None

	runtime_state['loop_index'] = next_loop_idx
	runtime_state['loop_value'] = loop_values[next_loop_idx]
	_reset_excel_cursors(runtime_state)

	target_idx = None
	for back_i in range(int(current_index) - 1, -1, -1):
		st = str(steps[back_i].get('step_type', '') or '').strip()
		if st == SapProcessStep.TYPE_SAP_FILL:
			target_idx = back_i
			break
	if target_idx is None:
		target_idx = 0

	msg = f'Döngüde sonraki kayıt: {next_loop_idx + 1}/{len(loop_values)} ({runtime_state.get("loop_value", "")}) | adıma dön: {target_idx + 1}'
	return True, msg, target_idx


def _advance_excel_loop_runtime(steps, runtime_state, cfg):
	excel_path = str(cfg.get('excel_file_path', '') or '').strip()
	sheet_name = str(cfg.get('excel_sheet_name', '') or '').strip()
	try:
		header_row = max(1, int(cfg.get('excel_header_row') or 1))
	except (TypeError, ValueError):
		header_row = 1

	if not excel_path and isinstance(steps, list):
		for cand in steps:
			if not isinstance(cand, dict):
				continue
			if str(cand.get('step_type', '') or '').strip() != SapProcessStep.TYPE_SAP_FILL_INPUT:
				continue
			ccfg = cand.get('config', {}) if isinstance(cand.get('config'), dict) else {}
			if str(ccfg.get('value_source', 'static') or 'static').strip() != 'excel_column':
				continue
			cand_path = str(ccfg.get('excel_file_path', '') or '').strip()
			if cand_path:
				excel_path = cand_path
				sheet_name = sheet_name or str(ccfg.get('excel_sheet_name', '') or '').strip()
				try:
					header_row = max(1, int(ccfg.get('excel_header_row') or header_row))
				except (TypeError, ValueError):
					pass
				break

	if not excel_path:
		return False, 'Excel dosya yolu boş.', None
	if not os.path.isfile(excel_path):
		return False, f'Excel dosyası bulunamadı: {excel_path}', None
	cfg_resolved = dict(cfg or {})
	cfg_resolved['excel_file_path'] = excel_path
	cfg_resolved['excel_sheet_name'] = sheet_name
	cfg_resolved['excel_header_row'] = header_row
	ok_count, total_rows, count_err = _get_excel_total_data_rows(excel_path, sheet_name, header_row, runtime_state=runtime_state, cfg=cfg_resolved)
	if not ok_count:
		return False, count_err, None
	if total_rows <= 0:
		return False, 'Excel veri satırı bulunamadı.', None

	current_idx = _get_excel_cursor_index(runtime_state, cfg_resolved)
	next_idx = current_idx + 1
	if next_idx >= total_rows:
		_set_excel_cursor_index(runtime_state, cfg_resolved, 0)
		runtime_state['excel_loop_index'] = 0
		return False, f'Excel satırları tamamlandı ({total_rows}/{total_rows})', None

	_set_excel_cursor_index(runtime_state, cfg_resolved, next_idx)
	runtime_state['excel_loop_index'] = next_idx
	target_idx = _resolve_step_target_index(
		steps,
		cfg,
		step_no_key='target_step_no',
		step_id_key='target_step_id',
	)
	if target_idx is None:
		return False, 'Excel satırı ilerletildi fakat hedef adım no geçersiz.', None
	msg = f'Excelde sonraki satıra geçildi: {next_idx + 1}/{total_rows} | adıma dön: {target_idx + 1}'
	return True, msg, target_idx


def _ftp_list_files(account, remote_path='.', file_pattern='*'):
	remote_path = str(remote_path or '.').strip() or '.'
	file_pattern = str(file_pattern or '*').strip() or '*'
	items = []

	if account.protocol == 'sftp':
		if paramiko is None:
			raise RuntimeError('SFTP için paramiko kurulu değil.')
		transport = paramiko.Transport((account.host, int(account.port or 22)))
		try:
			transport.connect(username=account.username, password=account.get_password())
			sftp = paramiko.SFTPClient.from_transport(transport)
			for attr in sftp.listdir_attr(remote_path):
				name = attr.filename
				if fnmatch.fnmatch(name, file_pattern):
					items.append(name)
			sftp.close()
		finally:
			transport.close()
		return items

	if account.protocol == 'ftps':
		ftp = ftplib.FTP_TLS()
		ftp.connect(account.host, int(account.port or 21), timeout=20)
		ftp.login(account.username, account.get_password())
		ftp.prot_p()
		ftp.cwd(remote_path)
		try:
			items = [n for n in ftp.nlst() if fnmatch.fnmatch(os.path.basename(n), file_pattern)]
		finally:
			ftp.quit()
		return items

	ftp = ftplib.FTP()
	ftp.connect(account.host, int(account.port or 21), timeout=20)
	ftp.login(account.username, account.get_password())
	ftp.cwd(remote_path)
	try:
		items = [n for n in ftp.nlst() if fnmatch.fnmatch(os.path.basename(n), file_pattern)]
	finally:
		ftp.quit()
	return items


def _ftp_download(account, remote_path, local_path, file_pattern='*', limit=0):
	remote_path = str(remote_path or '.').strip() or '.'
	local_path = str(local_path or '').strip()
	if not local_path:
		raise RuntimeError('local_path zorunlu.')
	os.makedirs(local_path, exist_ok=True)

	files = _ftp_list_files(account, remote_path=remote_path, file_pattern=file_pattern)
	if limit and limit > 0:
		files = files[:limit]

	downloaded = []
	if account.protocol == 'sftp':
		if paramiko is None:
			raise RuntimeError('SFTP için paramiko kurulu değil.')
		transport = paramiko.Transport((account.host, int(account.port or 22)))
		try:
			transport.connect(username=account.username, password=account.get_password())
			sftp = paramiko.SFTPClient.from_transport(transport)
			for name in files:
				remote_file = f"{remote_path.rstrip('/')}/{os.path.basename(name)}"
				local_file = os.path.join(local_path, os.path.basename(name))
				sftp.get(remote_file, local_file)
				downloaded.append(local_file)
			sftp.close()
		finally:
			transport.close()
		return downloaded

	if account.protocol == 'ftps':
		ftp = ftplib.FTP_TLS()
		ftp.connect(account.host, int(account.port or 21), timeout=20)
		ftp.login(account.username, account.get_password())
		ftp.prot_p()
		ftp.cwd(remote_path)
		try:
			for name in files:
				base = os.path.basename(name)
				local_file = os.path.join(local_path, base)
				with open(local_file, 'wb') as fp:
					ftp.retrbinary(f'RETR {base}', fp.write)
				downloaded.append(local_file)
		finally:
			ftp.quit()
		return downloaded

	ftp = ftplib.FTP()
	ftp.connect(account.host, int(account.port or 21), timeout=20)
	ftp.login(account.username, account.get_password())
	ftp.cwd(remote_path)
	try:
		for name in files:
			base = os.path.basename(name)
			local_file = os.path.join(local_path, base)
			with open(local_file, 'wb') as fp:
				ftp.retrbinary(f'RETR {base}', fp.write)
			downloaded.append(local_file)
	finally:
		ftp.quit()
	return downloaded


def _ftp_upload(account, local_file, remote_path):
	local_file = str(local_file or '').strip()
	remote_path = str(remote_path or '.').strip() or '.'
	if not local_file:
		raise RuntimeError('local_file zorunlu.')
	if not os.path.isfile(local_file):
		raise RuntimeError(f'Yerel dosya bulunamadı: {local_file}')
	base = os.path.basename(local_file)

	if account.protocol == 'sftp':
		if paramiko is None:
			raise RuntimeError('SFTP için paramiko kurulu değil.')
		transport = paramiko.Transport((account.host, int(account.port or 22)))
		try:
			transport.connect(username=account.username, password=account.get_password())
			sftp = paramiko.SFTPClient.from_transport(transport)
			remote_file = f"{remote_path.rstrip('/')}/{base}"
			sftp.put(local_file, remote_file)
			sftp.close()
		finally:
			transport.close()
		return remote_file

	if account.protocol == 'ftps':
		ftp = ftplib.FTP_TLS()
		ftp.connect(account.host, int(account.port or 21), timeout=20)
		ftp.login(account.username, account.get_password())
		ftp.prot_p()
		ftp.cwd(remote_path)
		try:
			with open(local_file, 'rb') as fp:
				ftp.storbinary(f'STOR {base}', fp)
		finally:
			ftp.quit()
		return f"{remote_path.rstrip('/')}/{base}"

	ftp = ftplib.FTP()
	ftp.connect(account.host, int(account.port or 21), timeout=20)
	ftp.login(account.username, account.get_password())
	ftp.cwd(remote_path)
	try:
		with open(local_file, 'rb') as fp:
			ftp.storbinary(f'STOR {base}', fp)
	finally:
		ftp.quit()
	return f"{remote_path.rstrip('/')}/{base}"


@require_POST
def sap_process_run_preview(request, process_id):
	"""Süreci (veya belirtilen adıma kadar) gerçek SAP oturumunda çalıştır."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	async_mode = _as_bool(body.get('async_mode', False), False)
	if async_mode:
		state = _runtime_get(process_id)
		if state and bool(state.get('running')):
			# Stop istendi, force_reset talebi geldi veya updated_at 30+ saniyedir
			# güncellenmediyse runtime kaydını otomatik sıfırla.
			_force_reset = bool(body.get('_force_reset')) or bool(state.get('stop_requested'))
			if not _force_reset:
				try:
					from datetime import datetime as _dt
					_upd = state.get('updated_at', '')
					_age = (_dt.now() - _dt.fromisoformat(_upd)).total_seconds() if _upd else 999
					_force_reset = _age > 30
				except Exception:
					_force_reset = True
			if _force_reset:
				_runtime_finish(process_id)
			else:
				return JsonResponse({
					'ok': False,
					'error': 'Bu süreç zaten çalışıyor. Eğer takıldıysa "Zorla Sıfırla" deneyin.',
					'can_force_reset': True,
				}, status=409)

		def _async_runner(snapshot_body, req_meta, req_user):
			try:
				child_req = HttpRequest()
				child_req.method = 'POST'
				child_req.META = dict(req_meta or {})
				if req_user is not None:
					child_req.user = req_user
				payload = dict(snapshot_body or {})
				payload['async_mode'] = False
				child_req._body = json.dumps(payload).encode('utf-8')
				sap_process_run_preview(child_req, process_id)
			except Exception as ex:
				_runtime_push_log(process_id, f'Async runner hatası: {ex}')
				_runtime_finish(process_id)

		_runtime_init(process_id, proc.name, 0)
		_runtime_push_log(process_id, 'Süreç başlatılıyor (arka plan)...')
		threading.Thread(
			target=_async_runner,
			args=(body, getattr(request, 'META', {}).copy(), getattr(request, 'user', None)),
			daemon=True,
		).start()
		return JsonResponse({'ok': True, 'started': True, 'message': 'Süreç arka planda başlatıldı.'})

	_runtime_init(process_id, proc.name, 0)
	overlay = _GhostOverlayWindow(enabled=proc.ghost_overlay_enabled, process_name=proc.name, process_id=process_id)
	overlay.push_log('Süreç başlatıldı')
	try:

		# Alt süreç zincirinde döngüsel çağrıları engelle (A -> B -> A)
		process_chain = []
		for raw_pid in (body.get('_process_chain', []) if isinstance(body.get('_process_chain', []), list) else []):
			try:
				process_chain.append(int(raw_pid))
			except Exception:
				continue
		if int(process_id) in process_chain:
			overlay.close()
			_runtime_finish(process_id)
			return JsonResponse({'ok': False, 'error': 'Döngüsel alt süreç çağrısı engellendi.'}, status=400)

		steps = _extract_runtime_steps(body, proc)
		if not steps:
			overlay.close()
			_runtime_finish(process_id)
			return JsonResponse({'ok': False, 'error': 'Çalıştırılacak adım yok.'}, status=400)
		_runtime_set_step(process_id, 0, len(steps), '')

		conn, conn_err = _resolve_connection_from_steps(steps)
		if conn_err:
			overlay.close()
			_runtime_finish(process_id)
			return JsonResponse({'ok': False, 'error': conn_err}, status=400)

		# Bildirim konfigürasyonu
		# Bildirim konfigürasyonu — tüm sap_fill adımlarındaki şablonlar taranır;
		# telegram_bot_id ve telegram_group_id içeren ilk şablon kullanılır.
		_notify_cfg = {}
		_notify_setup_notes = []
		_template_notify = {}
		try:
			# Önce bağlantı şablonunu dene, sonra diğer sap_fill adımlarını tara
			candidate_tpl_names = []
			_conn_tpl = str(conn.get('template_name', '') or '').strip()
			if _conn_tpl:
				candidate_tpl_names.append(_conn_tpl)
			for _step in steps:
				_stype = str(_step.get('step_type', '') or '').strip()
				if _stype == SapProcessStep.TYPE_SAP_FILL:
					_scfg = _step.get('config', {}) if isinstance(_step.get('config'), dict) else {}
					_stpl = str(_scfg.get('template_name', '') or '').strip()
					if _stpl and _stpl not in candidate_tpl_names:
						candidate_tpl_names.append(_stpl)
			for _tpl_name in candidate_tpl_names:
				_tpl = SAPTemplateService.get_template(_tpl_name)
				if not isinstance(_tpl, dict):
					continue
				_state = _tpl.get('state', {}) if isinstance(_tpl.get('state'), dict) else {}
				_notif = _state.get('notification', {}) if isinstance(_state.get('notification'), dict) else {}
				_bot_raw = str(_notif.get('telegram_bot_id', '') or '').strip()
				_grp_raw = str(_notif.get('telegram_group_id', '') or '').strip()
				if _bot_raw and _grp_raw:
					_template_notify = _notif
					break
				if not _template_notify and (_notif.get('mail_account_id')):
					_template_notify = _notif  # mail-only fallback; keep scanning for telegram
		except Exception:
			_template_notify = {}

		def _as_int(value):
			try:
				v = str(value or '').strip()
				return int(v) if v else None
			except Exception:
				return None

		if proc.telegram_notifications_enabled:
			tg_bot_id = _as_int(_template_notify.get('telegram_bot_id'))
			tg_group_id = _as_int(_template_notify.get('telegram_group_id'))
			if tg_bot_id and tg_group_id:
				_notify_cfg['telegram_bot_id'] = tg_bot_id
				_notify_cfg['telegram_group_id'] = tg_group_id
				_notify_cfg['telegram_voice_enabled'] = bool(proc.telegram_voice_enabled)
				for key in ('telegram_start_message', 'telegram_end_message'):
					value = str(_template_notify.get(key, '') or '').strip()
					if value:
						_notify_cfg[key] = value
			else:
				_notify_setup_notes.append('Telegram bildirim atlandı: şablonda Telegram bot/grup seçili değil.')

		if proc.mail_notifications_enabled:
			mail_id = _as_int(_template_notify.get('mail_account_id'))
			if mail_id:
				_notify_cfg['mail_account_id'] = mail_id
				for key in ('mail_to', 'mail_subject', 'mail_start_message', 'mail_end_message'):
					value = str(_template_notify.get(key, '') or '').strip()
					if value:
						_notify_cfg[key] = value
			else:
				_notify_setup_notes.append('Mail bildirimi atlandı: şablonda mail hesabı seçili değil.')


		# Başlangıç bildirimi
		if ('telegram_bot_id' in _notify_cfg and 'telegram_group_id' in _notify_cfg) or ('mail_account_id' in _notify_cfg):
			try:
				start_notes = _notify_sap_event(_notify_cfg, 'start')
			except Exception as _ex:
				start_notes = []
				_notify_setup_notes.append(f'Başlangıç bildirimi sırasında beklenmedik hata: {_ex}')
			for n in (start_notes or []):
				ch = n.get('channel', '')
				ok_flag = n.get('ok', False)
				msg = n.get('msg', '')
				if not ok_flag:
					_notify_setup_notes.append(f"{ch} hatası: {msg}")
				else:
					_notify_setup_notes.append(f"{ch} gönderildi.")

		try:
			upto_index = int(body.get('upto_index', len(steps) - 1))
		except (TypeError, ValueError):
			upto_index = len(steps) - 1
		try:
			start_index = int(body.get('start_index', 0))
		except (TypeError, ValueError):
			start_index = 0
		if start_index < 0:
			start_index = 0
		if start_index > len(steps) - 1:
			start_index = len(steps) - 1
		if upto_index < 0:
			upto_index = 0
		if upto_index > len(steps) - 1:
			upto_index = len(steps) - 1
		if start_index > upto_index:
			start_index = upto_index

		service = SAPScanService()
		logs = []
		runtime_state = {}
		runtime_state['row_results'] = []  # satır bazlı sonuç raporu için
		for note in (_notify_setup_notes or []):
			logs.append({'step': 0, 'type': 'notification', 'label': 'Bildirim', 'ok': False if 'atlandı' in str(note).casefold() else True, 'msg': str(note)})
			overlay.push_log(str(note))

		# ─── SAP Bağlantı Hazırlık (Retry) ───────────────────────────────
		# Süreç ayarlarında "SAP bağlantı yoksa tekrar dene" açıksa, ana
		# döngüye girmeden önce SAP'ın gerçekten ulaşılabilir olduğundan
		# emin oluruz. Aksi halde verilen aralıkta tekrar tekrar deneriz.
		if bool(getattr(proc, 'sap_retry_enabled', True)):
			_target_sys_id = str(conn.get('sys_id', '') or '').strip()
			if _target_sys_id:
				_retry_interval_min = max(1, int(getattr(proc, 'sap_retry_interval_minutes', 10) or 10))
				_retry_max_min = int(getattr(proc, 'sap_retry_max_duration_minutes', 180) or 0)
				_retry_interval_sec = _retry_interval_min * 60
				_retry_max_sec = _retry_max_min * 60 if _retry_max_min > 0 else 0

				def _retry_stop_check():
					_st = _runtime_get(process_id) or {}
					if bool(_st.get('stop_requested')):
						return True
					if getattr(overlay, 'stop_requested', False):
						return True
					return False

				def _retry_on_attempt(attempt_no, ok_attempt, err_attempt, next_wait_sec):
					if ok_attempt:
						_msg = f'SAP bağlantısı kuruldu (deneme #{attempt_no}).'
						overlay.push_log(_msg)
						logs.append({'step': 0, 'type': 'sap_retry', 'label': 'SAP Bağlantı', 'ok': True, 'msg': _msg})
					else:
						_min = max(1, int(round(next_wait_sec / 60)))
						_msg = f'SAP bağlantı denemesi #{attempt_no} başarısız: {err_attempt or "bilinmeyen hata"}. {_min} dk sonra tekrar denenecek.'
						overlay.push_log(_msg)
						logs.append({'step': 0, 'type': 'sap_retry', 'label': 'SAP Bağlantı', 'ok': False, 'msg': _msg})

				overlay.push_log(f'SAP bağlantısı kontrol ediliyor: {_target_sys_id}')
				_ok_conn, _attempts, _last_err = service.wait_for_sap_available(
					_target_sys_id,
					retry_interval_sec=_retry_interval_sec,
					max_duration_sec=_retry_max_sec,
					stop_check=_retry_stop_check,
					on_attempt=_retry_on_attempt,
				)
				if not _ok_conn:
					_err_summary = _last_err or 'Bilinmeyen SAP bağlantı hatası'
					return JsonResponse({
						'ok': False,
						'error': f'SAP bağlantısı sağlanamadı ({_attempts} deneme). Son hata: {_err_summary}',
						'logs': logs,
						'failed_at': 0,
					}, status=503)

		_end_notify_sent = False
		def _send_end_notify_once():
			nonlocal _end_notify_sent
			if _end_notify_sent:
				return
			if ('telegram_bot_id' in _notify_cfg and 'telegram_group_id' in _notify_cfg) or ('mail_account_id' in _notify_cfg):
				try:
					_notify_sap_event(_notify_cfg, 'end', logs)
				except Exception as _ex:
					try:
						logs.append({'step': 0, 'type': 'notification', 'label': 'Bildirim', 'ok': False, 'msg': f'Bitiş bildirimi hatası: {_ex}'})
					except Exception:
						pass
			_end_notify_sent = True

		def _ensure_session_ready():
			ok, payload = service.apply_to_screen(
				sys_id=conn.get('sys_id', ''),
				client=conn.get('client', ''),
				user=conn.get('user', ''),
				pwd=conn.get('pwd', ''),
				lang=conn.get('lang', 'TR'),
				t_code='',
				root_id=conn.get('root_id', 'wnd[0]'),
				extra_wait=conn.get('extra_wait', 0),
				actions=[],
				execute_f8=False,
			)
			return ok, payload

		def _wait_until_idle_or_popup(timeout_sec=30):
			"""Wait until session is idle, but return early if a popup window appears."""
			deadline = time.time() + max(1, min(int(timeout_sec or 30), 120))
			while time.time() < deadline:
				session = getattr(service, 'session', None)
				if session is None:
					return False, 'session_missing'

				try:
					children = getattr(session, 'Children', None)
					wnd_count = int(getattr(children, 'Count', 0) or 0) if children is not None else 0
					if wnd_count > 1:
						return True, 'popup_detected'
				except Exception:
					pass

				try:
					if not bool(getattr(session, 'Busy', False)):
						return True, 'idle'
				except Exception:
					return True, 'unknown'

				time.sleep(0.15)

			return False, 'timeout'

		def _handle_overlay_controls(failed_at_index):
			state = _runtime_get(process_id) or {}
			web_stop = bool(state.get('stop_requested'))
			web_paused = bool(state.get('paused'))
			if web_stop:
				overlay.stop_requested = True
			if web_paused != bool(overlay.paused):
				overlay.paused = web_paused
			if overlay.poll_controls():
				overlay.push_log('Kullanıcı süreci durdurdu')
				overlay.close()
				_runtime_finish(process_id)
				return JsonResponse({'ok': False, 'error': 'Süreç kullanıcı tarafından durduruldu.', 'logs': logs, 'failed_at': failed_at_index}, status=409)
			if overlay.wait_if_paused():
				overlay.push_log('Kullanıcı süreci durdurdu')
				overlay.close()
				_runtime_finish(process_id)
				return JsonResponse({'ok': False, 'error': 'Süreç kullanıcı tarafından durduruldu.', 'logs': logs, 'failed_at': failed_at_index}, status=409)
			return None

		i = start_index
		iteration_count = 0
		max_iterations = max(100, len(steps) * 20)
		while i < len(steps) and i <= upto_index:
			stop_response = _handle_overlay_controls(i)
			if stop_response is not None:
				return stop_response
			overlay.set_step(i + 1, len(steps), str(steps[i].get('label') or steps[i].get('step_type') or 'Adım'))
			pre_step_type = str(steps[i].get('step_type', '') or '').strip()
			if proc.office_express_auto_close and pre_step_type not in (
				SapProcessStep.TYPE_WINDOWS_DIALOG_ACTION,
				SapProcessStep.TYPE_WINDOWS_SCAN_DIALOGS,
				SapProcessStep.TYPE_CONVERT_SAP_EXPORT,
			):
				closed_before = _close_office_express_popups(service)
				if closed_before > 0:
					msg = f'Ofis Ekspres popup kapatıldı (adım öncesi): {closed_before}'
					logs.append({'step': i + 1, 'type': 'office_popup', 'label': 'Office Popup', 'ok': True, 'msg': msg})
					overlay.push_log(msg)
			if iteration_count >= max_iterations:
				overlay.close()
				_runtime_finish(process_id)
				return JsonResponse({'ok': False, 'error': 'Süreç maksimum iterasyon sınırına ulaştı.', 'logs': logs, 'failed_at': i}, status=500)
			iteration_count += 1
			step = steps[i]
			next_i = i + 1

			step_type = str(step.get('step_type', '') or '').strip()
			label = str(step.get('label', '') or '').strip()
			cfg = step.get('config', {}) if isinstance(step.get('config'), dict) else {}
			step_name = label or step_type
			continue_on_error = _as_bool(cfg.get('continue_on_error'))
			skip_step_end_auto_close = False

			if bool(cfg.get('disabled', False)):
				logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': 'Adım pasif olduğu için atlandı'})
				overlay.push_log(f'{step_name}: pasif adım atlandı')
				i = next_i
				continue

			try:
				if step_type == SapProcessStep.TYPE_SAP_FILL:
					tpl_name = str(cfg.get('template_name', '') or '').strip()
					if not tpl_name:
						return JsonResponse({'ok': False, 'error': f'{i + 1}. adımda şablon adı boş.', 'logs': logs, 'failed_at': i}, status=400)
					tpl = SAPTemplateService.get_template(tpl_name)
					# Şablon uygulamadan önce açık popup pencereleri (wnd[1]+) kapat
					if service.session is not None:
						try:
							_children = getattr(service.session, 'Children', None)
							_wnd_count = int(getattr(_children, 'Count', 0) or 0) if _children is not None else 0
							for _wi in range(_wnd_count - 1, 0, -1):
								try:
									_wnd = _children(_wi)
									_wnd.sendVKey(12)  # ESC
									service._wait_until_idle(service.session, timeout_sec=3, stable_checks=1)
								except Exception:
									pass
						except Exception:
							pass
					if not isinstance(tpl, dict):
						return JsonResponse({'ok': False, 'error': f'Şablon bulunamadı: {tpl_name}', 'logs': logs, 'failed_at': i}, status=404)

					state = tpl.get('state', {}) if isinstance(tpl.get('state'), dict) else {}
					_ensure_runtime_loop_state(runtime_state, state, cfg)
					actions = _build_actions_from_template_state_with_runtime(state, runtime_state=runtime_state)
					form = state.get('form', {}) if isinstance(state.get('form'), dict) else {}
					t_code = str(cfg.get('t_code_override', '') or form.get('t_code', '') or conn.get('t_code', '') or '').strip()

					ok, payload = service.apply_to_screen(
						sys_id=conn.get('sys_id', ''),
						client=conn.get('client', ''),
						user=conn.get('user', ''),
						pwd=conn.get('pwd', ''),
						lang=conn.get('lang', 'TR'),
						t_code=t_code,
						root_id=conn.get('root_id', 'wnd[0]'),
						extra_wait=conn.get('extra_wait', 0),
						actions=actions,
						execute_f8=False,
					)
					if not ok:
						return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)
					# Doldurma verilerini raporlama için biriktir
					filled_bucket = runtime_state.setdefault('_current_row_filled', [])
					for _act in actions:
						_act_type = str(_act.get('action_type', '') or '')
						if _act_type in ('sabit', 'dinamik', 'dongu', 'selectbox'):
							filled_bucket.append({
								'alan': str(_act.get('element_id', '') or ''),
								'kaynak': _act_type,
								'deger': str(_act.get('value', '') or ''),
							})
					loop_msg = ''
					if runtime_state.get('loop_values'):
						current_loop = str(runtime_state.get('loop_value', '') or '')
						loop_idx = int(runtime_state.get('loop_index', 0) or 0) + 1
						loop_total = len(runtime_state.get('loop_values') or [])
						loop_msg = f' | döngü: {loop_idx}/{loop_total} ({current_loop})'
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Şablon uygulandı ({tpl_name}), aksiyon: {len(actions)}{loop_msg}'})

				elif step_type == SapProcessStep.TYPE_SAP_RUN:
					key = str(cfg.get('key', 'F8') or 'F8').strip()
					if key.upper() == 'F8':
						ok, payload = service.apply_to_screen(
							sys_id=conn.get('sys_id', ''),
							client=conn.get('client', ''),
							user=conn.get('user', ''),
							pwd=conn.get('pwd', ''),
							lang=conn.get('lang', 'TR'),
							t_code='',
							root_id=conn.get('root_id', 'wnd[0]'),
							extra_wait=conn.get('extra_wait', 0),
							actions=[],
							execute_f8=True,
						)
						if not ok:
							return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': 'F8 çalıştırıldı'})
					else:
						ok, payload = _ensure_session_ready()
						if not ok:
							return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)
						vkey_map = {'ENTER': 0, 'F5': 5, 'F6': 6, 'F7': 7, 'F8': 8, 'ESCAPE': 12, 'ESC': 12}
						vk = vkey_map.get(key.upper())
						if vk is None:
							return JsonResponse({'ok': False, 'error': f'Desteklenmeyen tuş: {key}', 'logs': logs, 'failed_at': i}, status=400)
						service._wait_until_idle(service.session, timeout_sec=15)
						service.session.findById('wnd[0]').sendVKey(vk)
						service._wait_until_idle(service.session, timeout_sec=30)
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'{key} gönderildi'})

					delay_after_ms = int(cfg.get('delay_after_ms') or 0)
					if delay_after_ms > 0:
						time.sleep(min(delay_after_ms / 1000.0, 60.0))

				elif step_type == SapProcessStep.TYPE_SAP_KEY_PRESS:
					ok, payload = _ensure_session_ready()
					if not ok:
						return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)

					combo_text = str(cfg.get('combo_text', '') or '').strip()
					key_name = str(cfg.get('key_name', '') or '').strip()
					use_ctrl = bool(cfg.get('use_ctrl', False))
					use_alt = bool(cfg.get('use_alt', False))
					use_shift = bool(cfg.get('use_shift', False))
					use_win = bool(cfg.get('use_win', False))
					try:
						repeat_count = int(cfg.get('repeat_count') or 1)
					except (TypeError, ValueError):
						repeat_count = 1
					repeat_count = max(1, min(repeat_count, 20))
					try:
						delay_between_ms = int(cfg.get('delay_between_ms') or 120)
					except (TypeError, ValueError):
						delay_between_ms = 120
					delay_between_ms = max(0, min(delay_between_ms, 5000))

					last_combo = ''
					for r in range(repeat_count):
						ok_key, key_msg = _send_sap_hotkey(
							service,
							combo_text=combo_text,
							key=key_name,
							use_ctrl=use_ctrl,
							use_alt=use_alt,
							use_shift=use_shift,
							use_win=use_win,
						)
						if not ok_key:
							if continue_on_error:
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Tuş gönderimi hatası (devam): {key_msg}'})
								break
							return JsonResponse({'ok': False, 'error': f'Tuş gönderimi hatası: {key_msg}', 'logs': logs, 'failed_at': i}, status=500)
						last_combo = key_msg
						if r < repeat_count - 1 and delay_between_ms > 0:
							time.sleep(delay_between_ms / 1000.0)

					if last_combo:
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Tuş gönderildi: {last_combo} | tekrar: {repeat_count}'})

				elif step_type == SapProcessStep.TYPE_SAP_SELECT_OPTION:
					ok, payload = _ensure_session_ready()
					if not ok:
						return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)

					element_id = _normalize_session_element_id(cfg.get('element_id', ''))
					select_mode = str(cfg.get('select_mode', 'auto') or 'auto').strip().casefold()
					wait_timeout_sec = max(1, min(int(cfg.get('wait_timeout_sec') or 20), 60))
					obj = service._wait_for_element(service.session, element_id, timeout_sec=wait_timeout_sec)
					if not obj:
						return JsonResponse({'ok': False, 'error': f'Secim alanı bulunamadı: {element_id}', 'logs': logs, 'failed_at': i}, status=404)

					obj_type = str(getattr(obj, 'Type', '') or '').casefold()
					is_checkbox = ('checkbox' in obj_type) or ('/chk' in element_id.casefold())
					is_radio = ('radiobutton' in obj_type) or ('/rad' in element_id.casefold())

					if is_checkbox:
						if select_mode in ('toggle', 'auto'):
							obj.selected = not bool(getattr(obj, 'selected', False))
						elif select_mode in ('set_on', 'check', 'true'):
							obj.selected = True
						elif select_mode in ('set_off', 'uncheck', 'false'):
							obj.selected = False
						else:
							obj.selected = not bool(getattr(obj, 'selected', False))
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Checkbox uygulandı: {element_id} | mod: {select_mode}'})
					elif is_radio:
						try:
							obj.select()
						except Exception:
							try:
								obj.selected = True
							except Exception:
								obj.setFocus()
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Radio seçildi: {element_id}'})
					else:
						try:
							obj.select()
						except Exception:
							try:
								obj.selected = True
							except Exception:
								obj.setFocus()
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Secim uygulandı: {element_id}'})

					service._wait_until_idle(service.session, timeout_sec=5)

				elif step_type == SapProcessStep.TYPE_SAP_FILL_INPUT:
					ok, payload = _ensure_session_ready()
					if not ok:
						return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)

					element_id = _normalize_session_element_id(cfg.get('element_id', ''))
					value_source = str(cfg.get('value_source', 'static') or 'static').strip().casefold()
					value_text = str(cfg.get('value_text', '') or '')
					if value_source == 'excel_column':
						ok_excel, excel_value, excel_msg = _resolve_fill_input_excel_value(cfg, runtime_state)
						if not ok_excel:
							return JsonResponse({'ok': False, 'error': f'Excel değeri alınamadı: {excel_msg}', 'logs': logs, 'failed_at': i}, status=400)
						value_text = excel_value
						_runtime_push_log(process_id, f'{step_name}: {excel_msg}')
						# Excel'den okunan input değerini raporlama için biriktir
						_col_ref = str(cfg.get('excel_column', '') or cfg.get('column_ref', '') or '')
						runtime_state.setdefault('_current_row_filled', []).append({
							'alan': str(element_id or ''),
							'kaynak': f'excel({_col_ref})',
							'deger': str(excel_value or ''),
						})
					wait_timeout_sec = max(1, min(int(cfg.get('wait_timeout_sec') or 20), 60))
					obj = service._wait_for_element(service.session, element_id, timeout_sec=wait_timeout_sec)
					if not obj:
						return JsonResponse({'ok': False, 'error': f'Input alanı bulunamadı: {element_id}', 'logs': logs, 'failed_at': i}, status=404)

					ok_fill, fill_msg = fill_popup_input_value(obj, value_text)
					if not ok_fill:
						return JsonResponse({'ok': False, 'error': f'Input doldurulamadı: {element_id} | {fill_msg}', 'logs': logs, 'failed_at': i}, status=500)

					# Export diyalogunda (DY_PATH / DY_FILENAME) dosya zaten varsa önceden sil,
					# böylece "already exists" onay popup'ı açılmadan yazma devam eder.
					element_id_norm = str(element_id or '').casefold()
					filled_value = str(value_text or '').strip()
					if element_id_norm.endswith('/ctxtdy_path') or '/ctxtdy_path' in element_id_norm:
						runtime_state['_sap_export_path'] = filled_value
					if element_id_norm.endswith('/ctxtdy_filename') or '/ctxtdy_filename' in element_id_norm:
						runtime_state['_sap_export_filename'] = filled_value

					export_path = str(runtime_state.get('_sap_export_path') or '').strip()
					export_name = str(runtime_state.get('_sap_export_filename') or '').strip()
					if export_name and (
						element_id_norm.endswith('/ctxtdy_filename')
						or '/ctxtdy_filename' in element_id_norm
						or element_id_norm.endswith('/ctxtdy_path')
						or '/ctxtdy_path' in element_id_norm
					):
						candidate = export_name if os.path.isabs(export_name) else (os.path.join(export_path, export_name) if export_path else '')
						candidate = os.path.normpath(os.path.expandvars(os.path.expanduser(candidate))) if candidate else ''
						if candidate:
							try:
								if os.path.isfile(candidate):
									os.remove(candidate)
									del_msg = f'Mevcut dosya silindi (overwrite için): {candidate}'
									logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': del_msg})
									_runtime_push_log(process_id, f'{step_name}: {del_msg}')
							except Exception as del_ex:
								warn_msg = f'Mevcut dosya silinemedi, popup ile devam edilecek: {candidate} | {del_ex}'
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': warn_msg})
								_runtime_push_log(process_id, f'{step_name}: {warn_msg}')
					service._wait_until_idle(service.session, timeout_sec=5)
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Input dolduruldu: {element_id} = {value_text}'})

				elif step_type == SapProcessStep.TYPE_CONVERT_SAP_EXPORT:
					input_path = str(cfg.get('input_path', '') or '').strip()
					output_path = str(cfg.get('output_path', '') or '').strip()
					delimiter = str(cfg.get('delimiter', '\\t') or '\\t')
					encodings = str(cfg.get('encodings', 'utf-16,utf-16-le,cp1254,latin-1,utf-8') or 'utf-16,utf-16-le,cp1254,latin-1,utf-8').strip()
					drop_top_rows = max(0, int(cfg.get('drop_top_rows') or 0))
					drop_left_cols = max(0, int(cfg.get('drop_left_cols') or 0))
					drop_rows = str(cfg.get('drop_rows', '') or '').strip()
					drop_cols = str(cfg.get('drop_cols', '') or '').strip()
					skip_empty_lines = _as_bool(cfg.get('skip_empty_lines', True), True)
					trim_cells = _as_bool(cfg.get('trim_cells', True), True)
					strip_quotes = _as_bool(cfg.get('strip_quotes', True), True)

					if not input_path:
						msg = 'Dönüştürme için input_path zorunludur.'
						if continue_on_error:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': msg})
							continue
						return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=400)

					# SAP dosyasını yazmasını bekle (max 15 saniye)
					_file_wait_deadline = time.time() + 15
					_file_ready = False
					_last_size = -1
					_stable_count = 0
					while time.time() < _file_wait_deadline:
						try:
							if os.path.isfile(input_path):
								_cur_size = os.path.getsize(input_path)
								if _cur_size > 0 and _cur_size == _last_size:
									_stable_count += 1
									if _stable_count >= 2:
										_file_ready = True
										break
								else:
									_stable_count = 0
								_last_size = _cur_size
						except OSError:
							pass
						time.sleep(0.5)
					if not _file_ready:
						msg = f'SAP export dosyası 15sn içinde stabil olmadı: {input_path}'
						_runtime_push_log(process_id, f'{step_name}: {msg}')
						if continue_on_error:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': msg})
							continue
						return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=404)

					_runtime_push_log(process_id, f'{step_name}: dönüştürme başlıyor: {input_path}')
					# In-process dönüştürme (subprocess Windows + ağ sürücüsünde bloklayabiliyor)
					try:
						from core.management.commands.convert_sap_export import (
							_decode_escapes as _conv_decode,
							_parse_index_spec as _conv_parse_idx,
							_read_text_with_fallback as _conv_read,
							_remove_noise_rows as _conv_remove_noise_rows,
							_write_xlsx as _conv_write,
						)
						from pathlib import Path as _ConvPath
						_in_path = _ConvPath(input_path).expanduser().resolve()
						if output_path:
							_out_path = _ConvPath(output_path).expanduser().resolve()
						else:
							_out_path = _in_path.with_suffix('.xlsx')
						_delim = _conv_decode(delimiter or '\\t')
						_encs = [x.strip() for x in (encodings or '').split(',') if x.strip()] or ['utf-16','cp1254','latin-1','utf-8']
						_drop_rows_set = _conv_parse_idx(drop_rows)
						_drop_cols_set = _conv_parse_idx(drop_cols)
						_raw_text, _used_enc = _conv_read(_in_path, _encs)
						_raw_lines = _raw_text.splitlines()
						_rows = []
						_max_cols = 0
						for _line in _raw_lines:
							if skip_empty_lines and (not str(_line or '').strip()):
								continue
							_parts = str(_line).split(_delim)
							if trim_cells:
								_parts = [p.strip() for p in _parts]
							if strip_quotes:
								_parts = [p.replace('"', '') for p in _parts]
							_rows.append(_parts)
							if len(_parts) > _max_cols:
								_max_cols = len(_parts)
						if not _rows:
							raise RuntimeError('Okunan veri boş.')
						for _r in _rows:
							if len(_r) < _max_cols:
								_r.extend([''] * (_max_cols - len(_r)))
						if drop_top_rows > 0:
							_rows = _rows[drop_top_rows:]
						if drop_left_cols > 0:
							_rows = [r[drop_left_cols:] if len(r) > drop_left_cols else [] for r in _rows]
						if _drop_rows_set:
							_rows = [r for idx, r in enumerate(_rows) if idx not in _drop_rows_set]
						if _drop_cols_set:
							_rows = [[c for idx, c in enumerate(r) if idx not in _drop_cols_set] for r in _rows]
						_rows = _conv_remove_noise_rows(_rows)
						if not _rows:
							raise RuntimeError('Temizlikten sonra veri kalmadı.')
						_out_path.parent.mkdir(parents=True, exist_ok=True)
						_conv_write(_out_path, _rows)
						_conv_msg = f'Donusum tamamlandi | enc={_used_enc} | satir={len(_rows)} | sutun={max((len(r) for r in _rows), default=0)}'
					except Exception as conv_ex:
						msg = f'SAP export dönüştürme başarısız: {conv_ex}'
						_runtime_push_log(process_id, f'{step_name}: {msg}')
						if continue_on_error:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': msg})
							continue
						return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=500)

					msg = f'SAP export dönüştürüldü: {input_path} -> {_out_path} | {_conv_msg}'
					_runtime_push_log(process_id, f'{step_name}: {msg}')
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': msg})

				elif step_type == SapProcessStep.TYPE_WINDOWS_SCAN_DIALOGS:
					# Diagnostic: tüm açık Windows popup'larını tara ve logla
					title_filter = str(cfg.get('title_filter', '') or '').strip()
					scan_timeout_sec = max(1, min(int(cfg.get('scan_timeout_sec') or 10), 60))
					import time as _time_mod
					scan_msg = f'Windows popup tanı taraması başlıyor... (filter="{title_filter or "*"}", timeout={scan_timeout_sec}s)'
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': scan_msg})
					_runtime_push_log(process_id, f'{step_name}: {scan_msg}')

					deadline = _time_mod.time() + scan_timeout_sec
					found_any = False
					while _time_mod.time() < deadline:
						dialogs = scan_visible_dialogs(title_filter=title_filter)
						if dialogs:
							found_any = True
							break
						_time_mod.sleep(0.3)

					if not dialogs:
						no_result_msg = f'Hiç Windows popup bulunamadı (filter="{title_filter or "*"}", {scan_timeout_sec}s beklendi)'
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': no_result_msg})
						_runtime_push_log(process_id, f'{step_name}: {no_result_msg}')
					else:
						for dlg in dialogs:
							dlg_title = dlg.get('title', '?')
							dlg_hwnd = dlg.get('hwnd', '?')
							dlg_cls = dlg.get('class_name', '')
							btns = [b.get('text', '') for b in dlg.get('buttons', []) if b.get('text')]
							chks = [c.get('text', '') for c in dlg.get('checkboxes', []) if c.get('text')]
							rads = [r.get('text', '') for r in dlg.get('radios', []) if r.get('text')]
							inps = [inp.get('text', inp.get('class_name', '')) for inp in dlg.get('inputs', [])]
							detail_parts = []
							if btns:
								detail_parts.append(f'Butonlar: {btns}')
							if chks:
								detail_parts.append(f'Checkbox: {chks}')
							if rads:
								detail_parts.append(f'Radio: {rads}')
							if inps:
								detail_parts.append(f'Input: {inps}')
							detail = ' | '.join(detail_parts) if detail_parts else '(kontrol yok)'
							dlg_msg = f'DIALOG BULUNDU: hwnd={dlg_hwnd} class="{dlg_cls}" title="{dlg_title}" => {detail}'
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': dlg_msg})
							_runtime_push_log(process_id, f'{step_name}: {dlg_msg}')

				elif step_type == SapProcessStep.TYPE_WINDOWS_DIALOG_ACTION:
					window_title = str(cfg.get('window_title', '') or '').strip()
					button_text = str(cfg.get('button_text', '') or '').strip()
					checkbox_text = str(cfg.get('checkbox_text', '') or '').strip()
					if '?' in window_title and 'SAP' in window_title.upper():
						# Bozuk karakter içeren SAP GUI başlığı için genel eşleşme
						window_title = 'SAP GUI'
					title_match_mode = str(cfg.get('title_match_mode', 'contains') or 'contains').strip().casefold()
					button_match_mode = str(cfg.get('button_match_mode', 'contains') or 'contains').strip().casefold()
					checkbox_match_mode = str(cfg.get('checkbox_match_mode', 'contains') or 'contains').strip().casefold()
					apply_checkbox = _as_bool(cfg.get('apply_checkbox', False))
					checkbox_state_raw = cfg.get('checkbox_state', True)
					checkbox_state = _as_bool(checkbox_state_raw, True)
					# "Kararımı hatırla" benzeri akışlarda popup bir kez çıktıktan sonra tekrar gelmeyebilir.
					# Bu durumda adımın süreci kırmaması için not-found durumunu opsiyonel başarı kabul et.
					allow_not_found = _as_bool(cfg.get('allow_not_found', True), True)
					timeout_sec = max(1, min(int(cfg.get('timeout_sec') or 25), 120))
					poll_ms = max(50, min(int(cfg.get('poll_ms') or 250), 5000))
					ready_delay_ms = max(0, min(int(cfg.get('ready_delay_ms') or 350), 5000))

					if not window_title:
						msg = 'Windows popup başlığı boş bırakılamaz.'
						if continue_on_error:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': msg})
							continue
						return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=400)

					if not apply_checkbox:
						checkbox_text = ''

					wait_msg = f'Windows popup bekleniyor: {window_title} | timeout={timeout_sec}s'
					if checkbox_text:
						wait_msg = f'{wait_msg} | kontrol={checkbox_text}'
					if button_text:
						wait_msg = f'{wait_msg} | buton={button_text}'
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': wait_msg})
					_runtime_push_log(process_id, f'{step_name}: {wait_msg}')

					_popup_progress_state = {'last_remaining': None, 'last_status': ''}

					def _popup_progress(payload):
						if not isinstance(payload, dict):
							return
						remaining = payload.get('remaining_sec')
						status = str(payload.get('status', '') or '')
						if status == 'searching':
							if remaining == _popup_progress_state['last_remaining']:
								return
							_popup_progress_state['last_remaining'] = remaining
							_runtime_push_log(process_id, f'{step_name}: popup aranıyor... kalan ~{remaining}s')
						elif status == 'found' and _popup_progress_state['last_status'] != 'found':
							_popup_progress_state['last_status'] = 'found'
							title = str(payload.get('title', '') or window_title)
							_runtime_push_log(process_id, f'{step_name}: popup bulundu: {title}')

					dialog_kwargs = {
						'window_title': window_title,
						'button_text': button_text,
						'checkbox_text': checkbox_text,
						'checkbox_state': checkbox_state,
						'title_match_mode': title_match_mode,
						'button_match_mode': button_match_mode,
						'checkbox_match_mode': checkbox_match_mode,
						'allow_control_fallback': True,
						'timeout_sec': timeout_sec,
						'poll_ms': poll_ms,
						'ready_delay_ms': ready_delay_ms,
					}

					ok_dialog = False
					dialog_msg = 'Windows popup işlemi başlatılamadı.'
					worker_proc = None
					try:
						worker_script = os.path.join(os.path.dirname(__file__), 'windows_dialog_worker.py')
						worker_cmd = [
							sys.executable,
							worker_script,
							json.dumps(dialog_kwargs, ensure_ascii=False),
						]
						worker_proc = subprocess.Popen(
							worker_cmd,
							stdout=subprocess.PIPE,
							stderr=subprocess.PIPE,
							text=True,
						)

						worker_deadline = time.time() + timeout_sec + 4
						last_remain = None
						while time.time() < worker_deadline:
							if worker_proc is not None and worker_proc.poll() is not None:
								out, err = worker_proc.communicate(timeout=0.2)
								parsed = None
								try:
									parsed = json.loads(str(out or '').strip() or '{}')
								except Exception:
									parsed = None
								if isinstance(parsed, dict):
									ok_dialog = bool(parsed.get('ok'))
									dialog_msg = str(parsed.get('msg') or '').strip()
								else:
									ok_dialog = False
									dialog_msg = str(err or out or 'Windows popup worker çıktısı okunamadı.').strip()
								break

							remain = max(0, int(round(worker_deadline - time.time())))
							if remain != last_remain:
								last_remain = remain
								_popup_progress({'remaining_sec': remain, 'status': 'searching'})
							time.sleep(0.2)

						if not ok_dialog and (not dialog_msg):
							dialog_msg = 'Windows popup worker zaman aşımı veya yanıt vermedi.'
					except Exception as worker_ex:
							ok_dialog = False
							dialog_msg = f'Windows popup worker başlatılamadı: {worker_ex}'
					finally:
						try:
							if worker_proc is not None and worker_proc.poll() is None:
								worker_proc.terminate()
								worker_proc.wait(timeout=1)
						except Exception:
							pass
					if not ok_dialog:
						dialog_msg_n = str(dialog_msg or '').strip().casefold()
						dialog_not_found = (
							('dialog bulunamadi' in dialog_msg_n)
							or ('dialog bulunamadi.' in dialog_msg_n)
							or ('dialog bulunamadı' in dialog_msg_n)
							or ('dialog bulunamadı.' in dialog_msg_n)
							or ('dialog not found' in dialog_msg_n)
						)
						if dialog_not_found and allow_not_found:
							skip_msg = f'Popup bulunamadı, adım opsiyonel kabul edildi: {window_title}'
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': skip_msg})
							_runtime_push_log(process_id, f'{step_name}: {skip_msg}')
							continue
						_runtime_push_log(process_id, f'{step_name}: Windows popup işlemi başarısız: {dialog_msg}')
						if continue_on_error:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Windows popup işlemi başarısız (devam): {dialog_msg}'})
							continue
						return JsonResponse({'ok': False, 'error': f'Windows popup işlemi başarısız: {dialog_msg}', 'logs': logs, 'failed_at': i}, status=500)

					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': dialog_msg})
					_runtime_push_log(process_id, f'{step_name}: {dialog_msg}')

				elif step_type == SapProcessStep.TYPE_SAP_PRESS_BUTTON:
					ok, payload = _ensure_session_ready()
					if not ok:
						return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)
					next_active_i = next_i
					while next_active_i < len(steps):
						_candidate_step = steps[next_active_i] if isinstance(steps[next_active_i], dict) else {}
						_candidate_cfg = _candidate_step.get('config', {}) if isinstance(_candidate_step.get('config'), dict) else {}
						if bool(_candidate_cfg.get('disabled', False)):
							next_active_i += 1
							continue
						break
					next_step_type = ''
					if next_active_i < len(steps):
						next_step_type = str(steps[next_active_i].get('step_type', '') or '').strip()
					windows_handoff_types = {
						SapProcessStep.TYPE_WINDOWS_DIALOG_ACTION,
						SapProcessStep.TYPE_WINDOWS_SCAN_DIALOGS,
					}
					handoff_to_windows_dialog = (next_step_type in windows_handoff_types)
					raw_button_id = str(cfg.get('button_id', '') or '').strip()
					button_id = _normalize_session_element_id(raw_button_id)
					wait_timeout_sec = max(1, min(int(cfg.get('wait_timeout_sec') or 25), 60))
					_runtime_push_log(process_id, f'{step_name}: buton adımı başlıyor | hedef={button_id or raw_button_id} | sonraki={next_step_type or "-"}')
					ctx_cmd = _parse_toolbar_context_command(raw_button_id)
					if ctx_cmd:
						shell = service._wait_for_element(service.session, ctx_cmd['shell_id'], timeout_sec=wait_timeout_sec)
						if not shell:
							return JsonResponse({'ok': False, 'error': f'Context shell bulunamadı: {ctx_cmd["shell_id"]}', 'logs': logs, 'failed_at': i}, status=404)
						if ctx_cmd.get('command'):
							try:
								shell.pressToolbarContextButton(ctx_cmd['command'])
							except Exception as ex_ctx:
								return JsonResponse({'ok': False, 'error': f'Context toolbar komutu çalıştırılamadı ({ctx_cmd["command"]}): {ex_ctx}', 'logs': logs, 'failed_at': i}, status=500)
						if ctx_cmd.get('menu_item'):
							try:
								shell.selectContextMenuItem(ctx_cmd['menu_item'])
							except Exception as ex_item:
								return JsonResponse({'ok': False, 'error': f'Context menu seçimi yapılamadı ({ctx_cmd["menu_item"]}): {ex_item}', 'logs': logs, 'failed_at': i}, status=500)
						if not handoff_to_windows_dialog:
							_wait_until_idle_or_popup(timeout_sec=min(wait_timeout_sec, 30))
						ctx_msg = f'Context komutu uygulandı ({ctx_cmd["shell_id"]})'
						if ctx_cmd.get('command'):
							ctx_msg = f'{ctx_msg}: {ctx_cmd["command"]}'
						if ctx_cmd.get('menu_item'):
							ctx_msg = f'{ctx_msg} -> {ctx_cmd["menu_item"]}'
						wait_state = ''
						if handoff_to_windows_dialog:
							skip_step_end_auto_close = True
							ctx_msg = f'{ctx_msg} | windows popup adımına hızlı geçiş'
						else:
							_, wait_state = _wait_until_idle_or_popup(timeout_sec=1)
							if wait_state == 'popup_detected':
								ctx_msg = f'{ctx_msg} | popup algılandı, sonraki adıma geçiliyor'
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': ctx_msg})
					else:
						button = service._wait_for_element(service.session, button_id, timeout_sec=wait_timeout_sec)
						if not button:
							return JsonResponse({'ok': False, 'error': f'Buton bulunamadı: {button_id}', 'logs': logs, 'failed_at': i}, status=404)
						if handoff_to_windows_dialog:
							# SAP COM senkron: button.press() Windows güvenlik popup açılınca bloklar.
							# Çözüm: önce popup watcher subprocess'ini başlat, sonra button.press() çağır.
							# press() bloke olurken subprocess popup'ı bulur ve kapatır → press() unblock olur.
							next_step_cfg = steps[next_active_i].get('config', {}) if next_active_i < len(steps) else {}
							if not isinstance(next_step_cfg, dict):
								next_step_cfg = {}
							watcher_window_title = str(next_step_cfg.get('window_title', '') or '').strip()
							watcher_button_text  = str(next_step_cfg.get('button_text', '') or '').strip()
							watcher_checkbox_text = str(next_step_cfg.get('checkbox_text', '') or '').strip()
							watcher_apply_chk     = _as_bool(next_step_cfg.get('apply_checkbox', False))
							watcher_chk_state     = _as_bool(next_step_cfg.get('checkbox_state', True), True)
							watcher_timeout       = max(15, min(int(next_step_cfg.get('timeout_sec') or 30), 120))
							if '?' in watcher_window_title and 'SAP' in watcher_window_title.upper():
								watcher_window_title = 'SAP GUI'
							if not watcher_apply_chk:
								watcher_checkbox_text = ''
							watcher_kwargs = {
								'window_title':     watcher_window_title or 'SAP GUI',
								'button_text':      watcher_button_text,
								'checkbox_text':    watcher_checkbox_text,
								'checkbox_state':   watcher_chk_state,
								'title_match_mode': str(next_step_cfg.get('title_match_mode', 'contains') or 'contains').strip().casefold(),
								'button_match_mode': str(next_step_cfg.get('button_match_mode', 'contains') or 'contains').strip().casefold(),
								'checkbox_match_mode': str(next_step_cfg.get('checkbox_match_mode', 'contains') or 'contains').strip().casefold(),
								'allow_control_fallback': True,
								'timeout_sec':      watcher_timeout,
								'poll_ms':          max(50, min(int(next_step_cfg.get('poll_ms') or 250), 2000)),
								'ready_delay_ms':   max(0,  min(int(next_step_cfg.get('ready_delay_ms') or 300), 3000)),
							}
							worker_script = os.path.join(os.path.dirname(__file__), 'windows_dialog_worker.py')
							watcher_proc = subprocess.Popen(
								[sys.executable, worker_script, json.dumps(watcher_kwargs, ensure_ascii=False)],
								stdout=subprocess.PIPE,
								stderr=subprocess.PIPE,
								text=True,
							)
							_runtime_push_log(process_id, f'{step_name}: popup watcher başlatıldı ({watcher_kwargs["window_title"]}), şimdi butona basılıyor...')
							# Şimdi button.press() / menu.select() çağır — popup gelince subprocess kapatır, bu döner
							try:
								_invoke_button_or_menu(button, button_id)
							except Exception as _press_ex:
								_runtime_push_log(process_id, f'{step_name}: press()/select() exception: {_press_ex}')
							# press() döndü (popup kapandı veya hata), subprocess sonucunu oku
							try:
								watcher_parsed = None
								quick_deadline = time.time() + 2.5
								while time.time() < quick_deadline:
									if watcher_proc.poll() is not None:
										watcher_out, watcher_err = watcher_proc.communicate(timeout=0.2)
										watcher_parsed = json.loads(str(watcher_out or '').strip() or '{}')
										break
									time.sleep(0.1)
								if watcher_parsed is None:
									try:
										watcher_proc.terminate()
									except Exception:
										pass
									watcher_parsed = {'ok': False, 'msg': 'Dialog bulunamadı.'}
							except Exception:
								watcher_parsed = {'ok': False, 'msg': 'Dialog bulunamadı.'}
								try:
									watcher_proc.terminate()
								except Exception:
									pass
							watcher_ok  = bool(watcher_parsed.get('ok'))
							watcher_msg = str(watcher_parsed.get('msg') or '').strip() or 'watcher sonuç yok'
							watcher_msg_n = str(watcher_msg or '').strip().casefold()
							watcher_not_found = (
								('dialog bulunamadi' in watcher_msg_n)
								or ('dialog bulunamadi.' in watcher_msg_n)
								or ('dialog bulunamadı' in watcher_msg_n)
								or ('dialog bulunamadı.' in watcher_msg_n)
								or ('dialog not found' in watcher_msg_n)
							)
							skip_step_end_auto_close = True
							msg = f'Butona basıldı: {button_id} | popup watcher: {"OK" if watcher_ok else "FAIL"} | {watcher_msg}'
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': msg})
							_runtime_push_log(process_id, f'{step_name}: {msg}')
							# Handoff modunda watcher popup'ı paralel ele alır.
							# press() döndüyse aynı popup için ayrı windows_dialog_action bekletmesine girme.
							if next_step_type == SapProcessStep.TYPE_WINDOWS_DIALOG_ACTION:
								next_i = next_active_i + 1  # windows_dialog_action adımını atla
								_runtime_push_log(process_id, f'{step_name}: handoff tamamlandı, windows_dialog_action adımı atlanıyor')
						else:
							_invoke_button_or_menu(button, button_id)
							wait_state = ''
							_, wait_state = _wait_until_idle_or_popup(timeout_sec=min(wait_timeout_sec, 30))
							_action_word = 'Menü seçildi' if _is_menu_target(button, button_id) else 'Butona basıldı'
							msg = f'{_action_word}: {button_id}'
							if wait_state == 'popup_detected':
								msg = f'{msg} | popup algılandı, popup adımına geçiliyor'
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': msg})

					explicit_target_idx = _resolve_step_target_index(
						steps,
						cfg,
						step_no_key='next_step_no',
						step_id_key='next_step_id',
					)
					if explicit_target_idx is not None:
						next_i = explicit_target_idx
						logs.append({
							'step': i + 1,
							'type': step_type,
							'label': step_name,
							'ok': True,
							'msg': f'Buton sonrası yönlendirme aktif: hedef adım {explicit_target_idx + 1}'
						})
						_runtime_push_log(process_id, f'{step_name}: buton sonrası hedef adım {explicit_target_idx + 1}')

				elif step_type == SapProcessStep.TYPE_SAP_SELECT_ROW:
					ok, payload = _ensure_session_ready()
					if not ok:
						return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)
					grid_id = str(cfg.get('grid_id', '') or '').strip()
					row_text_contains = str(cfg.get('row_text_contains', '') or '').strip()
					normalized_grid_id = _normalize_session_element_id(grid_id)
					wait_timeout_sec = max(1, min(int(cfg.get('wait_timeout_sec') or 25), 60))
					popup_wait_sec = max(1, min(int(cfg.get('popup_wait_sec') or wait_timeout_sec), wait_timeout_sec))
					allow_main_fallback = bool(cfg.get('allow_main_fallback', True))
					grid = None
					grid_source = ''
					if normalized_grid_id:
						# Popup grid hedeflenmişse önce popup penceresini bekle.
						if normalized_grid_id.startswith('wnd[1]/'):
							service._wait_for_element(service.session, 'wnd[1]', timeout_sec=popup_wait_sec)
						grid = _find_grid(
							service,
							grid_id=normalized_grid_id,
							timeout_sec=wait_timeout_sec,
							grid_type='detail' if 'wnd[1]' in normalized_grid_id else 'main',
						)
						grid_source = normalized_grid_id
					else:
						# Grid ID yoksa önce popup detay tablosunu aktif olarak bekle.
						service._wait_for_element(service.session, 'wnd[1]', timeout_sec=popup_wait_sec)
						grid = _find_grid(service, grid_id='', timeout_sec=popup_wait_sec, grid_type='detail')
						if grid is not None:
							grid_source = 'auto-detail'
						elif allow_main_fallback:
							remain_timeout = max(1, wait_timeout_sec - popup_wait_sec)
							grid = _find_grid(service, grid_id='', timeout_sec=remain_timeout, grid_type='main')
							grid_source = 'auto-main'
						else:
							grid = None
							grid_source = 'auto-detail-only'
					if not grid:
						gid_msg = normalized_grid_id or 'auto-detect'
						return JsonResponse({'ok': False, 'error': f'Grid bulunamadı: {gid_msg}', 'logs': logs, 'failed_at': i}, status=404)
					resolved_by = 'row_index'
					row_index = 0
					if row_text_contains:
						resolved_row = _resolve_grid_row_by_text(grid, row_text_contains)
						if resolved_row is None:
							return JsonResponse({'ok': False, 'error': f'Satır metni bulunamadı: {row_text_contains}', 'logs': logs, 'failed_at': i}, status=404)
						row_index = int(resolved_row)
						resolved_by = 'row_text_contains'
					else:
						try:
							row_index = max(0, int(cfg.get('row_index') or 1) - 1)
						except (TypeError, ValueError):
							row_index = 0
					ok_select, applied_idx, err_msg = _select_row_on_grid(grid, row_index)
					if not ok_select:
						return JsonResponse({'ok': False, 'error': f'Grid satırı seçilemedi: {err_msg}', 'logs': logs, 'failed_at': i}, status=500)
					service._wait_until_idle(service.session, timeout_sec=5)
					copy_row_to_memory = bool(cfg.get('copy_row_to_memory', False))
					select_hint = f'{applied_idx + 1} ({grid_source})'
					if resolved_by == 'row_text_contains':
						select_hint = f'{select_hint} | metin: {row_text_contains}'
					if copy_row_to_memory:
						# Seçilen satırın verilerini runtime_state'e kopyala
						row_data = _read_grid_row_data(grid, applied_idx)
						runtime_state.update(row_data)
						row_data_hint = ', '.join(f'{k}={v}' for k, v in row_data.items()) if row_data else 'veri okunamadı'
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Grid satırı seçildi: {select_hint} | hafızaya kopyalandı | {row_data_hint}'})
					else:
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Grid satırı seçildi: {select_hint}'})

				elif step_type == SapProcessStep.TYPE_SAP_POPUP_DECIDE:
					ok, payload = _ensure_session_ready()
					if not ok:
						return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)
					popup_root_id = _normalize_session_element_id(cfg.get('popup_root_id', 'wnd[1]') or 'wnd[1]')
					popup = service._wait_for_element(service.session, popup_root_id, timeout_sec=max(1, min(int(cfg.get('timeout_sec') or 5), 30)))
					if not popup:
						if bool(cfg.get('fail_if_not_found')):
							return JsonResponse({'ok': False, 'error': 'Beklenen popup bulunamadı.', 'logs': logs, 'failed_at': i}, status=404)
						if bool(cfg.get('next_if_not_found', True)):
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': 'Popup bulunamadı, sonraki adıma geçildi'})
							i = next_i
							continue
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': 'Popup bulunamadı, süreç bu adımda sonlandırıldı'})
						overlay.push_log('Popup bulunamadı, süreç sonlandırıldı')
						overlay.close()
						_runtime_finish(process_id)
						return JsonResponse({'ok': True, 'logs': logs, 'ran_until': i, 'connection_template': conn.get('template_name', '')})
					title = str(getattr(popup, 'Text', '') or '').strip()
					popup_legacy_text = _collect_popup_text_legacy(service.session, popup_root_id, limit=220)
					popup_message_text = _collect_popup_message_text(service.session, popup_root_id, limit=120)
					popup_deep_text = _collect_node_text(popup, limit=220)
					popup_text = ' | '.join([p for p in [popup_legacy_text, popup_message_text, popup_deep_text] if p])
					title_contains = _normalize_match_text(str(cfg.get('popup_title_contains', '') or ''))
					text_contains = _normalize_match_text(str(cfg.get('popup_text_contains', '') or ''))
					title_norm = _normalize_match_text(title)
					popup_text_norm = _normalize_match_text(popup_text)
					matched = True
					if title_contains and title_contains not in title_norm:
						matched = False
					if text_contains and text_contains not in popup_text_norm:
						matched = False
					fallback_used = False
					if not matched and bool(cfg.get('allow_question_popup_fallback', True)):
						title_ok = (not title_contains) or (title_contains in title_norm)
						has_yes = _popup_has_button_by_text(popup, ['evet', 'yes'])
						has_no = _popup_has_button_by_text(popup, ['hayır', 'hayir', 'no'])
						if title_ok and has_yes and has_no:
							# Eski süreçteki gibi soru popup'ını başlık+buton deseniyle yakala.
							# Özellikle HTML kontrol içinde metin dönmeyen popup'lar için.
							matched = True
							fallback_used = True
					if not matched:
						if bool(cfg.get('fail_if_not_match')):
							return JsonResponse({'ok': False, 'error': f'Popup eşleşmedi. Başlık: {title} | Metin: {popup_text}', 'logs': logs, 'failed_at': i}, status=400)
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Popup geldi ama eşleşmedi. Başlık: {title}'})
						i = next_i
						continue
					if fallback_used:
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': 'Popup metni alınamadı; başlık ve Evet/Hayır buton desenine göre eşleşti (fallback).'})
					action = str(cfg.get('popup_action', '') or '').strip().casefold()
					if action == 'close_escape':
						popup.sendVKey(12)
					elif action == 'close_enter':
						popup.sendVKey(0)
					elif action == 'press_yes':
						ok_btn, btn_msg = _press_popup_button_by_text(popup, ['evet', 'yes', 'ok', 'tamam'])
						if not ok_btn:
							return JsonResponse({'ok': False, 'error': f'Popup evet butonu bulunamadı: {btn_msg}', 'logs': logs, 'failed_at': i}, status=404)
					elif action == 'press_no':
						ok_btn, btn_msg = _press_popup_button_by_text(popup, ['hayır', 'hayir', 'no'])
						if not ok_btn:
							return JsonResponse({'ok': False, 'error': f'Popup hayır/no butonu bulunamadı: {btn_msg}', 'logs': logs, 'failed_at': i}, status=404)
					elif action == 'press_cancel':
						ok_btn, btn_msg = _press_popup_button_by_text(popup, ['iptal', 'cancel', 'vazgeç', 'vazgec'])
						if not ok_btn:
							return JsonResponse({'ok': False, 'error': f'Popup iptal butonu bulunamadı: {btn_msg}', 'logs': logs, 'failed_at': i}, status=404)
					elif action == 'press_button_id':
						button_id = _normalize_session_element_id(cfg.get('popup_button_id', ''))
						button = service._wait_for_element(service.session, button_id, timeout_sec=5)
						if not button:
							return JsonResponse({'ok': False, 'error': f'Popup butonu bulunamadı: {button_id}', 'logs': logs, 'failed_at': i}, status=404)
						button.press()
					elif action == 'select_radio_id':
						radio_id = _normalize_session_element_id(cfg.get('popup_radio_id', ''))
						radio = service._wait_for_element(service.session, radio_id, timeout_sec=5)
						if not radio:
							return JsonResponse({'ok': False, 'error': f'Popup radio bulunamadı: {radio_id}', 'logs': logs, 'failed_at': i}, status=404)
						ok_radio, radio_msg = select_popup_radio_by_id(radio)
						if not ok_radio:
							return JsonResponse({'ok': False, 'error': f'Popup radio seçilemedi: {radio_id} | {radio_msg}', 'logs': logs, 'failed_at': i}, status=500)
					elif action == 'fill_input_id':
						input_id = _normalize_session_element_id(cfg.get('popup_input_id', ''))
						input_value = str(cfg.get('popup_input_value', '') or '')
						popup_input = service._wait_for_element(service.session, input_id, timeout_sec=5)
						if not popup_input:
							return JsonResponse({'ok': False, 'error': f'Popup input bulunamadı: {input_id}', 'logs': logs, 'failed_at': i}, status=404)
						ok_input, input_msg = fill_popup_input_value(popup_input, input_value)
						if not ok_input:
							return JsonResponse({'ok': False, 'error': f'Popup input doldurulamadı: {input_id} | {input_msg}', 'logs': logs, 'failed_at': i}, status=500)
					service._wait_until_idle(service.session, timeout_sec=10)
					mail_msg = _safe_send_popup_mail(
						cfg,
						mail_enabled=proc.mail_notifications_enabled,
						runtime_state=runtime_state,
						notification_cfg=_notify_cfg,
						popup_title=title,
						popup_text=popup_text,
					)
					log_msg = f'Popup işlendi. Başlık: {title}'
					if mail_msg:
						log_msg = f'{log_msg} | {mail_msg}'
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': log_msg})
					on_match_action = str(cfg.get('on_match_action', 'next_step') or 'next_step').strip().casefold()
					if on_match_action == 'goto_step':
						target_idx = _resolve_step_target_index(
							steps,
							cfg,
							step_no_key='popup_target_step',
							step_id_key='popup_target_step_id',
						)
						if target_idx is not None:
							next_i = target_idx
					elif on_match_action == 'press_button':
						btn_id_raw = str(cfg.get('popup_next_button_id', '') or '').strip()
						btn_id = _normalize_session_element_id(btn_id_raw)
						if not btn_id:
							return JsonResponse({'ok': False, 'error': 'Popup sonrası basılacak buton ID boş.', 'logs': logs, 'failed_at': i}, status=400)
						btn = service._wait_for_element(service.session, btn_id, timeout_sec=8)
						if not btn:
							return JsonResponse({'ok': False, 'error': f'Popup sonrası buton bulunamadı: {btn_id}', 'logs': logs, 'failed_at': i}, status=404)
						try:
							btn.press()
						except Exception as ex_btn:
							return JsonResponse({'ok': False, 'error': f'Popup sonrası butona basılamadı ({btn_id}): {ex_btn}', 'logs': logs, 'failed_at': i}, status=500)
						service._wait_until_idle(service.session, timeout_sec=10)
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Popup sonrası butona basıldı: {btn_id}'})
					elif on_match_action == 'loop_next':
						loop_idx = _find_loop_next_step_index(steps, i)
						if loop_idx is not None:
							next_i = loop_idx
						else:
							advanced, loop_msg, target_idx = _advance_loop_runtime(steps, runtime_state, i)
							if advanced:
								next_i = target_idx
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Popup sonrası loop_next adımı olmadan döngü ilerletildi: {loop_msg}'})
							else:
								if loop_msg == 'Döngü değerleri tamamlandı':
									logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': 'Döngü değerleri tamamlandı, süreç bitirildi'})
									overlay.push_log('Döngü değerleri tamamlandı, süreç bitirildi')
									overlay.close()
									_runtime_finish(process_id)
									_send_end_notify_once()
									return JsonResponse({'ok': True, 'logs': logs, 'ran_until': i, 'connection_template': conn.get('template_name', '')})
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Popup sonrası loop_next adımı bulunamadı ve döngü ilerletilemedi: {loop_msg}'})
					elif on_match_action == 'stop':
						overlay.push_log('Popup adımı sonrası süreç durduruldu')
						overlay.close()
						_runtime_finish(process_id)
						_send_end_notify_once()
						return JsonResponse({'ok': True, 'logs': logs, 'ran_until': i, 'connection_template': conn.get('template_name', '')})
					elif on_match_action == 'fail':
						return JsonResponse({'ok': False, 'error': f'Popup eşleşti ve hata aksiyonu tetiklendi. Başlık: {title}', 'logs': logs, 'failed_at': i}, status=400)

				elif step_type == SapProcessStep.TYPE_SAP_WAIT:
					ok, payload = _ensure_session_ready()
					if not ok:
						return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)

					timeout_sec = max(1, min(int(cfg.get('timeout_sec') or 60), 600))
					poll_ms = max(200, min(int(cfg.get('poll_ms') or 1000), 10000))
					screen_title = str(cfg.get('screen_title', '') or '').strip().casefold()

					if not screen_title:
						service._wait_until_idle(service.session, timeout_sec=min(timeout_sec, 30))
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': 'Ekran hazır (idle) olarak algılandı'})
					else:
						deadline = time.time() + timeout_sec
						found = False
						last_countdown = None
						while time.time() < deadline:
							stop_response = _handle_overlay_controls(i)
							if stop_response is not None:
								return stop_response
							remain = max(0, int(deadline - time.time()))
							if remain != last_countdown:
								last_countdown = remain
								wait_msg = f'Ekran bekleniyor: {cfg.get("screen_title", "")} | kalan: {remain}s'
								overlay.push_log(wait_msg)
							service._wait_until_idle(service.session, timeout_sec=3, stable_checks=1)
							title = service._get_window_title(service.session).casefold()
							if screen_title in title:
								found = True
								overlay.push_log(f'Ekran geldi: {cfg.get("screen_title", "")}')
								break
							time.sleep(poll_ms / 1000.0)
						if not found:
							timeout_action = str(cfg.get('on_timeout_action', 'fail') or 'fail').strip().casefold()
							if timeout_action == 'next_step':
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Ekran gelmedi, sonraki adıma geçildi: {cfg.get("screen_title", "")}'})
							elif timeout_action == 'goto_step':
								target_idx = _resolve_step_target_index(
									steps,
									cfg,
									step_no_key='timeout_target_step',
									step_id_key='timeout_target_step_id',
								)
								if target_idx is not None:
									next_i = target_idx
									logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Ekran gelmedi, {target_idx + 1}. adıma dönüldü: {cfg.get("screen_title", "")}'})
								else:
									logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Ekran gelmedi, hedef adım bulunamadı: {cfg.get("screen_title", "")}'})
							elif timeout_action == 'loop_next':
								loop_idx = _find_loop_next_step_index(steps, i)
								if loop_idx is not None:
									next_i = loop_idx
									logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Ekran gelmedi, döngüde sonraki elemana geçildi ({loop_idx + 1}. adım): {cfg.get("screen_title", "")}'})
								else:
									advanced, loop_msg, target_idx = _advance_loop_runtime(steps, runtime_state, i)
									if advanced:
										next_i = target_idx
										logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Ekran gelmedi, loop_next adımı olmadan döngü ilerletildi: {loop_msg}'})
									else:
										if loop_msg == 'Döngü değerleri tamamlandı':
											logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': 'Döngü değerleri tamamlandı, süreç bitirildi'})
											overlay.push_log('Döngü değerleri tamamlandı, süreç bitirildi')
											overlay.close()
											_runtime_finish(process_id)
											_send_end_notify_once()
											return JsonResponse({'ok': True, 'logs': logs, 'ran_until': i, 'connection_template': conn.get('template_name', '')})
										logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Ekran gelmedi, loop_next adımı bulunamadı ve döngü ilerletilemedi: {loop_msg}'})
							else:
								return JsonResponse({'ok': False, 'error': f'Beklenen ekran başlığı zaman aşımına uğradı: {cfg.get("screen_title", "")}', 'logs': logs, 'failed_at': i}, status=408)
						else:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Ekran geldi: {cfg.get("screen_title", "")}'})

				elif step_type == SapProcessStep.TYPE_SAP_ACTION:
					raw_actions = cfg.get('actions', []) if isinstance(cfg.get('actions'), list) else []
					ok, payload = _ensure_session_ready()
					if not ok:
						return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)

					for a in raw_actions:
						stop_response = _handle_overlay_controls(i)
						if stop_response is not None:
							return stop_response
						if not isinstance(a, dict):
							continue
						a_type = str(a.get('type', '') or '').strip()
						a_value = str(a.get('value', '') or '').strip()

						if a_type == 'grid_select_row':
							try:
								grid = find_alv_grid(service.session, None, grid_type="main")
								if grid:
									if a_value.lower() == 'current':
										# User'ın SAP'ta tıkladığı satırı al
										row_idx = int(getattr(grid, 'currentCellRow', 0) or 0)
									else:
										row_idx = int(a_value or 0)
									ok_select, applied_idx, err_msg = _select_row_on_grid(grid, row_idx)
									if ok_select:
										logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Grid satırı seçildi: {applied_idx + 1}'})
									else:
										logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Grid seçim hatası: {err_msg}'})
								else:
									logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': 'Grid bulunamadı'})
							except Exception as ge:
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Grid seçim hatası: {ge}'})

						elif a_type == 'click_button' and a_value:
							try:
								btn = service.session.findById(a_value)
								if btn:
									btn.press()
									logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Buton tıklandı: {a_value}'})
								else:
									logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Buton bulunamadı: {a_value}'})
							except Exception as be:
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Buton tıklama hatası: {be}'})

						elif a_type == 'wait_ms':
							try:
								wait_ms = int(a_value or '0')
							except ValueError:
								wait_ms = 0
							if wait_ms > 0:
								time.sleep(min(wait_ms / 1000.0, 60.0))

						elif a_type in ('key_press', 'popup_close'):
							vkey_map = {'ENTER': 0, 'F5': 5, 'F6': 6, 'F7': 7, 'F8': 8, 'ESCAPE': 12, 'ESC': 12}
							key_name = a_value or ('Escape' if a_type == 'popup_close' else 'Enter')
							vk = vkey_map.get(key_name.upper(), 0)
							service._wait_until_idle(service.session, timeout_sec=10)
							service.session.findById('wnd[0]').sendVKey(vk)
							service._wait_until_idle(service.session, timeout_sec=20)

						elif a_type == 'select_variant' and ':' in a_value:
							element_id, key_val = a_value.split(':', 1)
							ok, payload = service.apply_to_screen(
								sys_id=conn.get('sys_id', ''),
								client=conn.get('client', ''),
								user=conn.get('user', ''),
								pwd=conn.get('pwd', ''),
								lang=conn.get('lang', 'TR'),
								t_code='',
								root_id=conn.get('root_id', 'wnd[0]'),
								extra_wait=conn.get('extra_wait', 0),
								actions=[{'element_id': element_id.strip(), 'action_type': 'selectbox', 'value': key_val.strip()}],
								execute_f8=False,
							)
							if not ok:
								return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)

					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Aksiyon adımı uygulandı ({len(raw_actions)} işlem)'})

				elif step_type == SapProcessStep.TYPE_FTP_LIST:
					account_id = cfg.get('ftp_account_id')
					account = FTPAccount.objects.filter(pk=account_id, is_active=True).first()
					if not account:
						msg = f'Geçerli FTP hesabı bulunamadı (id={account_id}).'
						if continue_on_error:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Hata (devam): {msg}'})
							continue
						return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=400)
					remote_path = str(cfg.get('remote_path', '') or account.remote_base_path or '.').strip()
					file_pattern = str(cfg.get('file_pattern', '*') or '*').strip()
					items = _ftp_list_files(account, remote_path=remote_path, file_pattern=file_pattern)
					preview = ', '.join(items[:6])
					if len(items) > 6:
						preview += ' ...'
					if not preview:
						preview = 'liste boş'
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'{len(items)} dosya listelendi ({account.name}) | {preview}'})

				elif step_type == SapProcessStep.TYPE_FTP_DOWNLOAD:
					account_id = cfg.get('ftp_account_id')
					account = FTPAccount.objects.filter(pk=account_id, is_active=True).first()
					if not account:
						msg = f'Geçerli FTP hesabı bulunamadı (id={account_id}).'
						if continue_on_error:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Hata (devam): {msg}'})
							continue
						return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=400)
					remote_path = str(cfg.get('remote_path', '') or account.remote_base_path or '.').strip()
					local_path = str(cfg.get('local_path', '') or '').strip()
					file_pattern = str(cfg.get('file_pattern', '*') or '*').strip()
					try:
						limit = int(cfg.get('limit') or 0)
					except (TypeError, ValueError):
						limit = 0
					downloaded = _ftp_download(account, remote_path=remote_path, local_path=local_path, file_pattern=file_pattern, limit=limit)
					files_preview = ', '.join([os.path.basename(x) for x in downloaded[:6]])
					if len(downloaded) > 6:
						files_preview += ' ...'
					if not files_preview:
						files_preview = 'indirilen dosya yok'
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'{len(downloaded)} dosya indirildi ({account.name}) | {files_preview}'})

				elif step_type == SapProcessStep.TYPE_FTP_UPLOAD:
					account_id = cfg.get('ftp_account_id')
					account = FTPAccount.objects.filter(pk=account_id, is_active=True).first()
					if not account:
						msg = f'Geçerli FTP hesabı bulunamadı (id={account_id}).'
						if continue_on_error:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Hata (devam): {msg}'})
							continue
						return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=400)
					local_file = str(cfg.get('local_file', '') or '').strip()
					remote_path = str(cfg.get('remote_path', '') or account.remote_base_path or '.').strip()
					uploaded = _ftp_upload(account, local_file=local_file, remote_path=remote_path)
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Dosya yüklendi ({account.name}): {uploaded}'})

				elif step_type == SapProcessStep.TYPE_SAP_CLOSE:
					ok, payload = _ensure_session_ready()
					if not ok:
						return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)
					try:
						service.session.findById('wnd[0]').close()
					except Exception:
						pass
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': 'SAP kapatma komutu gönderildi'})

				elif step_type == SapProcessStep.TYPE_SHOW_MESSAGE:
					msg_title = str(cfg.get('title', '') or '').strip() or 'Bilgi'
					msg_body = str(cfg.get('message', '') or '').strip() or 'Devam etmek için Tamam butonuna basın.'
					overlay.push_log(f'Mesaj gösteriliyor: {msg_title}')
					msg_ok, msg_err = overlay.show_message(msg_title, msg_body)
					if not msg_ok:
						return JsonResponse({'ok': False, 'error': msg_err, 'logs': logs, 'failed_at': i}, status=500)
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Mesaj gösterildi ve onaylandı: {msg_title}'})

				elif step_type == SapProcessStep.TYPE_EXCEL_LOOP_NEXT:
					advanced, excel_msg, target_idx = _advance_excel_loop_runtime(steps, runtime_state, cfg)
					# Yeni satıra geçildiğinde doldurma geçmişini sıfırla
					runtime_state['_current_row_filled'] = []
					if advanced:
						next_i = target_idx
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': excel_msg})
					else:
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': excel_msg})

				elif step_type == SapProcessStep.TYPE_EXCEL_ROW_LOG:
					row_status = str(cfg.get('status', '') or 'ok').strip().lower() or 'ok'
					row_reason = _resolve_placeholders(str(cfg.get('reason', '') or '').strip(), runtime_state)
					row_entry = {
						'cari': str(runtime_state.get('loop_value', '') or ''),
						'excel_row': (runtime_state.get('excel_loop_index', 0) or 0) + 1,
						'status': row_status,
						'reason': row_reason,
						'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
						'filled_data': list(runtime_state.get('_current_row_filled') or []),
					}
					runtime_state.setdefault('row_results', []).append(row_entry)
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True,
						'msg': f'Satır sonucu yazıldı → Cari: {row_entry["cari"]}, Satır: {row_entry["excel_row"]}, Durum: {row_status.upper()}, Neden: {row_reason or "-"}'})
					overlay.push_log(f'Satır kaydedildi: {row_entry["cari"]} / {row_entry["excel_row"]} → {row_status.upper()}')

				elif step_type == SapProcessStep.TYPE_SEND_REPORT_MAIL:
					row_results = runtime_state.get('row_results', [])
					output_dir = str(cfg.get('output_dir', '') or '').strip()
					stamp_now = datetime.now().strftime('%Y%m%d_%H%M%S')
					output_path = None
					if output_dir:
						import os as _os_r
						_os_r.makedirs(output_dir, exist_ok=True)
						output_path = _os_r.path.join(output_dir, f'rapor_{stamp_now}.xlsx')
					try:
						xlsx_path = _generate_row_report_xlsx(row_results, output_path=output_path)
					except Exception as _xe:
						err_msg = f'Rapor xlsx üretilemedi: {_xe}'
						if continue_on_error:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': err_msg})
						else:
							return JsonResponse({'ok': False, 'error': err_msg, 'logs': logs, 'failed_at': i}, status=500)
					else:
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True,
							'msg': f'Rapor xlsx oluşturuldu: {xlsx_path} ({len(row_results)} satır)'})
						overlay.push_log(f'Rapor oluşturuldu: {xlsx_path}')
						runtime_state['last_report_xlsx'] = xlsx_path

						mail_id = cfg.get('mail_account_id')
						if mail_id:
							_mail_acc = MailAccount.objects.filter(pk=mail_id, is_active=True).first()
							if _mail_acc:
								ok_count = sum(1 for r in row_results if str(r.get('status', '')).lower() == 'ok')
								err_count = sum(1 for r in row_results if str(r.get('status', '')).lower() == 'error')
								skip_count = sum(1 for r in row_results if str(r.get('status', '')).lower() == 'skip')
								_to = str(cfg.get('to', '') or '').strip() or _mail_acc.email
								_subj = _resolve_placeholders(str(cfg.get('subject', '') or '').strip(), runtime_state) \
									or f'Saggio RPA Raporu – {stamp_now}'
								_body_tpl = str(cfg.get('body', '') or '').strip() or (
									f'Merhaba,\n\nSüreç tamamlandı.\n\nToplam satır: {len(row_results)}\n'
									f'Başarılı: {ok_count}\nHatalı: {err_count}\nAtlandı: {skip_count}\n\n'
									f'Rapor ektedir.\n\nZaman: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
								)
								_body = _resolve_placeholders(_body_tpl, runtime_state)
								mail_ok, mail_msg = _send_mail_message(_mail_acc, _to, _subj, _body, attachment_path=xlsx_path)
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': mail_ok,
									'msg': f'Rapor maili → {"Gönderildi" if mail_ok else "Hata"}: {mail_msg}'})
								overlay.push_log(f'Rapor maili: {"Gönderildi" if mail_ok else "HATA – " + mail_msg}')
							else:
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False,
									'msg': f'Mail hesabı bulunamadı (id={mail_id})'})

				elif step_type == SapProcessStep.TYPE_LOOP_NEXT:
					advanced, loop_msg, target_idx = _advance_loop_runtime(steps, runtime_state, i)
					if not advanced:
						if loop_msg == 'Döngü değerleri tamamlandı':
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': 'Döngü değerleri tamamlandı, süreç bitirildi'})
							overlay.push_log('Döngü değerleri tamamlandı, süreç bitirildi')
							overlay.close()
							_runtime_finish(process_id)
							_send_end_notify_once()
							return JsonResponse({'ok': True, 'logs': logs, 'ran_until': i, 'connection_template': conn.get('template_name', '')})
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': loop_msg})
					else:
						next_i = target_idx
						logs.append({
							'step': i + 1,
							'type': step_type,
							'label': step_name,
							'ok': True,
							'msg': loop_msg
						})

				elif step_type == SapProcessStep.TYPE_SAP_SCAN:
					ok, payload = _ensure_session_ready()
					if not ok:
						return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)
				
					try:
						# Detail grid'i bul
						detail_grid = find_alv_grid(service.session, None, grid_type="detail")
						if not detail_grid:
							return JsonResponse({'ok': False, 'error': 'Detail grid bulunamadı', 'logs': logs, 'failed_at': i}, status=500)
					
						# Yöntem deteksiyonu (detail grid ID'sine bakarak)
						d_id = str(getattr(detail_grid, 'ID', '') or '')
						yontem = 1
						if 'ALV_HT_FR/' in d_id:
							yontem = 2
						elif 'ALV_HT_FRUG/' in d_id:
							yontem = 3
					
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Detail grid bulundu (Yöntem: {yontem}, Grid ID: {d_id[:50]})'})
					
						# Yöntem 3 için malzeme kodları ve ürün grubu kodları oku
						if yontem == 3:
							try:
								# Column headers'ı al
								d_col_headers = str(getattr(detail_grid, 'colHeaders', '') or '')
								d_col_names = d_col_headers.split('\t')
							
								# Satır sayısı al
								d_row_count = int(getattr(detail_grid, 'visibleRowCount', 0) or 0)
							
								malzeme_kodlari = []
								logs_detail = []
							
								for dr in range(d_row_count):
									try:
										# Scroll işlemi: satırı görüntülemek için grid scroll'ını ayarla
										vis_count = int(getattr(detail_grid, 'visibleRowCount', 0) or 0)
										first_vis = int(getattr(detail_grid, 'firstVisibleRow', 0) or 0)
										if dr < first_vis or dr >= first_vis + vis_count:
											detail_grid.firstVisibleRow = max(0, dr - 1)
											service._wait_until_idle(service.session, timeout_sec=3)
									
										# Column 2 ve 3'ten değerleri al
										if len(d_col_names) > 3:
											v1 = str(detail_grid.getCellValue(dr, d_col_names[2])).strip() if d_col_names[2] else ''
											v2 = str(detail_grid.getCellValue(dr, d_col_names[3])).strip() if len(d_col_names) > 3 and d_col_names[3] else ''
											if v1:
												malzeme_kodlari.append({'malzeme': v1.lstrip('0'), 'urun_grubu': v2.lstrip('0')})
												logs_detail.append(f'Satır {dr}: malzeme={v1} urungrup={v2}')
									except Exception as row_err:
										logs_detail.append(f'Satır {dr} okuma hatası: {row_err}')
							
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'{len(malzeme_kodlari)} malzeme kodu okudu', 'detail': logs_detail[:10]})
							except Exception as detail_err:
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Detail grid okuma hatası: {detail_err}'})
						else:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Yöntem {yontem} için scan henüz implement edilmedi'})
				
					except Exception as scan_err:
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Scan hatası: {scan_err}'})

				elif step_type == SapProcessStep.TYPE_RUN_PROCESS:
					try:
						target_process_id = int(cfg.get('target_process_id') or 0)
					except (TypeError, ValueError):
						target_process_id = 0
					if target_process_id <= 0:
						msg = 'Çalıştırılacak alt süreç seçilmedi.'
						if continue_on_error:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Hata (devam): {msg}'})
						else:
							return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=400)
						continue
					if target_process_id == int(process_id):
						msg = 'Bir süreç kendisini alt süreç olarak çağıramaz.'
						if continue_on_error:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Hata (devam): {msg}'})
						else:
							return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=400)
						continue

					target_proc = SapProcess.objects.filter(pk=target_process_id).first()
					if not target_proc:
						msg = f'Alt süreç bulunamadı (id={target_process_id}).'
						if continue_on_error:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Hata (devam): {msg}'})
						else:
							return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=404)
						continue

					child_req = HttpRequest()
					child_req.method = 'POST'
					child_req.META = getattr(request, 'META', {}).copy()
					if hasattr(request, 'user'):
						child_req.user = request.user
					child_body = {
						'_process_chain': process_chain + [int(process_id)],
					}
					child_req._body = json.dumps(child_body).encode('utf-8')

					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Alt süreç başlatıldı: {target_proc.name}'})
					child_resp = sap_process_run_preview(child_req, target_process_id)
					child_status = int(getattr(child_resp, 'status_code', 500) or 500)
					try:
						child_data = json.loads((child_resp.content or b'{}').decode('utf-8'))
					except Exception:
						child_data = {}
					child_ok = bool(child_data.get('ok')) and child_status < 400
					if child_ok:
						child_logs = child_data.get('logs', []) if isinstance(child_data.get('logs'), list) else []
						logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'Alt süreç tamamlandı: {target_proc.name} | log: {len(child_logs)}'})
					else:
						child_err = str(child_data.get('error') or f'Alt süreç hata döndürdü (HTTP {child_status})')
						if continue_on_error:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Alt süreç hatası (devam): {child_err}'})
						else:
							return JsonResponse({'ok': False, 'error': f'Alt süreç başarısız: {child_err}', 'logs': logs, 'failed_at': i}, status=500 if child_status < 400 else child_status)

				elif step_type == SapProcessStep.TYPE_IF_ELSE:
					condition_type = str(cfg.get('condition_type', 'element_value') or 'element_value').strip()

					if condition_type == 'popup_exists':
						# ── Popup varlık / içerik kontrolü ──────────────────────────
						ok, payload = _ensure_session_ready()
						if not ok:
							# Oturum yoksa popup da yok
							popup_title, popup_text_raw = None, None
						else:
							popup_title, popup_text_raw = None, None
							try:
								children = getattr(service.session, 'Children', None)
								count = int(getattr(children, 'Count', 0) or 0) if children is not None else 0
								for _pidx in range(1, count):
									_wnd = None
									try:
										_wnd = children(_pidx)
									except Exception:
										try:
											_wnd = children.Item(_pidx)
										except Exception:
											pass
									if _wnd is None:
										continue
									_pid = _normalize_session_element_id(str(getattr(_wnd, 'Id', '') or '').strip())
									popup_title = str(getattr(_wnd, 'Text', '') or '').strip()
									_lt = _collect_popup_text_legacy(service.session, _pid or 'wnd[1]', limit=120)
									_mt = _collect_popup_message_text(service.session, _pid or 'wnd[1]', limit=80)
									popup_text_raw = ' | '.join(p for p in [_lt, _mt] if p)
									break  # İlk popup yeterli
							except Exception as _pe:
								overlay.push_log(f'IF/ELSE popup okuma hatası: {_pe}')

						popup_full = ' '.join(p for p in [popup_title, popup_text_raw] if p).strip()

						if _as_bool(cfg.get('save_popup_text', False)):
							runtime_state['last_popup_text'] = popup_full

						target_idx = None
						base_msg = ''
						if popup_full:
							# Kural listesini tara: ilk eşleşen kazanır
							rules = cfg.get('popup_rules') or []
							if isinstance(rules, str):
								try:
									import json as _json
									rules = _json.loads(rules)
								except Exception:
									rules = []
							matched_rule = None
							for _rule in (rules or []):
								kw = str(_rule.get('keyword', '') or '').strip()
								if not kw or kw.casefold() in popup_full.casefold():
									matched_rule = _rule
									break
							if matched_rule:
								target_idx = _resolve_rule_target_index(steps, matched_rule)
								rule_label = str(matched_rule.get('label', '') or matched_rule.get('keyword', '') or '?')
								base_msg = f'IF/ELSE Popup: "{popup_title or "?"}" → kural eşleşti: {rule_label}'
							else:
								target_idx = _resolve_step_target_index(
									steps,
									cfg,
									step_no_key='popup_nomatch_step_no',
									step_id_key='popup_nomatch_step_id',
								)
								base_msg = f'IF/ELSE Popup: açık popup var ama kural eşleşmedi ("{popup_title or "?"}")'
						else:
							target_idx = _resolve_step_target_index(
								steps,
								cfg,
								step_no_key='no_popup_step_no',
								step_id_key='no_popup_step_id',
							)
							base_msg = 'IF/ELSE Popup: Açık popup penceresi bulunamadı'

						if target_idx is None:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True,
								'msg': f'{base_msg} | hedef tanımsız, sonraki adıma geçildi'})
						else:
							next_i = target_idx
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True,
								'msg': f'{base_msg} | hedef adım: {target_idx + 1}'})
						overlay.push_log(base_msg)

					elif condition_type == 'status_message':
						ok, payload = _ensure_session_ready()
						if not ok:
							return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)

						s_ok, s_text, s_type, s_err = _read_sap_statusbar(service.session)
						if not s_ok:
							msg = f'IF/ELSE status bar okunamadı: {s_err}'
							if continue_on_error:
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Hata (devam): {msg}'})
							else:
								return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=400)
							continue

						runtime_state['last_status_text'] = s_text
						runtime_state['last_status_type'] = s_type

						status_category = str(cfg.get('status_category', 'any') or 'any').strip().casefold()
						needle = str(cfg.get('status_text_contains', '') or '').strip().casefold()
						msg_fold = str(s_text or '').casefold()

						cat_ok = True
						if status_category == 'positive':
							cat_ok = (s_type in ('S',))
						elif status_category == 'negative':
							cat_ok = (s_type in ('E', 'A'))
						elif status_category == 'info':
							cat_ok = (s_type in ('W', 'I'))

						text_ok = (not needle) or (needle in msg_fold)
						matched = bool(cat_ok and text_ok)

						if matched:
							target_idx = _resolve_step_target_index(
								steps,
								cfg,
								step_no_key='status_match_step_no',
								step_id_key='status_match_step_id',
							)
							base_msg = f'IF/ELSE Status: [{s_type or "?"}] {s_text} | eşleşti'
						else:
							target_idx = _resolve_step_target_index(
								steps,
								cfg,
								step_no_key='status_nomatch_step_no',
								step_id_key='status_nomatch_step_id',
							)
							base_msg = f'IF/ELSE Status: [{s_type or "?"}] {s_text} | eşleşmedi'

						if target_idx is None:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'{base_msg} | hedef tanımsız, sonraki adıma geçildi'})
						else:
							next_i = target_idx
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'{base_msg} | hedef adım: {target_idx + 1}'})
						overlay.push_log(base_msg)

					else:
						# ── Mevcut: SAP alan değeri sayısal karşılaştırması ─────────
						ok, payload = _ensure_session_ready()
						if not ok:
							return JsonResponse({'ok': False, 'error': payload, 'logs': logs, 'failed_at': i}, status=500)

						element_id = str(cfg.get('element_id', '') or '').strip()
						if not element_id:
							msg = 'IF/ELSE için element_id zorunlu.'
							if continue_on_error:
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Hata (devam): {msg}'})
							else:
								return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=400)
							continue

						try:
							threshold = _parse_decimal_text(cfg.get('threshold', '1'))
						except Exception:
							threshold = Decimal('1')

						read_ok, raw_value, read_err = _read_sap_element_text(service.session, element_id)
						if not read_ok:
							msg = f'IF/ELSE alan okuma hatasi: {read_err}'
							if continue_on_error:
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Hata (devam): {msg}'})
							else:
								return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=400)
							continue

						try:
							numeric_value = _parse_decimal_text(raw_value)
						except Exception as parse_ex:
							msg = f'IF/ELSE sayi parse hatasi ({raw_value}): {parse_ex}'
							if continue_on_error:
								logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Hata (devam): {msg}'})
							else:
								return JsonResponse({'ok': False, 'error': msg, 'logs': logs, 'failed_at': i}, status=400)
							continue

						abs_value = abs(numeric_value)
						if abs_value > threshold:
							branch = 'gt'
							target_idx = _resolve_step_target_index(steps, cfg, step_no_key='gt_step_no', step_id_key='gt_step_id')
						elif abs_value < threshold:
							branch = 'lt'
							target_idx = _resolve_step_target_index(steps, cfg, step_no_key='lt_step_no', step_id_key='lt_step_id')
						else:
							branch = 'eq'
							target_idx = _resolve_step_target_index(steps, cfg, step_no_key='eq_step_no', step_id_key='eq_step_id')

						branch_map = {'gt': '>', 'lt': '<', 'eq': '='}
						base_msg = f'IF/ELSE: {element_id}={raw_value} | |deger|={abs_value} {branch_map.get(branch, "=")} {threshold}'
						if target_idx is None:
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'{base_msg} | hedef tanimsiz, sonraki adima gecildi'})
						else:
							next_i = target_idx
							logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': f'{base_msg} | hedef adim: {target_idx + 1}'})

				elif step_type == SapProcessStep.TYPE_IF_END:
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': True, 'msg': 'IF bloğu kapatıldı'})

				else:
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Bilinmeyen step_type: {step_type}'})

			except Exception as ex:
				if continue_on_error and step_type in (SapProcessStep.TYPE_FTP_LIST, SapProcessStep.TYPE_FTP_DOWNLOAD, SapProcessStep.TYPE_FTP_UPLOAD):
					logs.append({'step': i + 1, 'type': step_type, 'label': step_name, 'ok': False, 'msg': f'Hata (devam): {ex}'})
					overlay.push_log(f'Hata (devam): {ex}')
					continue
				overlay.close()
				_runtime_finish(process_id)
				return JsonResponse({'ok': False, 'error': str(ex), 'logs': logs, 'failed_at': i}, status=500)

			if proc.office_express_auto_close and not skip_step_end_auto_close and step_type not in (
				SapProcessStep.TYPE_WINDOWS_DIALOG_ACTION,
				SapProcessStep.TYPE_WINDOWS_SCAN_DIALOGS,
				SapProcessStep.TYPE_CONVERT_SAP_EXPORT,
			):
				closed_after = _close_office_express_popups(service)
				if closed_after > 0:
					msg = f'Ofis Ekspres popup kapatıldı (adım sonrası): {closed_after}'
					logs.append({'step': i + 1, 'type': 'office_popup', 'label': 'Office Popup', 'ok': True, 'msg': msg})
					overlay.push_log(msg)

			if logs:
				overlay.push_log(logs[-1].get('msg', 'Adım tamamlandı'))

			i = next_i
		overlay.push_log('Süreç tamamlandı')
		overlay.close()
		_runtime_finish(process_id)
		# Bitiş bildirimi
		_send_end_notify_once()
		return JsonResponse({'ok': True, 'logs': logs, 'ran_until': upto_index, 'connection_template': conn.get('template_name', '')})
	finally:
		try:
			overlay.close()
		except Exception:
			pass
		try:
			_runtime_finish(process_id)
		except Exception:
			pass
		# Tüm SAP ekranlarını kapat ve oturumu sonlandır
		try:
			if 'service' in locals() and service is not None:
				service.close_all_sap_windows()
		except Exception:
			pass


# ─── Telegram Bot Studio ─────────────────────────────────────────────────────

def telegram_bot_studio(request):
    """Bot menü yönetimi + sohbet simülatörü ana sayfası."""
    bots_qs = TelegramBot.objects.filter(is_active=True).order_by('name')
    bots = []
    for b in bots_qs:
        bots.append({
            'id': b.pk,
            'name': b.name,
            'bot_username': b.bot_username,
            'allowed_user_ids': b.allowed_user_ids or '',
            'webhook_secret': b.webhook_secret or '',
            'webhook_registered_url': b.webhook_registered_url or '',
        })
    sap_processes = list(SapProcess.objects.order_by('name').values('id', 'name'))
    return render(request, 'core/telegram_bot_studio.html', {
        'page_title': 'Telegram Bot Stüdyo',
        'page_subtitle': 'Bot menüleri oluşturun, buton-süreç bağlantısı kurun ve sohbet simülatörüyle test edin.',
        'bots_json': json.dumps(bots, ensure_ascii=False),
        'sap_processes_json': json.dumps(sap_processes, ensure_ascii=False),
    })


@require_POST
def telegram_bot_studio_menu_save(request):
    """Bir bot menüsünü (ve butonlarını) kaydeder ya da günceller."""
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

    bot_id       = body.get('bot_id')
    menu_id      = body.get('menu_id')
    name         = str(body.get('name', '')).strip()
    trigger      = str(body.get('trigger_command', '/start')).strip() or '/start'
    welcome      = str(body.get('welcome_message', '')).strip()
    buttons_raw  = body.get('buttons', [])

    if not bot_id or not name:
        return JsonResponse({'ok': False, 'error': 'bot_id ve name zorunludur.'}, status=400)

    bot = TelegramBot.objects.filter(pk=bot_id).first()
    if not bot:
        return JsonResponse({'ok': False, 'error': 'Bot bulunamadı.'}, status=404)

    if menu_id:
        menu = TelegramBotMenu.objects.filter(pk=menu_id, bot=bot).first()
        if not menu:
            return JsonResponse({'ok': False, 'error': 'Menü bulunamadı.'}, status=404)
        menu.name = name
        menu.trigger_command = trigger
        menu.welcome_message = welcome
        menu.save()
    else:
        menu = TelegramBotMenu.objects.create(
            bot=bot, name=name, trigger_command=trigger, welcome_message=welcome
        )

    # Mevcut butonları sil ve yeniden oluştur
    menu.buttons.all().delete()
    for btn in (buttons_raw if isinstance(buttons_raw, list) else []):
        label      = str(btn.get('label', '')).strip()
        row        = int(btn.get('row', 0))
        col        = int(btn.get('col', 0))
        process_id = btn.get('sap_process_id')
        if not label:
            continue
        proc = SapProcess.objects.filter(pk=process_id).first() if process_id else None
        TelegramBotButton.objects.create(menu=menu, label=label, sap_process=proc, row=row, col=col)

    return JsonResponse({'ok': True, 'menu_id': menu.pk})


def telegram_bot_studio_menus(request):
    """Bir bota ait menü listesini JSON olarak döner."""
    bot_id = request.GET.get('bot_id')
    if not bot_id:
        return JsonResponse({'ok': False, 'error': 'bot_id gerekli.'}, status=400)
    menus = TelegramBotMenu.objects.filter(bot_id=bot_id).prefetch_related('buttons__sap_process')
    result = []
    for m in menus:
        btns = []
        for b in m.buttons.all():
            btns.append({
                'id': b.pk,
                'label': b.label,
                'row': b.row,
                'col': b.col,
                'sap_process_id': b.sap_process_id,
                'sap_process_name': b.sap_process.name if b.sap_process else None,
            })
        result.append({
            'id': m.pk,
            'name': m.name,
            'trigger_command': m.trigger_command,
            'welcome_message': m.welcome_message,
            'is_active': m.is_active,
            'buttons': btns,
        })
    return JsonResponse({'ok': True, 'menus': result})


@require_POST
def telegram_bot_studio_menu_delete(request):
    """Bir menüyü siler."""
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)
    menu_id = body.get('menu_id')
    deleted, _ = TelegramBotMenu.objects.filter(pk=menu_id).delete()
    if not deleted:
        return JsonResponse({'ok': False, 'error': 'Menü bulunamadı.'}, status=404)
    return JsonResponse({'ok': True})


@require_POST
def telegram_bot_studio_simulate(request):
    """
    Sohbet simülatörü için mesaj/buton basımını işler.
    - type='command' : komuta göre menü döner (örn. /start)
    - type='button'  : butona bağlı sürecin adını döner
    """
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

    action_type = body.get('type')
    bot_id      = body.get('bot_id')

    if not bot_id:
        return JsonResponse({'ok': False, 'error': 'bot_id gerekli.'}, status=400)

    if action_type == 'command':
        command = str(body.get('command', '/start')).strip()
        menu = TelegramBotMenu.objects.filter(bot_id=bot_id, trigger_command=command, is_active=True).prefetch_related('buttons__sap_process').first()
        if not menu:
            return JsonResponse({'ok': True, 'response': {'type': 'text', 'text': f'Bu komut için kayıtlı bir menü bulunamadı: {command}'}})
        btns_by_row = {}
        for b in menu.buttons.all():
            btns_by_row.setdefault(b.row, []).append({'id': b.pk, 'label': b.label, 'sap_process_id': b.sap_process_id, 'sap_process_name': b.sap_process.name if b.sap_process else None})
        keyboard = [btns_by_row[r] for r in sorted(btns_by_row)]
        return JsonResponse({'ok': True, 'response': {'type': 'menu', 'text': menu.welcome_message, 'keyboard': keyboard, 'menu_id': menu.pk}})

    elif action_type == 'button':
        button_id = body.get('button_id')
        btn = TelegramBotButton.objects.select_related('sap_process').filter(pk=button_id, menu__bot_id=bot_id).first()
        if not btn:
            return JsonResponse({'ok': False, 'error': 'Buton bulunamadı.'}, status=404)
        if btn.sap_process:
            return JsonResponse({'ok': True, 'response': {
                'type': 'process_trigger',
                'text': f'✅ Süreç başlatılıyor: <b>{btn.sap_process.name}</b>',
                'process_id': btn.sap_process.pk,
                'process_name': btn.sap_process.name,
            }})
        else:
            return JsonResponse({'ok': True, 'response': {'type': 'text', 'text': '⚠️ Bu butona henüz bir süreç bağlanmamış.'}})

    return JsonResponse({'ok': False, 'error': 'Geçersiz type parametresi.'}, status=400)


# ─── Telegram Webhook + Yetkilendirme ─────────────────────────────────────────

def _normalize_allowed_user_ids(text):
    """Serbest metni (satır/virgül/boşluk ayrılmış) numerik user ID set'ine çevirir."""
    if not text:
        return set()
    out = set()
    for chunk in str(text).replace(',', '\n').replace(';', '\n').split('\n'):
        token = chunk.strip()
        if not token:
            continue
        # Sadece rakamları al (negatif chat_id de olabilir, başında '-')
        sign = ''
        if token.startswith('-'):
            sign = '-'
            token = token[1:]
        digits = ''.join(c for c in token if c.isdigit())
        if digits:
            try:
                out.add(int(sign + digits))
            except ValueError:
                pass
    return out


def _is_user_allowed_for_bot(bot, user_id):
    """Telegram user_id (int) izinli listesinde mi?"""
    if user_id is None:
        return False
    allowed = _normalize_allowed_user_ids(bot.allowed_user_ids or '')
    if not allowed:
        return False  # boş liste = kapalı bot (güvenlik öncelikli)
    try:
        return int(user_id) in allowed
    except (TypeError, ValueError):
        return False


def _telegram_api_call(token, method, payload):
    """Telegram Bot API'sini JSON POST ile çağırır."""
    if not token:
        return False, {'description': 'token_missing'}
    try:
        req = Request(
            f'https://api.telegram.org/bot{token}/{method}',
            data=json.dumps(payload).encode('utf-8'),
            method='POST',
            headers={'Content-Type': 'application/json'},
        )
        with urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode('utf-8') or '{}')
        return bool(data.get('ok')), data
    except HTTPError as e:
        try:
            body = json.loads(e.read().decode('utf-8') or '{}')
        except Exception:
            body = {'description': str(e)}
        return False, body
    except URLError as e:
        return False, {'description': f'baglanti_hatasi: {e.reason}'}
    except Exception as e:
        return False, {'description': str(e)}


def _telegram_send_with_keyboard(token, chat_id, text, inline_keyboard=None, parse_mode=None):
    payload = {'chat_id': chat_id, 'text': text or ''}
    if parse_mode:
        payload['parse_mode'] = parse_mode
    if inline_keyboard:
        payload['reply_markup'] = {'inline_keyboard': inline_keyboard}
    return _telegram_api_call(token, 'sendMessage', payload)


def _telegram_answer_callback(token, callback_id, text='', show_alert=False):
    payload = {'callback_query_id': callback_id, 'text': text or '', 'show_alert': bool(show_alert)}
    return _telegram_api_call(token, 'answerCallbackQuery', payload)


def _render_menu_inline_keyboard(menu):
    """TelegramBotMenu -> Telegram inline_keyboard listesi."""
    rows = {}
    for b in menu.buttons.all().order_by('row', 'col'):
        rows.setdefault(b.row, []).append({
            'text': b.label,
            'callback_data': f'btn:{b.pk}',
        })
    return [rows[r] for r in sorted(rows)]


def _trigger_sap_process_for_telegram(process_id, source_label=''):
    """
    Telegram callback'inden tetiklenen SAP süreci için arka plan başlatma.
    Mevcut runtime endpoint'ini iç HTTP isteği ile çalıştırır; CSRF/auth'tan etkilenmez
    çünkü Django test client kullanıyoruz.
    """
    import threading

    def _runner():
        try:
            from django.test import Client
            client = Client(enforce_csrf_checks=False)
            client.post(f'/sap-process/{process_id}/run/', data={}, content_type='application/json')
        except Exception:
            pass

    t = threading.Thread(target=_runner, name=f'tg-trigger-{process_id}', daemon=True)
    t.start()


@csrf_exempt
@require_POST
def telegram_bot_webhook(request, bot_id, secret):
    """Telegram'ın gönderdiği update'leri işler. URL: /tg/webhook/<bot_id>/<secret>/"""
    bot = TelegramBot.objects.filter(pk=bot_id, is_active=True).first()
    if not bot:
        return JsonResponse({'ok': False}, status=404)
    if not bot.webhook_secret or secret != bot.webhook_secret:
        return JsonResponse({'ok': False}, status=403)

    try:
        update = json.loads(request.body.decode('utf-8') or '{}')
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'ok': False}, status=400)

    process_telegram_update(bot, update)
    return JsonResponse({'ok': True})


def process_telegram_update(bot, update):
    """Tek bir Telegram update'ini (webhook veya polling) tutarlı şekilde işler."""
    token = bot.get_bot_token()

    # 1) Inline buton tıklaması
    cb = update.get('callback_query')
    if cb:
        from_user = (cb.get('from') or {})
        user_id = from_user.get('id')
        chat_id = (((cb.get('message') or {}).get('chat')) or {}).get('id')
        callback_id = cb.get('id')
        data = str(cb.get('data') or '')

        if not _is_user_allowed_for_bot(bot, user_id):
            _telegram_answer_callback(token, callback_id, 'Yetkiniz yok.', show_alert=True)
            return

        if data.startswith('btn:'):
            try:
                btn_pk = int(data.split(':', 1)[1])
            except (ValueError, IndexError):
                btn_pk = 0
            btn = TelegramBotButton.objects.select_related('sap_process', 'menu').filter(
                pk=btn_pk, menu__bot=bot
            ).first()
            if not btn:
                _telegram_answer_callback(token, callback_id, 'Buton bulunamadı.')
                return
            if not btn.sap_process:
                _telegram_answer_callback(token, callback_id, 'Bu butona süreç bağlanmamış.')
                return
            _telegram_answer_callback(token, callback_id, 'Süreç başlatılıyor…')
            _telegram_send_with_keyboard(
                token, chat_id,
                f'✅ Süreç başlatma talebi alındı:\n<b>{btn.sap_process.name}</b>\n\nİstek: {from_user.get("first_name") or from_user.get("username") or user_id}',
                parse_mode='HTML',
            )
            _trigger_sap_process_for_telegram(btn.sap_process.pk, source_label=str(user_id))
            return

        _telegram_answer_callback(token, callback_id, '')
        return

    # 2) Mesaj (komut)
    msg = update.get('message') or update.get('edited_message')
    if msg:
        from_user = (msg.get('from') or {})
        user_id = from_user.get('id')
        chat_id = ((msg.get('chat') or {}).get('id'))
        text = str(msg.get('text') or '').strip()

        if not _is_user_allowed_for_bot(bot, user_id):
            if chat_id:
                _telegram_send_with_keyboard(
                    token, chat_id,
                    'Bu botu kullanmaya yetkiniz yok.\nKullanıcı ID\'niz: <code>' + str(user_id) + '</code>',
                    parse_mode='HTML',
                )
            return

        if text:
            menu = TelegramBotMenu.objects.filter(
                bot=bot, trigger_command=text, is_active=True
            ).prefetch_related('buttons').first()
            if menu:
                kb = _render_menu_inline_keyboard(menu)
                _telegram_send_with_keyboard(token, chat_id, menu.welcome_message, inline_keyboard=kb)
                return
            _telegram_send_with_keyboard(
                token, chat_id,
                'Bu komut için tanımlı bir menü yok. Kullanılabilir komutları yöneticinizden öğrenin.'
            )


@require_POST
def telegram_bot_studio_bot_save(request):
    """Bot bazlı erişim ayarları (allowed_user_ids) kaydeder."""
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)
    bot_id = body.get('bot_id')
    bot = TelegramBot.objects.filter(pk=bot_id).first()
    if not bot:
        return JsonResponse({'ok': False, 'error': 'Bot bulunamadı.'}, status=404)

    raw = str(body.get('allowed_user_ids', '') or '')
    parsed = sorted(_normalize_allowed_user_ids(raw))
    bot.allowed_user_ids = '\n'.join(str(x) for x in parsed)
    bot.save()
    return JsonResponse({'ok': True, 'allowed_user_ids': bot.allowed_user_ids, 'count': len(parsed)})


@require_POST
def telegram_bot_studio_set_webhook(request):
    """Bu bot için Telegram'a setWebhook çağrısı yapar.
    Body: { bot_id, base_url } - base_url örn 'https://example.com' (sondaki / olmadan).
    """
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)
    bot_id = body.get('bot_id')
    base_url = str(body.get('base_url') or '').strip().rstrip('/')
    if not base_url:
        return JsonResponse({'ok': False, 'error': 'base_url gerekli (örn https://alanim.com).'}, status=400)
    if not (base_url.startswith('https://') or base_url.startswith('http://')):
        return JsonResponse({'ok': False, 'error': 'base_url http(s):// ile başlamalı.'}, status=400)

    bot = TelegramBot.objects.filter(pk=bot_id).first()
    if not bot:
        return JsonResponse({'ok': False, 'error': 'Bot bulunamadı.'}, status=404)
    if not bot.webhook_secret:
        import secrets as _sec
        bot.webhook_secret = _sec.token_urlsafe(32)
        bot.save()

    webhook_url = f"{base_url}/tg/webhook/{bot.pk}/{bot.webhook_secret}/"
    token = bot.get_bot_token()
    ok, data = _telegram_api_call(token, 'setWebhook', {
        'url': webhook_url,
        'allowed_updates': ['message', 'edited_message', 'callback_query'],
        'drop_pending_updates': True,
    })
    if ok:
        bot.webhook_registered_url = webhook_url
        bot.save()
        return JsonResponse({'ok': True, 'webhook_url': webhook_url, 'telegram': data})
    return JsonResponse({'ok': False, 'error': data.get('description') or 'telegram_api_error', 'telegram': data}, status=502)


def telegram_bot_studio_webhook_info(request):
    """Telegram'dan getWebhookInfo döner (durum kontrolü)."""
    bot_id = request.GET.get('bot_id')
    bot = TelegramBot.objects.filter(pk=bot_id).first()
    if not bot:
        return JsonResponse({'ok': False, 'error': 'Bot bulunamadı.'}, status=404)
    token = bot.get_bot_token()
    ok, data = _telegram_api_call(token, 'getWebhookInfo', {})
    if ok:
        return JsonResponse({'ok': True, 'info': data.get('result') or {}, 'expected_url': bot.webhook_registered_url})
    return JsonResponse({'ok': False, 'error': data.get('description') or 'telegram_api_error'}, status=502)


@require_POST
def telegram_bot_studio_delete_webhook(request):
    """Telegram'da deleteWebhook çağrısı yapar."""
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)
    bot = TelegramBot.objects.filter(pk=body.get('bot_id')).first()
    if not bot:
        return JsonResponse({'ok': False, 'error': 'Bot bulunamadı.'}, status=404)
    token = bot.get_bot_token()
    ok, data = _telegram_api_call(token, 'deleteWebhook', {'drop_pending_updates': True})
    if ok:
        bot.webhook_registered_url = ''
        bot.save()
        return JsonResponse({'ok': True})
    return JsonResponse({'ok': False, 'error': data.get('description') or 'telegram_api_error'}, status=502)


def _read_json_body(request):
	try:
		return json.loads(request.body or b'{}')
	except (json.JSONDecodeError, TypeError):
		return None


def _request_ip(request):
	forwarded = str(request.META.get('HTTP_X_FORWARDED_FOR', '') or '').strip()
	if forwarded:
		return forwarded.split(',')[0].strip()
	return str(request.META.get('REMOTE_ADDR', '') or '').strip()


def _authenticate_agent(request, body):
	agent_code = str(
		body.get('agent_code')
		or request.headers.get('X-Agent-Code')
		or ''
	).strip()
	token = str(
		body.get('token')
		or request.headers.get('X-Agent-Token')
		or ''
	).strip()

	if not agent_code or not token:
		return None, JsonResponse({'ok': False, 'error': 'agent_code ve token gerekli.'}, status=401)

	agent = RobotAgent.objects.filter(code=agent_code, is_enabled=True).first()
	if not agent:
		return None, JsonResponse({'ok': False, 'error': 'Ajan bulunamadı veya pasif.'}, status=404)

	if not agent.verify_token(token):
		return None, JsonResponse({'ok': False, 'error': 'Token doğrulaması başarısız.'}, status=401)

	return agent, None


def _serialize_job(job):
	return {
		'job_id': job.pk,
		'command_type': job.command_type,
		'sap_process_id': job.sap_process_id,
		'priority': job.priority,
		'payload': job.payload or {},
		'created_at': job.created_at.isoformat() if job.created_at else None,
	}


def _safe_release_version(version):
	return re.sub(r'[^0-9A-Za-z._-]+', '_', str(version or '')).strip('_') or 'release'


def _package_zip_path(version):
	safe_version = _safe_release_version(version)
	package_dir = os.path.join(str(settings.MEDIA_ROOT), 'robot_agent', 'packages')
	os.makedirs(package_dir, exist_ok=True)
	file_name = f'SaggioRobotAgentPackage_{safe_version}.zip'
	full_path = os.path.join(package_dir, file_name)
	return full_path, file_name


def _create_agent_event(agent, level, message, *, job=None, extra=None):
	level_norm = str(level or 'info').strip().lower()
	if level_norm not in {'info', 'warning', 'error'}:
		level_norm = 'info'
	RobotAgentEvent.objects.create(
		agent=agent,
		job=job,
		level=level_norm,
		message=str(message or '').strip()[:2000],
		extra=extra if isinstance(extra, dict) else {},
	)


@csrf_exempt
@require_POST
def agent_register(request):
	body = _read_json_body(request)
	if body is None:
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	agent, err = _authenticate_agent(request, body)
	if err:
		return err

	agent.name = str(body.get('name') or agent.name or agent.code).strip()[:180]
	agent.machine_name = str(body.get('machine_name') or '').strip()[:120]
	agent.host_name = str(body.get('host_name') or request.get_host() or '').strip()[:180]
	agent.os_user = str(body.get('os_user') or '').strip()[:120]
	agent.agent_version = str(body.get('agent_version') or '').strip()[:40]
	agent.ip_address = _request_ip(request)[:64]
	capabilities = body.get('capabilities')
	if isinstance(capabilities, dict):
		agent.capabilities = capabilities
	agent.status = 'online'
	agent.mark_seen(startup=True)
	agent.save()

	pending = RobotJob.objects.filter(status='queued').filter(
		Q(target_agent__isnull=True) | Q(target_agent=agent)
	).count()
	return JsonResponse({
		'ok': True,
		'agent': {
			'code': agent.code,
			'name': agent.name,
			'status': agent.status,
			'last_seen_at': agent.last_seen_at.isoformat() if agent.last_seen_at else None,
		},
		'pending_jobs': pending,
	})


@csrf_exempt
@require_POST
def agent_heartbeat(request):
	body = _read_json_body(request)
	if body is None:
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	agent, err = _authenticate_agent(request, body)
	if err:
		return err

	if agent.status != 'busy':
		agent.status = 'online'
	agent_version = str(body.get('agent_version') or '').strip()
	if agent_version:
		agent.agent_version = agent_version[:40]
	agent.ip_address = _request_ip(request)[:64]
	agent.mark_seen(startup=False)
	agent.save(update_fields=['status', 'agent_version', 'ip_address', 'last_seen_at', 'updated_at'])

	return JsonResponse({
		'ok': True,
		'server_time': timezone.now().isoformat(),
		'agent_status': agent.status,
		'desired_version': agent.desired_version,
	})


@csrf_exempt
@require_POST
def agent_check_update(request):
	body = _read_json_body(request)
	if body is None:
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	agent, err = _authenticate_agent(request, body)
	if err:
		return err

	current_version = str(body.get('current_version') or agent.agent_version or '').strip()
	if current_version:
		agent.agent_version = current_version[:40]

	latest = RobotAgentRelease.objects.filter(is_active=True).order_by('-created_at').first()
	desired = str(agent.desired_version or '').strip() or (latest.version if latest else '')
	available = bool(desired and current_version and desired != current_version)

	agent.mark_seen(startup=False)
	agent.save(update_fields=['agent_version', 'last_seen_at', 'updated_at'])

	result = {
		'ok': True,
		'agent_code': agent.code,
		'current_version': current_version,
		'desired_version': desired,
		'update_available': available,
	}
	if latest:
		release_download_url = latest.download_url or f"/api/robot-agent/releases/download/{latest.version}/"
		result['release'] = {
			'version': latest.version,
			'download_url': release_download_url,
			'checksum_sha256': latest.checksum_sha256,
			'is_mandatory': bool(latest.is_mandatory),
			'release_notes': latest.release_notes,
		}
	return JsonResponse(result)


@csrf_exempt
@require_POST
def agent_log_event(request):
	body = _read_json_body(request)
	if body is None:
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	agent, err = _authenticate_agent(request, body)
	if err:
		return err

	message = str(body.get('message') or '').strip()
	if not message:
		return JsonResponse({'ok': False, 'error': 'message gerekli.'}, status=400)

	level = str(body.get('level') or 'info').strip().lower()
	job_id = body.get('job_id')
	job = None
	if job_id:
		job = RobotJob.objects.filter(pk=job_id).first()

	_create_agent_event(agent, level, message, job=job, extra=body.get('extra'))
	agent.mark_seen(startup=False)
	agent.save(update_fields=['last_seen_at', 'updated_at'])
	return JsonResponse({'ok': True})


@csrf_exempt
@require_POST
def agent_pull_job(request):
	body = _read_json_body(request)
	if body is None:
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	agent, err = _authenticate_agent(request, body)
	if err:
		return err

	now = timezone.now()
	with transaction.atomic():
		job = RobotJob.objects.select_for_update().filter(
			status='queued'
		).filter(
			Q(target_agent__isnull=True) | Q(target_agent=agent)
		).order_by('-priority', 'created_at').first()

		if not job:
			agent.mark_seen(startup=False)
			if agent.status != 'busy':
				agent.status = 'online'
			agent.save(update_fields=['status', 'last_seen_at', 'updated_at'])
			return JsonResponse({'ok': True, 'job': None})

		if not job.started_at:
			job.started_at = now
		job.target_agent = agent
		job.status = 'dispatched'
		job.last_heartbeat_at = now
		job.lease_expires_at = now + timedelta(minutes=10)
		job.save(update_fields=['target_agent', 'status', 'started_at', 'last_heartbeat_at', 'lease_expires_at', 'updated_at'])

	agent.status = 'busy'
	agent.mark_seen(startup=False)
	agent.save(update_fields=['status', 'last_seen_at', 'updated_at'])

	return JsonResponse({'ok': True, 'job': _serialize_job(job)})


@csrf_exempt
@require_POST
def agent_job_update(request):
	body = _read_json_body(request)
	if body is None:
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	agent, err = _authenticate_agent(request, body)
	if err:
		return err

	job_id = body.get('job_id')
	status_value = str(body.get('status') or '').strip().lower()
	if not job_id or status_value not in {'running', 'succeeded', 'failed', 'canceled'}:
		return JsonResponse({'ok': False, 'error': 'job_id ve geçerli status gerekli.'}, status=400)

	job = RobotJob.objects.filter(pk=job_id, target_agent=agent).first()
	if not job:
		return JsonResponse({'ok': False, 'error': 'İş bulunamadı veya ajan yetkisiz.'}, status=404)

	now = timezone.now()
	job.status = status_value
	job.last_heartbeat_at = now
	if status_value == 'running' and not job.started_at:
		job.started_at = now

	if status_value in {'succeeded', 'failed', 'canceled'}:
		job.finished_at = now
		job.lease_expires_at = None
		agent.status = 'online'
	else:
		agent.status = 'busy'

	result_message = body.get('result_message')
	if result_message is not None:
		job.result_message = str(result_message)
	result_payload = body.get('result_payload')
	if isinstance(result_payload, dict):
		job.result_payload = result_payload

	job.save()
	agent.mark_seen(startup=False)
	agent.save(update_fields=['status', 'last_seen_at', 'updated_at'])

	return JsonResponse({'ok': True})


def robot_agent_status(request):
	rows = []
	for agent in RobotAgent.objects.all().order_by('code'):
		queued = RobotJob.objects.filter(status='queued').filter(
			Q(target_agent__isnull=True) | Q(target_agent=agent)
		).count()
		version_state = 'ok'
		if agent.desired_version and agent.agent_version and agent.desired_version != agent.agent_version:
			version_state = 'outdated'
		elif agent.desired_version and not agent.agent_version:
			version_state = 'unknown'
		rows.append({
			'code': agent.code,
			'name': agent.name,
			'status': agent.status,
			'is_enabled': agent.is_enabled,
			'agent_version': agent.agent_version,
			'desired_version': agent.desired_version,
			'version_state': version_state,
			'last_seen_at': agent.last_seen_at.isoformat() if agent.last_seen_at else None,
			'machine_name': agent.machine_name,
			'ip_address': agent.ip_address,
			'pending_jobs': queued,
		})
	return JsonResponse({'ok': True, 'agents': rows})


def robot_job_list(request):
	status_filter = str(request.GET.get('status') or '').strip().lower()
	agent_code_filter = str(request.GET.get('agent_code') or '').strip()
	search = str(request.GET.get('q') or '').strip()
	limit_param = request.GET.get('limit', 50)
	try:
		limit = max(1, min(200, int(limit_param)))
	except (TypeError, ValueError):
		limit = 50

	jobs_qs = RobotJob.objects.select_related('target_agent', 'sap_process').order_by('-created_at')
	if status_filter:
		jobs_qs = jobs_qs.filter(status=status_filter)
	if agent_code_filter:
		jobs_qs = jobs_qs.filter(target_agent__code=agent_code_filter)
	if search:
		jobs_qs = jobs_qs.filter(
			Q(result_message__icontains=search)
			| Q(command_type__icontains=search)
			| Q(target_agent__code__icontains=search)
			| Q(sap_process__name__icontains=search)
		)

	jobs = jobs_qs[:limit]
	rows = []
	for job in jobs:
		rows.append({
			'job_id': job.pk,
			'status': job.status,
			'command_type': job.command_type,
			'priority': job.priority,
			'target_agent_code': job.target_agent.code if job.target_agent else '',
			'target_agent_name': job.target_agent.name if job.target_agent else '',
			'sap_process_id': job.sap_process_id,
			'sap_process_name': job.sap_process.name if job.sap_process else '',
			'result_message': job.result_message or '',
			'result_payload': job.result_payload if isinstance(job.result_payload, dict) else {},
			'created_at': job.created_at.isoformat() if job.created_at else None,
			'started_at': job.started_at.isoformat() if job.started_at else None,
			'finished_at': job.finished_at.isoformat() if job.finished_at else None,
		})

	return JsonResponse({'ok': True, 'jobs': rows})


def robot_agent_event_list(request):
	agent_code = str(request.GET.get('agent_code') or '').strip()
	if not agent_code:
		return JsonResponse({'ok': False, 'error': 'agent_code gerekli.'}, status=400)
	limit_param = request.GET.get('limit', 50)
	try:
		limit = max(1, min(500, int(limit_param)))
	except (TypeError, ValueError):
		limit = 50
	agent = RobotAgent.objects.filter(code=agent_code).first()
	if not agent:
		return JsonResponse({'ok': False, 'error': 'Ajan bulunamadı.'}, status=404)
	rows = []
	events = RobotAgentEvent.objects.select_related('job').filter(agent=agent).order_by('-created_at')[:limit]
	for ev in events:
		rows.append({
			'id': ev.pk,
			'level': ev.level,
			'message': ev.message,
			'job_id': ev.job_id,
			'created_at': ev.created_at.isoformat() if ev.created_at else None,
		})
	return JsonResponse({'ok': True, 'events': rows})


def robot_release_list(request):
	releases = RobotAgentRelease.objects.all().order_by('-created_at')
	rows = []
	for rel in releases:
		download_path = ''
		if rel.setup_file:
			download_path = f"/api/robot-agent/releases/download/{rel.version}/"
		package_full_path, _ = _package_zip_path(rel.version)
		package_download_path = f"/api/robot-agent/releases/download-package/{rel.version}/" if os.path.exists(package_full_path) else ''
		rows.append({
			'version': rel.version,
			'release_notes': rel.release_notes,
			'download_url': rel.download_url,
			'setup_file': rel.setup_file,
			'download_path': download_path,
			'package_download_path': package_download_path,
			'checksum_sha256': rel.checksum_sha256,
			'install_command': rel.install_command,
			'is_active': rel.is_active,
			'is_mandatory': rel.is_mandatory,
			'created_by': rel.created_by,
			'created_at': rel.created_at.isoformat() if rel.created_at else None,
		})
	return JsonResponse({'ok': True, 'releases': rows})


@require_POST
def robot_release_save(request):
	if request.content_type and request.content_type.startswith('multipart/form-data'):
		body = request.POST
		setup_upload = request.FILES.get('setup_file')
	else:
		body = _read_json_body(request)
		if body is None:
			return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)
		setup_upload = None

	version = str(body.get('version') or '').strip()
	if not version:
		return JsonResponse({'ok': False, 'error': 'version gerekli.'}, status=400)
	release, _ = RobotAgentRelease.objects.get_or_create(version=version)
	release.release_notes = str(body.get('release_notes') or '').strip()
	release.download_url = str(body.get('download_url') or '').strip()
	release.checksum_sha256 = str(body.get('checksum_sha256') or '').strip()
	release.install_command = str(body.get('install_command') or '').strip()
	release.is_active = _as_bool(body.get('is_active', True), True)
	release.is_mandatory = _as_bool(body.get('is_mandatory', False), False)
	created_by = str(body.get('created_by') or '').strip()
	if not created_by and getattr(request, 'user', None) and request.user.is_authenticated:
		created_by = str(request.user.get_username() or '').strip()
	release.created_by = created_by[:120]

	if setup_upload is not None:
		upload_name = str(getattr(setup_upload, 'name', '') or '').lower()
		if not upload_name.endswith('.exe'):
			return JsonResponse({'ok': False, 'error': 'Sadece .exe dosya yüklenebilir.'}, status=400)
		rel_dir = os.path.join(str(settings.MEDIA_ROOT), 'robot_agent', 'releases')
		os.makedirs(rel_dir, exist_ok=True)
		safe_version = re.sub(r'[^0-9A-Za-z._-]+', '_', version).strip('_') or 'release'
		file_name = f'robot-agent-{safe_version}.exe'
		full_path = os.path.join(rel_dir, file_name)
		with open(full_path, 'wb') as fp:
			for chunk in setup_upload.chunks():
				fp.write(chunk)
		release.setup_file = os.path.join('robot_agent', 'releases', file_name).replace('\\', '/')
		if not release.download_url:
			release.download_url = f"/api/robot-agent/releases/download/{version}/"

	release.save()
	return JsonResponse({'ok': True, 'version': release.version, 'download_url': release.download_url, 'setup_file': release.setup_file})


def robot_release_download(request, version):
	version_val = str(version or '').strip()
	release = RobotAgentRelease.objects.filter(version=version_val).first()
	if not release:
		return JsonResponse({'ok': False, 'error': 'Release bulunamadı.'}, status=404)
	if not release.setup_file:
		return JsonResponse({'ok': False, 'error': 'Bu release için setup dosyası yok.'}, status=404)
	full_path = os.path.join(str(settings.MEDIA_ROOT), release.setup_file)
	if not os.path.exists(full_path):
		return JsonResponse({'ok': False, 'error': 'Dosya bulunamadı.'}, status=404)
	with open(full_path, 'rb') as fp:
		content = fp.read()
	resp = HttpResponse(content, content_type='application/octet-stream')
	resp['Content-Disposition'] = f'attachment; filename="{os.path.basename(full_path)}"'
	return resp


def robot_release_download_package(request, version):
	version_val = str(version or '').strip()
	release = RobotAgentRelease.objects.filter(version=version_val).first()
	if not release:
		return JsonResponse({'ok': False, 'error': 'Release bulunamadı.'}, status=404)
	full_path, file_name = _package_zip_path(version_val)
	if not os.path.exists(full_path):
		return JsonResponse({'ok': False, 'error': 'Paket dosyası bulunamadı.'}, status=404)
	with open(full_path, 'rb') as fp:
		content = fp.read()
	resp = HttpResponse(content, content_type='application/zip')
	resp['Content-Disposition'] = f'attachment; filename="{file_name}"'
	return resp


@require_POST
def robot_release_deploy(request):
	body = _read_json_body(request)
	if body is None:
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	version = str(body.get('version') or '').strip()
	if not version:
		return JsonResponse({'ok': False, 'error': 'version gerekli.'}, status=400)
	release = RobotAgentRelease.objects.filter(version=version).first()
	if not release:
		return JsonResponse({'ok': False, 'error': 'Release bulunamadı.'}, status=404)

	scope = str(body.get('scope') or 'outdated').strip().lower()
	agent_codes = body.get('agent_codes') if isinstance(body.get('agent_codes'), list) else []

	agents_qs = RobotAgent.objects.filter(is_enabled=True)
	if scope == 'single':
		if not agent_codes:
			return JsonResponse({'ok': False, 'error': 'single scope için agent_codes gerekli.'}, status=400)
		agents_qs = agents_qs.filter(code__in=[str(x).strip() for x in agent_codes if str(x).strip()])
	elif scope == 'all':
		pass
	else:
		agents_qs = agents_qs.exclude(agent_version=version)

	agents = list(agents_qs)
	if not agents:
		return JsonResponse({'ok': False, 'error': 'Dağıtıma uygun ajan bulunamadı.'}, status=400)

	default_download = release.download_url or f"/api/robot-agent/releases/download/{release.version}/"
	if default_download.startswith('/'):
		default_download = request.build_absolute_uri(default_download)
	default_cmd = release.install_command or 'powershell -NoProfile -ExecutionPolicy Bypass -Command "$u=\"{download_url}\"; $v=\"{version}\"; $dst=\"C:/SaggioRobotAgent/agent-update-$v.exe\"; Invoke-WebRequest -Uri $u -OutFile $dst; Start-Process -FilePath $dst -ArgumentList \"/update\" -WindowStyle Hidden"'
	jobs = []
	for agent in agents:
		command = default_cmd.format(version=release.version, download_url=default_download)
		job = RobotJob.objects.create(
			command_type='run_command',
			target_agent=agent,
			status='queued',
			priority=300,
			requested_by=str(body.get('requested_by') or 'panel')[:120],
			payload={
				'command': command,
				'release_version': release.version,
				'download_url': default_download,
				'kind': 'agent_update',
			},
		)
		_create_agent_event(agent, 'info', f'Panelden update işi kuyruğa alındı: {release.version}', job=job)
		jobs.append(job.pk)

	return JsonResponse({'ok': True, 'jobs_created': len(jobs), 'job_ids': jobs, 'version': release.version})


@require_POST
def robot_build_setup_exe(request):
	body = _read_json_body(request)
	if body is None:
		body = {}

	version = str(body.get('version') or '').strip() or datetime.now().strftime('bootstrap-%Y%m%d-%H%M%S')
	force_rebuild = _as_bool(body.get('force_rebuild', False), False)

	safe_version = re.sub(r'[^0-9A-Za-z._-]+', '_', version).strip('_') or 'bootstrap'
	base_agent_dir = os.path.join(str(settings.BASE_DIR), 'robot_agent')
	service_script = os.path.join(base_agent_dir, 'robot_agent_service.py')
	runtime_script = os.path.join(base_agent_dir, 'agent_runtime.py')
	config_example = os.path.join(base_agent_dir, 'config.example.json')
	for required_path in (service_script, runtime_script, config_example):
		if not os.path.exists(required_path):
			return JsonResponse({'ok': False, 'error': f'Agent build dosyası bulunamadı: {required_path}'}, status=404)

	release_dir = os.path.join(str(settings.MEDIA_ROOT), 'robot_agent', 'releases')
	os.makedirs(release_dir, exist_ok=True)
	output_name = f'SaggioRobotAgentSetup_{safe_version}.exe'
	output_path = os.path.join(release_dir, output_name)
	rel_path = os.path.join('robot_agent', 'releases', output_name).replace('\\', '/')

	if os.path.exists(output_path) and not force_rebuild:
		release, _ = RobotAgentRelease.objects.get_or_create(version=version)
		release.setup_file = rel_path
		release.download_url = f'/api/robot-agent/releases/download/{version}/'
		if not release.install_command:
			release.install_command = 'powershell -NoProfile -ExecutionPolicy Bypass -Command "$u=\"{download_url}\"; $v=\"{version}\"; $dst=\"C:/SaggioRobotAgent/agent-update-$v.exe\"; Invoke-WebRequest -Uri $u -OutFile $dst; Start-Process -FilePath $dst -ArgumentList \"/update\" -WindowStyle Hidden"'
		release.is_active = True
		release.save()
		return JsonResponse({
			'ok': True,
			'built': False,
			'version': version,
			'download_url': release.download_url,
			'message': 'Var olan setup.exe kullanıldı.',
		})

	build_root = os.path.join(str(settings.MEDIA_ROOT), 'robot_agent', '_pyinstaller')
	work_dir = os.path.join(build_root, 'work', safe_version)
	dist_dir = os.path.join(build_root, 'dist', safe_version)
	spec_dir = os.path.join(build_root, 'spec', safe_version)
	src_dir = os.path.join(build_root, 'src', safe_version)
	os.makedirs(work_dir, exist_ok=True)
	os.makedirs(dist_dir, exist_ok=True)
	os.makedirs(spec_dir, exist_ok=True)
	os.makedirs(src_dir, exist_ok=True)

	try:
		pip_cmd = [sys.executable, '-m', 'pip', 'install', 'pyinstaller']
		pip_run = subprocess.run(pip_cmd, capture_output=True, text=True, timeout=420)
		if pip_run.returncode != 0:
			err = (pip_run.stderr or pip_run.stdout or 'pyinstaller kurulamadı')[-2000:]
			return JsonResponse({'ok': False, 'error': f'PyInstaller kurulum hatası: {err}'}, status=500)

		service_dist = os.path.join(dist_dir, 'service')
		service_work = os.path.join(work_dir, 'service')
		service_spec = os.path.join(spec_dir, 'service')
		installer_dist = os.path.join(dist_dir, 'setup')
		installer_work = os.path.join(work_dir, 'setup')
		installer_spec = os.path.join(spec_dir, 'setup')
		for p in (service_dist, service_work, service_spec, installer_dist, installer_work, installer_spec):
			os.makedirs(p, exist_ok=True)

		service_build_cmd = [
			sys.executable,
			'-m',
			'PyInstaller',
			'--noconfirm',
			'--clean',
			'--onefile',
			'--name',
			'SaggioRobotAgentService',
			service_script,
			'--distpath',
			service_dist,
			'--workpath',
			service_work,
			'--specpath',
			service_spec,
		]
		service_run = subprocess.run(service_build_cmd, capture_output=True, text=True, timeout=1800)
		if service_run.returncode != 0:
			err = (service_run.stderr or service_run.stdout or 'service build başarısız')[-4000:]
			return JsonResponse({'ok': False, 'error': f'servis exe üretilemedi: {err}'}, status=500)

		service_exe = os.path.join(service_dist, 'SaggioRobotAgentService.exe')
		if not os.path.exists(service_exe):
			return JsonResponse({'ok': False, 'error': 'Servis exe dosyası bulunamadı.'}, status=500)

		installer_script = os.path.join(src_dir, 'setup_installer_entry.py')
		installer_code = f'''import os
import shutil
import subprocess
import sys

SERVICE_NAME = "SaggioRobotAgent"
BASE_DIR = r"C:/SaggioRobotAgent"
SERVICE_EXE_NAME = "SaggioRobotAgentService.exe"


def _run(cmd):
    return subprocess.run(cmd, capture_output=True, text=True)


def _payload_dir():
    return os.path.join(getattr(sys, "_MEIPASS", os.path.dirname(__file__)), "payload")


def install_or_update():
    os.makedirs(BASE_DIR, exist_ok=True)
    payload = _payload_dir()
    service_src = os.path.join(payload, SERVICE_EXE_NAME)
    config_src = os.path.join(payload, "config.example.json")
    if not os.path.exists(service_src):
        raise RuntimeError("Payload içinde service exe yok.")

    service_dst = os.path.join(BASE_DIR, SERVICE_EXE_NAME)
    config_dst = os.path.join(BASE_DIR, "config.json")
    config_example_dst = os.path.join(BASE_DIR, "config.example.json")

    _run([service_dst, "stop"])
    _run([service_dst, "remove"])

    shutil.copyfile(service_src, service_dst)
    if os.path.exists(config_src):
        shutil.copyfile(config_src, config_example_dst)
    if not os.path.exists(config_dst) and os.path.exists(config_src):
        shutil.copyfile(config_src, config_dst)

    install_run = _run([service_dst, "--startup", "auto", "install"])
    if install_run.returncode != 0:
        err = (install_run.stderr or install_run.stdout or "service install failed")[-2000:]
        raise RuntimeError(err)

    start_run = _run([service_dst, "start"])
    if start_run.returncode != 0:
        err = (start_run.stderr or start_run.stdout or "service start failed")[-2000:]
        raise RuntimeError(err)

    print("Saggio Robot Agent kuruldu/guncellendi. Version: {safe_version}")
    print("Config: C:/SaggioRobotAgent/config.json")


def main():
    try:
        install_or_update()
    except Exception as ex:
        print(f"[ERROR] {{ex}}")
        sys.exit(1)


if __name__ == "__main__":
    main()
'''
		with open(installer_script, 'w', encoding='utf-8') as fp:
			fp.write(installer_code)

		installer_build_cmd = [
			sys.executable,
			'-m',
			'PyInstaller',
			'--noconfirm',
			'--clean',
			'--onefile',
			'--name',
			'SaggioRobotAgentSetup',
			'--add-data',
			f'{service_exe}{os.pathsep}payload',
			'--add-data',
			f'{config_example}{os.pathsep}payload',
			installer_script,
			'--distpath',
			installer_dist,
			'--workpath',
			installer_work,
			'--specpath',
			installer_spec,
		]
		installer_run = subprocess.run(installer_build_cmd, capture_output=True, text=True, timeout=1800)
		if installer_run.returncode != 0:
			err = (installer_run.stderr or installer_run.stdout or 'setup build başarısız')[-4000:]
			return JsonResponse({'ok': False, 'error': f'setup.exe üretilemedi: {err}'}, status=500)

		built_exe = os.path.join(installer_dist, 'SaggioRobotAgentSetup.exe')
		if not os.path.exists(built_exe):
			return JsonResponse({'ok': False, 'error': 'Build tamamlandı ama exe dosyası bulunamadı.'}, status=500)

		shutil.copyfile(built_exe, output_path)
		release, _ = RobotAgentRelease.objects.get_or_create(version=version)
		release.setup_file = rel_path
		release.download_url = f'/api/robot-agent/releases/download/{version}/'
		if not release.install_command:
			release.install_command = 'powershell -NoProfile -ExecutionPolicy Bypass -Command "$u=\"{download_url}\"; $v=\"{version}\"; $dst=\"C:/SaggioRobotAgent/agent-update-$v.exe\"; Invoke-WebRequest -Uri $u -OutFile $dst; Start-Process -FilePath $dst -ArgumentList \"/update\" -WindowStyle Hidden"'
		release.is_active = True
		release.created_by = 'panel-builder'
		release.save()

		return JsonResponse({
			'ok': True,
			'built': True,
			'version': version,
			'download_url': release.download_url,
			'message': 'setup.exe başarıyla üretildi.',
		})
	except subprocess.TimeoutExpired:
		return JsonResponse({'ok': False, 'error': 'setup.exe üretimi zaman aşımına uğradı.'}, status=500)


@require_POST
def robot_build_install_package(request):
	body = _read_json_body(request)
	if body is None:
		body = {}

	version = str(body.get('version') or '').strip() or datetime.now().strftime('bootstrap-%Y%m%d-%H%M%S')
	force_rebuild = _as_bool(body.get('force_rebuild', False), False)
	safe_version = _safe_release_version(version)

	base_agent_dir = os.path.join(str(settings.BASE_DIR), 'robot_agent')
	files_to_include = [
		('agent_runtime.py', 'agent_runtime.py'),
		('robot_agent_service.py', 'robot_agent_service.py'),
		('config.example.json', 'config.example.json'),
	]
	for src_name, _ in files_to_include:
		src_path = os.path.join(base_agent_dir, src_name)
		if not os.path.exists(src_path):
			return JsonResponse({'ok': False, 'error': f'Paket için gerekli dosya yok: {src_path}'}, status=404)

	zip_full_path, zip_file_name = _package_zip_path(version)
	if os.path.exists(zip_full_path) and not force_rebuild:
		release, _ = RobotAgentRelease.objects.get_or_create(version=version)
		release.is_active = True
		if not release.download_url:
			release.download_url = f'/api/robot-agent/releases/download-package/{version}/'
		release.save(update_fields=['is_active', 'download_url'])
		return JsonResponse({
			'ok': True,
			'built': False,
			'version': version,
			'package_url': f'/api/robot-agent/releases/download-package/{version}/',
			'message': 'Var olan kurulum paketi kullanıldı.',
		})

	install_ps1 = '''$ErrorActionPreference = "Stop"
$Base = "C:/SaggioRobotAgent"
$Venv = "$Base/.venv"
$TaskName = "SaggioRobotAgent"

New-Item -ItemType Directory -Force -Path $Base | Out-Null
Copy-Item -Force "$PSScriptRoot/agent_runtime.py" "$Base/agent_runtime.py"
Copy-Item -Force "$PSScriptRoot/robot_agent_service.py" "$Base/robot_agent_service.py"
if (Test-Path "$PSScriptRoot/update_agent.ps1") {
	Copy-Item -Force "$PSScriptRoot/update_agent.ps1" "$Base/update_agent.ps1"
}
if (-not (Test-Path "$Base/config.json") -and (Test-Path "$PSScriptRoot/config.example.json")) {
	Copy-Item -Force "$PSScriptRoot/config.example.json" "$Base/config.json"
}

if (-not (Test-Path $Venv)) {
	$pyLauncher = Get-Command py -ErrorAction SilentlyContinue
	$pythonExe = Get-Command python -ErrorAction SilentlyContinue
	if ($pyLauncher) {
		& py -3.11 -m venv $Venv
		if (-not (Test-Path $Venv)) {
			& py -3 -m venv $Venv
		}
	} elseif ($pythonExe) {
		& python -m venv $Venv
	} else {
		throw "Python bulunamadı."
	}
}

& "$Venv/Scripts/pip.exe" install --upgrade pip
& "$Venv/Scripts/pip.exe" install requests

$PythonExe = "$Venv/Scripts/python.exe"
$ScriptPath = "$Base/agent_runtime.py"
$RunnerPath = "$Base/run_agent.ps1"
if (-not (Test-Path $PythonExe)) {
	throw "Python venv bulunamadı: $PythonExe"
}

$runnerContent = @"
$ErrorActionPreference = "Stop"
$Base = "C:/SaggioRobotAgent"
$env:SAGGIO_AGENT_CONFIG = "$Base/config.json"
& "$PythonExe" -u "$ScriptPath" *>> "$Base/agent_runner.log"
"@
Set-Content -Path $RunnerPath -Value $runnerContent -Encoding UTF8

try {
	Unregister-ScheduledTask -TaskName $TaskName -Confirm:$false -ErrorAction SilentlyContinue | Out-Null
} catch {
}

$action = New-ScheduledTaskAction -Execute "powershell.exe" -Argument ('-NoProfile -ExecutionPolicy Bypass -File "' + $RunnerPath + '"')
$trigger = New-ScheduledTaskTrigger -AtLogOn
$settings = New-ScheduledTaskSettingsSet -StartWhenAvailable -RestartCount 999 -RestartInterval (New-TimeSpan -Minutes 1)
$userId = ([System.Security.Principal.WindowsIdentity]::GetCurrent()).Name
try {
	$principal = New-ScheduledTaskPrincipal -UserId $userId -LogonType Interactive -RunLevel Highest
	Register-ScheduledTask -TaskName $TaskName -Action $action -Trigger $trigger -Settings $settings -Principal $principal -Force | Out-Null
} catch {
	Register-ScheduledTask -TaskName $TaskName -Action $action -Trigger $trigger -Settings $settings -User $userId -RunLevel Highest -Force | Out-Null
}
Start-ScheduledTask -TaskName $TaskName

Write-Host "Saggio Robot Agent kurulumu tamamlandı. (Task Scheduler)"
'''

	update_ps1 = '''param(
	[Parameter(Mandatory=$true)][string]$Version,
	[Parameter(Mandatory=$true)][string]$Url
)
$ErrorActionPreference = "Stop"
$Base = "C:/SaggioRobotAgent"
$Venv = "$Base/.venv"
$TaskName = "SaggioRobotAgent"
$TmpRoot = "$env:TEMP/SaggioRobotAgentUpdate"
$ZipPath = "$TmpRoot/agent-$Version.zip"
$ExtractPath = "$TmpRoot/extract-$Version"

New-Item -ItemType Directory -Force -Path $Base | Out-Null
New-Item -ItemType Directory -Force -Path $TmpRoot | Out-Null
if (Test-Path $ExtractPath) {
	Remove-Item -Recurse -Force $ExtractPath
}

Invoke-WebRequest -Uri $Url -OutFile $ZipPath
Expand-Archive -Path $ZipPath -DestinationPath $ExtractPath -Force

try {
	Stop-ScheduledTask -TaskName $TaskName -ErrorAction SilentlyContinue
} catch {
}

Copy-Item -Force "$ExtractPath/agent_runtime.py" "$Base/agent_runtime.py"
Copy-Item -Force "$ExtractPath/robot_agent_service.py" "$Base/robot_agent_service.py"
Copy-Item -Force "$ExtractPath/update_agent.ps1" "$Base/update_agent.ps1"
if (-not (Test-Path "$Base/config.json") -and (Test-Path "$ExtractPath/config.example.json")) {
	Copy-Item -Force "$ExtractPath/config.example.json" "$Base/config.json"
}

if (-not (Test-Path $Venv)) {
	$pyLauncher = Get-Command py -ErrorAction SilentlyContinue
	$pythonExe = Get-Command python -ErrorAction SilentlyContinue
	if ($pyLauncher) {
		& py -3.11 -m venv $Venv
		if (-not (Test-Path $Venv)) {
			& py -3 -m venv $Venv
		}
	} elseif ($pythonExe) {
		& python -m venv $Venv
	} else {
		throw "Python bulunamadı."
	}
}

& "$Venv/Scripts/pip.exe" install --upgrade pip
& "$Venv/Scripts/pip.exe" install requests

$PythonExe = "$Venv/Scripts/python.exe"
$ScriptPath = "$Base/agent_runtime.py"
$RunnerPath = "$Base/run_agent.ps1"
if (-not (Test-Path $PythonExe)) {
	throw "Python venv bulunamadı: $PythonExe"
}

$runnerContent = @"
$ErrorActionPreference = "Stop"
$Base = "C:/SaggioRobotAgent"
$env:SAGGIO_AGENT_CONFIG = "$Base/config.json"
& "$PythonExe" -u "$ScriptPath" *>> "$Base/agent_runner.log"
"@
Set-Content -Path $RunnerPath -Value $runnerContent -Encoding UTF8

try {
	Unregister-ScheduledTask -TaskName $TaskName -Confirm:$false -ErrorAction SilentlyContinue | Out-Null
} catch {
}

$action = New-ScheduledTaskAction -Execute "powershell.exe" -Argument ('-NoProfile -ExecutionPolicy Bypass -File "' + $RunnerPath + '"')
$trigger = New-ScheduledTaskTrigger -AtLogOn
$settings = New-ScheduledTaskSettingsSet -StartWhenAvailable -RestartCount 999 -RestartInterval (New-TimeSpan -Minutes 1)
$userId = ([System.Security.Principal.WindowsIdentity]::GetCurrent()).Name
try {
	$principal = New-ScheduledTaskPrincipal -UserId $userId -LogonType Interactive -RunLevel Highest
	Register-ScheduledTask -TaskName $TaskName -Action $action -Trigger $trigger -Settings $settings -Principal $principal -Force | Out-Null
} catch {
	Register-ScheduledTask -TaskName $TaskName -Action $action -Trigger $trigger -Settings $settings -User $userId -RunLevel Highest -Force | Out-Null
}
Start-ScheduledTask -TaskName $TaskName

try {
	$cfgPath = "$Base/config.json"
	if (Test-Path $cfgPath) {
		$cfg = Get-Content $cfgPath -Raw | ConvertFrom-Json
		$cfg.agent_version = $Version
		$cfg | ConvertTo-Json -Depth 20 | Set-Content -Path $cfgPath -Encoding UTF8
	}
} catch {
}

Write-Host "Ajan güncelleme tamamlandı. Version: $Version"
'''

	uninstall_ps1 = '''$ErrorActionPreference = "SilentlyContinue"
$Base = "C:/SaggioRobotAgent"
$TaskName = "SaggioRobotAgent"
try {
	Stop-ScheduledTask -TaskName $TaskName -ErrorAction SilentlyContinue
} catch {
}
try {
	Unregister-ScheduledTask -TaskName $TaskName -Confirm:$false -ErrorAction SilentlyContinue | Out-Null
} catch {
}
Write-Host "Saggio Robot Agent gorev kaldirma tamamlandi."
'''

	readme_txt = f'''Saggio Robot Agent Kurumsal Paket\n\nVersion: {safe_version}\n\nCrowdStrike Uyumlu Akis (EXE yok, servis yok):\n1) Zip'i robota çıkart\n2) Yönetici PowerShell aç\n3) install_agent.ps1 çalıştır\n\nKurulum modeli:\n- Ajan, Windows Task Scheduler üzerinden mevcut kullanıcı ile OnLogon tetiklenir.\n- SAP/GUI otomasyonu için robot kullanıcısının oturumu açık olmalıdır.\n\nGerekli ayar:\n- C:/SaggioRobotAgent/config.json içine server_base_url, agent_code, token değerlerini gir.\n\nMerkezi guncelleme:\n- Panelden deploy yapıldığında C:/SaggioRobotAgent/update_agent.ps1 çağrılır ve zip paketi indirip ajanı yeniler.\n'''

	tmp_root = os.path.join(str(settings.MEDIA_ROOT), 'robot_agent', '_tmp_pkg', safe_version)
	os.makedirs(tmp_root, exist_ok=True)
	for src_name, out_name in files_to_include:
		shutil.copyfile(os.path.join(base_agent_dir, src_name), os.path.join(tmp_root, out_name))
	with open(os.path.join(tmp_root, 'install_agent.ps1'), 'w', encoding='utf-8') as fp:
		fp.write(install_ps1)
	with open(os.path.join(tmp_root, 'update_agent.ps1'), 'w', encoding='utf-8') as fp:
		fp.write(update_ps1)
	with open(os.path.join(tmp_root, 'uninstall_agent.ps1'), 'w', encoding='utf-8') as fp:
		fp.write(uninstall_ps1)
	with open(os.path.join(tmp_root, 'README.txt'), 'w', encoding='utf-8') as fp:
		fp.write(readme_txt)

	os.makedirs(os.path.dirname(zip_full_path), exist_ok=True)
	with zipfile.ZipFile(zip_full_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
		for item in os.listdir(tmp_root):
			item_path = os.path.join(tmp_root, item)
			if os.path.isfile(item_path):
				zf.write(item_path, arcname=item)

	release, _ = RobotAgentRelease.objects.get_or_create(version=version)
	release.is_active = True
	release.download_url = f'/api/robot-agent/releases/download-package/{version}/'
	release.install_command = 'powershell -NoProfile -ExecutionPolicy Bypass -File C:/SaggioRobotAgent/update_agent.ps1 -Version "{version}" -Url "{download_url}"'
	if not release.release_notes:
		release.release_notes = 'Kurumsal kurulum paketi (zip + powershell) üretildi.'
	release.save()

	return JsonResponse({
		'ok': True,
		'built': True,
		'version': version,
		'package_url': f'/api/robot-agent/releases/download-package/{version}/',
		'file_name': zip_file_name,
		'message': 'Kurumsal kurulum paketi hazır.',
	})


@require_POST
def robot_set_desired_version(request):
	body = _read_json_body(request)
	if body is None:
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)
	version = str(body.get('version') or '').strip()
	if not version:
		return JsonResponse({'ok': False, 'error': 'version gerekli.'}, status=400)
	if not RobotAgentRelease.objects.filter(version=version).exists():
		return JsonResponse({'ok': False, 'error': 'Release bulunamadı.'}, status=404)
	agent_code = str(body.get('agent_code') or '').strip()
	if agent_code:
		agent = RobotAgent.objects.filter(code=agent_code).first()
		if not agent:
			return JsonResponse({'ok': False, 'error': 'Ajan bulunamadı.'}, status=404)
		agent.desired_version = version
		agent.save(update_fields=['desired_version', 'updated_at'])
		return JsonResponse({'ok': True, 'updated': 1, 'scope': 'single'})
	updated = RobotAgent.objects.update(desired_version=version, updated_at=timezone.now())
	return JsonResponse({'ok': True, 'updated': int(updated), 'scope': 'all'})


@require_POST
def robot_agent_upsert(request):
	body = _read_json_body(request)
	if body is None:
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)
	code = str(body.get('code') or '').strip()
	name = str(body.get('name') or '').strip()
	if not code or not name:
		return JsonResponse({'ok': False, 'error': 'code ve name gerekli.'}, status=400)
	agent, created = RobotAgent.objects.get_or_create(
		code=code,
		defaults={
			'name': name,
			'token_hash': '',
			'is_enabled': True,
			'status': 'offline',
		},
	)
	agent.name = name[:180]
	agent.is_enabled = bool(body.get('is_enabled', True))
	desired_version = str(body.get('desired_version') or '').strip()
	if desired_version:
		agent.desired_version = desired_version[:40]
	token = str(body.get('token') or '').strip()
	if token:
		agent.set_token(token)
	agent.save()

	if created:
		_create_agent_event(agent, 'info', 'Ajan panelden oluşturuldu.')
	elif token:
		_create_agent_event(agent, 'warning', 'Ajan token değeri panelden yenilendi.')

	return JsonResponse({'ok': True, 'created': created, 'code': agent.code})


@require_POST
def robot_cancel_job(request):
	body = _read_json_body(request)
	if body is None:
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	job_id = body.get('job_id')
	if not job_id:
		return JsonResponse({'ok': False, 'error': 'job_id gerekli.'}, status=400)

	job = RobotJob.objects.filter(pk=job_id).first()
	if not job:
		return JsonResponse({'ok': False, 'error': 'İş bulunamadı.'}, status=404)

	if job.status in {'succeeded', 'failed', 'canceled'}:
		return JsonResponse({'ok': False, 'error': 'Tamamlanmış iş iptal edilemez.'}, status=400)

	job.status = 'canceled'
	job.finished_at = timezone.now()
	job.result_message = str(body.get('reason') or 'Kullanıcı tarafından iptal edildi.').strip()
	job.lease_expires_at = None
	job.save(update_fields=['status', 'finished_at', 'result_message', 'lease_expires_at', 'updated_at'])

	if job.target_agent_id:
		agent = job.target_agent
		active_count = RobotJob.objects.filter(target_agent=agent, status__in=['dispatched', 'running']).exclude(pk=job.pk).count()
		if active_count == 0:
			agent.status = 'online'
			agent.mark_seen(startup=False)
			agent.save(update_fields=['status', 'last_seen_at', 'updated_at'])

	return JsonResponse({'ok': True, 'job_id': job.pk, 'status': job.status})


@require_POST
def robot_dispatch_job(request):
	body = _read_json_body(request)
	if body is None:
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	command_type = str(body.get('command_type') or 'run_sap_process').strip()
	if command_type not in {'run_sap_process', 'run_command'}:
		return JsonResponse({'ok': False, 'error': 'Geçersiz command_type.'}, status=400)

	target_agent = None
	target_agent_code = str(body.get('target_agent_code') or '').strip()
	if target_agent_code:
		target_agent = RobotAgent.objects.filter(code=target_agent_code, is_enabled=True).first()
		if not target_agent:
			return JsonResponse({'ok': False, 'error': 'Hedef ajan bulunamadı veya pasif.'}, status=404)

	sap_process = None
	sap_process_id = body.get('sap_process_id')
	if command_type == 'run_sap_process':
		if not sap_process_id:
			return JsonResponse({'ok': False, 'error': 'run_sap_process için sap_process_id gerekli.'}, status=400)
		sap_process = SapProcess.objects.filter(pk=sap_process_id).first()
		if not sap_process:
			return JsonResponse({'ok': False, 'error': 'SAP süreç bulunamadı.'}, status=404)

	payload = body.get('payload') if isinstance(body.get('payload'), dict) else {}
	try:
		priority = int(body.get('priority', 100))
	except (TypeError, ValueError):
		priority = 100
	requested_by = str(body.get('requested_by') or '').strip()
	if not requested_by and getattr(request, 'user', None) and request.user.is_authenticated:
		requested_by = str(request.user.get_username() or '').strip()

	job = RobotJob.objects.create(
		command_type=command_type,
		sap_process=sap_process,
		target_agent=target_agent,
		payload=payload,
		priority=priority,
		status='queued',
		requested_by=requested_by[:120],
	)

	return JsonResponse({
		'ok': True,
		'job': {
			'job_id': job.pk,
			'status': job.status,
			'target_agent_code': target_agent.code if target_agent else '',
			'command_type': job.command_type,
			'sap_process_id': job.sap_process_id,
		}
	})
