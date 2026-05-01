"""Excel cursor + value resolution helpers for SAP runtime.

Originally lived in ``core/views.py``.
"""
from __future__ import annotations

import os
import re as _re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

try:
	from openpyxl import load_workbook
except ImportError:  # openpyxl optional
	load_workbook = None  # type: ignore

from ..utils.parsing import _as_bool


def _build_excel_cursor_key(cfg):
	if not isinstance(cfg, dict):
		return ''
	excel_path_raw = str(cfg.get('excel_file_path', '') or '').strip()
	# Windows'ta / ve \\ farkı aynı dosyaya işaret edebilir; cursor key'de tekilleştir.
	if excel_path_raw:
		try:
			excel_path = os.path.normpath(excel_path_raw).strip().lower()
		except Exception:
			excel_path = excel_path_raw.lower()
	else:
		excel_path = ''
	sheet_name = str(cfg.get('excel_sheet_name', '') or '').strip().lower()
	try:
		header_row = max(1, int(cfg.get('excel_header_row') or 1))
	except (TypeError, ValueError):
		header_row = 1
	if not excel_path:
		return ''
	return f'{excel_path}|{sheet_name}|{header_row}'


def _get_excel_cursor_index(runtime_state, cfg):
	if not isinstance(runtime_state, dict):
		return 0
	key = _build_excel_cursor_key(cfg)
	if not key:
		return 0
	cursors = runtime_state.get('excel_cursors')
	if not isinstance(cursors, dict):
		return 0
	try:
		return max(0, int(cursors.get(key, 0) or 0))
	except Exception:
		return 0


def _set_excel_cursor_index(runtime_state, cfg, index_value):
	if not isinstance(runtime_state, dict):
		return
	key = _build_excel_cursor_key(cfg)
	if not key:
		return
	cursors = runtime_state.get('excel_cursors')
	if not isinstance(cursors, dict):
		cursors = {}
		runtime_state['excel_cursors'] = cursors
	try:
		cursors[key] = max(0, int(index_value or 0))
	except Exception:
		cursors[key] = 0


def _reset_excel_cursors(runtime_state):
	if not isinstance(runtime_state, dict):
		return
	runtime_state['excel_cursors'] = {}


def _get_excel_row_index(cfg, runtime_state):
	row_source = str(cfg.get('excel_row_index_source', '') or '').strip().casefold()
	if not row_source:
		# Excel Satır Sonraki adımı kullanıldıysa (excel_cursors'ta bu dosya için kayıt varsa)
		# otomatik olarak excel_loop moduna geç
		_cursor_key = _build_excel_cursor_key(cfg)
		_cursors = runtime_state.get('excel_cursors') if isinstance(runtime_state, dict) else None
		if _cursor_key and isinstance(_cursors, dict) and _cursor_key in _cursors:
			row_source = 'excel_loop'
		else:
			row_source = 'outer_loop' if _as_bool(cfg.get('excel_use_loop_index', True), True) else 'fixed'
	if row_source == 'excel_loop':
		return _get_excel_cursor_index(runtime_state, cfg)
	if row_source == 'fixed':
		return 0
	return int(runtime_state.get('loop_index', 0) or 0) if isinstance(runtime_state, dict) else 0


def _get_excel_total_data_rows(excel_path, sheet_name, header_row, runtime_state=None, cfg=None):
	cache = runtime_state.get('excel_row_count_cache') if isinstance(runtime_state, dict) else None
	cache_key = None
	if isinstance(runtime_state, dict):
		cache_key = _build_excel_cursor_key(cfg or {
			'excel_file_path': excel_path,
			'excel_sheet_name': sheet_name,
			'excel_header_row': header_row,
		})
		if not isinstance(cache, dict):
			cache = {}
			runtime_state['excel_row_count_cache'] = cache
		if cache_key and cache_key in cache:
			return True, int(cache.get(cache_key, 0) or 0), ''

	if load_workbook is None:
		return False, 0, 'Excel okuma için openpyxl kurulu değil.'
	try:
		wb = load_workbook(excel_path, data_only=True, read_only=True)
	except Exception as ex:
		return False, 0, f'Excel açılamadı: {ex}'

	try:
		if sheet_name:
			if sheet_name not in wb.sheetnames:
				return False, 0, f'Sayfa bulunamadı: {sheet_name}'
			ws = wb[sheet_name]
		else:
			ws = wb[wb.sheetnames[0]]
		max_row = int(getattr(ws, 'max_row', 0) or 0)
		count = max(0, max_row - int(header_row))
		if count <= 0:
			# read_only çalışma modunda max_row güvenilmez olabiliyor; veri satırlarını fiilen say.
			count = 0
			for row in ws.iter_rows(min_row=int(header_row) + 1, values_only=True):
				if any((cell is not None and str(cell).strip() != '') for cell in (row or ())):
					count += 1
		if cache_key and isinstance(cache, dict):
			cache[cache_key] = count
		return True, count, ''
	finally:
		try:
			wb.close()
		except Exception:
			pass


def _excel_column_to_index(column_ref):
	text = str(column_ref or '').strip().upper()
	if not text:
		return None
	if text.isdigit():
		idx = int(text)
		return idx if idx > 0 else None
	if not _re.fullmatch(r'[A-Z]+', text):
		return None
	idx = 0
	for ch in text:
		idx = (idx * 26) + (ord(ch) - 64)
	return idx if idx > 0 else None


def _normalize_excel_value(value):
	if value is None:
		return ''

	def _format_decimal_tr(num, *, force_group_for_decimal=True):
		try:
			dec = Decimal(str(num))
		except (InvalidOperation, ValueError, TypeError):
			return str(num or '').strip()
		sign = '-' if dec < 0 else ''
		dec_abs = abs(dec)
		txt = format(dec_abs, 'f')
		if '.' in txt:
			txt = txt.rstrip('0').rstrip('.')
		if not txt:
			return '0'
		if '.' in txt:
			int_part, frac_part = txt.split('.', 1)
		else:
			int_part, frac_part = txt, ''
		if frac_part:
			if force_group_for_decimal:
				int_grouped = f"{int(int_part):,}".replace(',', '.')
			else:
				int_grouped = int_part
			return f"{sign}{int_grouped},{frac_part}"
		return f"{sign}{int_part}"

	# Tarih/datetime değeri TR formatına çevir.
	if isinstance(value, datetime):
		return value.strftime('%d.%m.%Y')
	if isinstance(value, date):
		return value.strftime('%d.%m.%Y')

	# Sayısal değerlerde ondalık ayracı korunur; Türk biçimine çevrilir.
	if isinstance(value, (int, float)):
		if isinstance(value, float) and value.is_integer():
			return str(int(value))
		return _format_decimal_tr(value, force_group_for_decimal=True)

	text = str(value or '').strip()
	if not text:
		return ''

	date_formats = [
		'%d.%m.%Y', '%d/%m/%Y', '%d-%m-%Y',
		'%Y-%m-%d', '%Y/%m/%d',
		'%d.%m.%Y %H:%M:%S', '%d/%m/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S',
	]
	for fmt in date_formats:
		try:
			return datetime.strptime(text, fmt).strftime('%d.%m.%Y')
		except Exception:
			pass

	# Harf içermeyen değerlerde Türk sayı formatını koru (ondalık ayracı: ,).
	if _re.search(r'\d', text) and not _re.search(r'[A-Za-zÇĞİÖŞÜçğıöşü]', text):
		num_raw = text.replace('\u00a0', '').replace(' ', '')
		sign = ''
		if num_raw.startswith(('+', '-')):
			sign = num_raw[0]
			num_raw = num_raw[1:]
		if num_raw.isdigit():
			return f'{sign}{num_raw}'
		if ('.' in num_raw) or (',' in num_raw):
			last_dot = num_raw.rfind('.')
			last_comma = num_raw.rfind(',')
			if last_dot >= 0 and last_comma >= 0:
				dec_sep = '.' if last_dot > last_comma else ','
			elif last_comma >= 0:
				dec_sep = ','
			else:
				dec_sep = '.'
			if dec_sep == ',':
				canon = num_raw.replace('.', '').replace(',', '.')
			else:
				canon = num_raw.replace(',', '')
			if _re.fullmatch(r'\d+(?:\.\d+)?', canon):
				return _format_decimal_tr(f'{sign}{canon}', force_group_for_decimal=True)
		return _re.sub(r'\D+', '', f'{sign}{num_raw}')

	return text


def _resolve_fill_input_excel_value(cfg, runtime_state):
	if load_workbook is None:
		return False, '', 'Excel okuma için openpyxl kurulu değil.'

	excel_path = str(cfg.get('excel_file_path', '') or '').strip()
	if not excel_path:
		return False, '', 'Excel dosya yolu boÅŸ.'
	if not os.path.isfile(excel_path):
		return False, '', f'Excel dosyası bulunamadı: {excel_path}'

	column_ref = str(cfg.get('excel_column', '') or '').strip()
	if not column_ref:
		return False, '', 'Excel sütunu boş.'

	try:
		header_row = max(1, int(cfg.get('excel_header_row') or 1))
	except (TypeError, ValueError):
		header_row = 1
	try:
		row_offset = int(cfg.get('excel_row_offset') or 0)
	except (TypeError, ValueError):
		row_offset = 0

	try:
		wb = load_workbook(excel_path, data_only=True, read_only=True)
	except Exception as ex:
		return False, '', f'Excel açılamadı: {ex}'

	try:
		sheet_name = str(cfg.get('excel_sheet_name', '') or '').strip()
		if sheet_name:
			if sheet_name not in wb.sheetnames:
				return False, '', f'Sayfa bulunamadı: {sheet_name}'
			ws = wb[sheet_name]
		else:
			ws = wb[wb.sheetnames[0]]

		col_idx = None
		headers = [str(c.value or '').strip() for c in ws[header_row]]
		needle = column_ref.casefold()
		for idx, head in enumerate(headers, start=1):
			if head.casefold() == needle:
				col_idx = idx
				break
		if col_idx is None:
			col_idx = _excel_column_to_index(column_ref)
		if col_idx is None:
			return False, '', f'Sütun bulunamadı: {column_ref}'

		loop_idx = _get_excel_row_index(cfg, runtime_state)
		data_row_start = header_row + 1
		target_row = data_row_start + loop_idx + row_offset
		if target_row < 1:
			target_row = 1

		cell_value = ws.cell(row=target_row, column=col_idx).value
		resolved = _normalize_excel_value(cell_value)
		if resolved == '':
			return False, '', f'Excel hücresi boş: {ws.title}!R{target_row}C{col_idx}'
		msg = f'Excel değer alındı: {ws.title}!R{target_row}C{col_idx} -> {resolved}'
		return True, resolved, msg
	finally:
		try:
			wb.close()
		except Exception:
			pass


