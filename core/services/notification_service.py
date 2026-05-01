"""Telegram / Mail / FTP test + notification helpers.

Originally lived in ``core/views.py``.
"""
from __future__ import annotations

import os
import json
import socket
import smtplib
import ftplib
import shutil
import subprocess
import tempfile
import uuid
from datetime import datetime
from email.mime.text import MIMEText
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

try:
	import paramiko
except Exception:  # pragma: no cover
	paramiko = None

from ..models import TelegramBot, TelegramGroup, MailAccount
from ..utils.parsing import _extract_telegram_http_error


def _send_telegram_group_test(group, payload=None):
	payload = payload or {}
	bot = group.default_bot
	if not bot or not bot.is_active:
		return False, 'Bu grup icin aktif varsayilan bot yok.'

	token = bot.get_bot_token()
	if not token:
		return False, 'Bot token cozulmedi veya bos.'

	custom_text = str(payload.get('test_message') or '').strip()
	text = custom_text or f"Saggio RPA test mesaji\nGrup: {group.name}\nBot: {bot.name}"
	ok, detail = _send_telegram_message(bot, group.chat_id, text)
	if ok:
		return True, 'Test mesaji gonderildi.'
	return False, f'Telegram hatasi: {detail}'


def _send_mail_test(account, payload=None):
	payload = payload or {}
	password = account.get_smtp_password()
	if not password:
		return False, 'SMTP şifresi çözülemedi veya boş.'

	from_display = account.from_name or account.name
	from_value = f'{from_display} <{account.email}>'
	to_value = str(payload.get('test_to') or '').strip() or account.email
	subject = str(payload.get('test_subject') or '').strip() or 'Saggio RPA - Test Maili'
	body = str(payload.get('test_body') or '').strip() or 'Bu e-posta Saggio RPA tarafindan test amacli gonderildi.'

	msg = MIMEText(body, 'plain', 'utf-8')
	msg['Subject'] = subject
	msg['From'] = from_value
	msg['To'] = to_value

	try:
		if account.use_ssl:
			server = smtplib.SMTP_SSL(account.smtp_host, account.smtp_port, timeout=15)
		else:
			server = smtplib.SMTP(account.smtp_host, account.smtp_port, timeout=15)
			if account.use_tls:
				server.starttls()

		with server:
			server.login(account.smtp_username, password)
			server.sendmail(account.email, [to_value], msg.as_string())
		return True, f'Test maili gonderildi: {to_value}'
	except Exception as e:
		return False, f'SMTP test hatası: {e}'


def _send_ftp_test(account, payload=None):
	payload = payload or {}
	password = account.get_password()
	if not password:
		return False, 'FTP sifresi çözülemedi veya boş.'

	remote_path = str(payload.get('test_path') or '').strip() or account.remote_base_path or '.'
	protocol = str(account.protocol or 'sftp').lower()

	try:
		if protocol == 'sftp':
			if paramiko is None:
				return False, 'SFTP testi için paramiko kurulu değil.'
			transport = paramiko.Transport((account.host, int(account.port or 22)))
			transport.connect(username=account.username, password=password)
			sftp = paramiko.SFTPClient.from_transport(transport)
			sftp.listdir(remote_path)
			sftp.close()
			transport.close()
			return True, f'SFTP bağlantısı başarılı. Yol erişimi: {remote_path}'

		if protocol == 'ftps':
			ftp = ftplib.FTP_TLS()
			ftp.connect(account.host, int(account.port or 21), timeout=15)
			ftp.login(account.username, password)
			ftp.prot_p()
			ftp.cwd(remote_path)
			ftp.quit()
			return True, f'FTPS bağlantısı başarılı. Yol erişimi: {remote_path}'

		ftp = ftplib.FTP()
		ftp.connect(account.host, int(account.port or 21), timeout=15)
		ftp.login(account.username, password)
		ftp.cwd(remote_path)
		ftp.quit()
		return True, f'FTP bağlantısı başarılı. Yol erişimi: {remote_path}'
	except (socket.timeout, ConnectionRefusedError) as e:
		return False, f'FTP bağlantı hatası: {e}'
	except Exception as e:
		return False, f'FTP test hatası: {e}'


def _send_telegram_message(bot, chat_id, text):
	if not bot or not bot.is_active:
		return False, 'Telegram bot aktif degil veya secilmedi.'
	if not chat_id:
		return False, 'Telegram chat id secilmedi.'
	token = bot.get_bot_token()
	if not token:
		return False, 'Telegram token cozulmedi.'

	msg_text = str(text or '')
	base_data = {'chat_id': str(chat_id), 'text': msg_text}
	data = dict(base_data)
	if bot.default_parse_mode:
		data['parse_mode'] = bot.default_parse_mode

	def _call_send_message(payload):
		req = Request(
			f'https://api.telegram.org/bot{token}/sendMessage',
			data=urlencode(payload).encode('utf-8'),
			method='POST',
		)
		with urlopen(req, timeout=10) as resp:
			return json.loads(resp.read().decode('utf-8') or '{}')

	try:
		payload = _call_send_message(data)
		if payload.get('ok'):
			return True, 'telegram_ok'
		desc = str(payload.get('description', 'telegram_api_error') or '')
		# parse_mode kaynaklı hata olursa sade metin ile yeniden dene.
		if data.get('parse_mode') and ('parse entities' in desc.casefold() or 'can\'t parse' in desc.casefold()):
			payload2 = _call_send_message(base_data)
			if payload2.get('ok'):
				return True, 'telegram_ok_plain'
			desc = str(payload2.get('description', 'telegram_api_error') or '')

		# Uzun mesaj veya genel API hatalarında güvenli fallback: düz metin + kısaltılmış içerik.
		if len(msg_text) > 3900:
			truncated_data = {'chat_id': str(chat_id), 'text': (msg_text[:3900] + '...')}
			payload3 = _call_send_message(truncated_data)
			if payload3.get('ok'):
				return True, 'telegram_ok_truncated'
			desc = str(payload3.get('description', desc or 'telegram_api_error') or '')
		return False, desc or 'telegram_api_error'
	except HTTPError as e:
		return False, _extract_telegram_http_error(e)
	except URLError as e:
		return False, f'Baglanti hatasi: {e.reason}'
	except Exception as e:
		return False, str(e)


def _send_telegram_voice_message(bot, chat_id, text):
	"""Windows SAPI ile ses üretip Telegram'a gerçek voice note olarak gönderir."""
	if not bot or not bot.is_active:
		return False, 'Telegram bot aktif degil veya secilmedi.'
	if not chat_id:
		return False, 'Telegram chat id secilmedi.'
	token = bot.get_bot_token()
	if not token:
		return False, 'Telegram token cozulmedi.'

	voice_text = str(text or '').strip()
	if not voice_text:
		return False, 'Sesli mesaj metni boş.'

	tmp_wav_path = None
	tmp_ogg_path = None
	tmp_mp3_path = None
	co_initialized = False
	errors = []

	def _send_voice_file(path_value, filename_value, content_type_value):
		with open(path_value, 'rb') as fp:
			file_data = fp.read()

		boundary = f'----SaggioBoundary{uuid.uuid4().hex}'
		parts = []

		def _add_text(name, value):
			parts.append(f'--{boundary}'.encode('utf-8'))
			parts.append(f'Content-Disposition: form-data; name="{name}"'.encode('utf-8'))
			parts.append(b'')
			parts.append(str(value).encode('utf-8'))

		_add_text('chat_id', str(chat_id))
		_add_text('caption', 'Saggio RPA sesli bildirim')

		parts.append(f'--{boundary}'.encode('utf-8'))
		parts.append(f'Content-Disposition: form-data; name="voice"; filename="{filename_value}"'.encode('utf-8'))
		parts.append(f'Content-Type: {content_type_value}'.encode('utf-8'))
		parts.append(b'')
		parts.append(file_data)

		parts.append(f'--{boundary}--'.encode('utf-8'))
		parts.append(b'')
		body = b'\r\n'.join(parts)

		req = Request(
			f'https://api.telegram.org/bot{token}/sendVoice',
			data=body,
			method='POST',
			headers={'Content-Type': f'multipart/form-data; boundary={boundary}'},
		)
		with urlopen(req, timeout=20) as resp:
			payload = json.loads(resp.read().decode('utf-8') or '{}')
			if payload.get('ok'):
				return True, 'telegram_voice_ok'
			return False, payload.get('description', 'telegram_voice_api_error')

	def _send_audio_file(path_value, filename_value, content_type_value):
		with open(path_value, 'rb') as fp:
			file_data = fp.read()

		boundary = f'----SaggioBoundary{uuid.uuid4().hex}'
		parts = []

		def _add_text(name, value):
			parts.append(f'--{boundary}'.encode('utf-8'))
			parts.append(f'Content-Disposition: form-data; name="{name}"'.encode('utf-8'))
			parts.append(b'')
			parts.append(str(value).encode('utf-8'))

		_add_text('chat_id', str(chat_id))
		_add_text('caption', 'Saggio RPA sesli bildirim')

		parts.append(f'--{boundary}'.encode('utf-8'))
		parts.append(f'Content-Disposition: form-data; name="audio"; filename="{filename_value}"'.encode('utf-8'))
		parts.append(f'Content-Type: {content_type_value}'.encode('utf-8'))
		parts.append(b'')
		parts.append(file_data)

		parts.append(f'--{boundary}--'.encode('utf-8'))
		parts.append(b'')
		body = b'\r\n'.join(parts)

		req = Request(
			f'https://api.telegram.org/bot{token}/sendAudio',
			data=body,
			method='POST',
			headers={'Content-Type': f'multipart/form-data; boundary={boundary}'},
		)
		with urlopen(req, timeout=20) as resp:
			payload = json.loads(resp.read().decode('utf-8') or '{}')
			if payload.get('ok'):
				return True, 'telegram_audio_ok'
			return False, payload.get('description', 'telegram_audio_api_error')
	try:
		try:
			import pythoncom
			import win32com.client

			pythoncom.CoInitialize()
			co_initialized = True
			with tempfile.NamedTemporaryFile(suffix='.wav', delete=False) as tmpf:
				tmp_wav_path = tmpf.name
			with tempfile.NamedTemporaryFile(suffix='.ogg', delete=False) as tmpf:
				tmp_ogg_path = tmpf.name

			voice = win32com.client.Dispatch('SAPI.SpVoice')
			# Türkçe sese öncelik ver (kuruluysa)
			try:
				voices = voice.GetVoices()
				selected_voice = None
				count = int(getattr(voices, 'Count', 0) or 0)
				for i in range(count):
					try:
						candidate = voices.Item(i)
						desc = str(candidate.GetDescription() or '').casefold()
						lang = str(candidate.GetAttribute('Language') or '').casefold()
						if 'turkish' in desc or 'türk' in desc or '041f' in lang:
							selected_voice = candidate
							break
					except Exception:
						continue
				if selected_voice is not None:
					voice.Voice = selected_voice
			except Exception:
				pass
			try:
				voice.Rate = -1
			except Exception:
				pass
			stream = win32com.client.Dispatch('SAPI.SpFileStream')
			# 3 = SSFMCreateForWrite
			stream.Open(tmp_wav_path, 3, False)
			voice.AudioOutputStream = stream
			voice.Speak(voice_text)
			stream.Close()

			ffmpeg_path = shutil.which('ffmpeg')
			if ffmpeg_path:
				convert_cmd = [
					ffmpeg_path,
					'-y',
					'-i',
					tmp_wav_path,
					'-c:a',
					'libopus',
					'-b:a',
					'24k',
					tmp_ogg_path,
				]
				convert_result = subprocess.run(convert_cmd, capture_output=True, text=True, encoding='utf-8', errors='replace')
				if convert_result.returncode == 0 and os.path.exists(tmp_ogg_path) and os.path.getsize(tmp_ogg_path) > 0:
					ok, detail = _send_voice_file(tmp_ogg_path, 'saggio_notification.ogg', 'audio/ogg')
					if ok:
						return True, detail
					errors.append(f'ogg_send_failed: {detail}')
				else:
					err = (convert_result.stderr or convert_result.stdout or 'ffmpeg convert hatasi').strip()
					errors.append(f'ffmpeg_convert_failed: {err}')
			else:
				errors.append('ffmpeg_not_found')
		except Exception as sapi_ex:
			errors.append(f'sapi_flow_failed: {sapi_ex}')

		# SAPI/ffmpeg başarısız olursa gTTS ile MP3 üretip sendVoice dene.
		try:
			from gtts import gTTS
			with tempfile.NamedTemporaryFile(suffix='.mp3', delete=False) as tmpf:
				tmp_mp3_path = tmpf.name
			clean_text = voice_text.replace('*', '').replace('_', '')
			gTTS(text=clean_text, lang='tr', slow=False).save(tmp_mp3_path)
			ok, detail = _send_voice_file(tmp_mp3_path, 'saggio_notification.mp3', 'audio/mpeg')
			if ok:
				return True, detail
			errors.append(f'gtts_send_failed: {detail}')
			# Voice endpoint reddederse normal audio olarak düşür.
			a_ok, a_msg = _send_audio_file(tmp_mp3_path, 'saggio_notification.mp3', 'audio/mpeg')
			if a_ok:
				return True, a_msg
			errors.append(f'gtts_audio_fallback_failed: {a_msg}')
		except Exception as ge:
			errors.append(f'gtts_fallback_failed: {ge}')

		return False, 'Voice note gonderilemedi: ' + ' | '.join(errors)
	except Exception as e:
		return False, str(e)
	finally:
		if co_initialized:
			try:
				pythoncom.CoUninitialize()
			except Exception:
				pass
		if tmp_wav_path and os.path.exists(tmp_wav_path):
			try:
				os.remove(tmp_wav_path)
			except Exception:
				pass
		if tmp_ogg_path and os.path.exists(tmp_ogg_path):
			try:
				os.remove(tmp_ogg_path)
			except Exception:
				pass
		if tmp_mp3_path and os.path.exists(tmp_mp3_path):
			try:
				os.remove(tmp_mp3_path)
			except Exception:
				pass




def _send_mail_message(account, to_value, subject, body, attachment_path=None):
	if not account or not account.is_active:
		return False, 'Mail hesabi aktif degil veya secilmedi.'
	password = account.get_smtp_password()
	if not password:
		return False, 'SMTP şifresi çözülemedi.'

	from_display = account.from_name or account.name
	from_value = f'{from_display} <{account.email}>'

	if attachment_path:
		from email.mime.multipart import MIMEMultipart
		from email.mime.base import MIMEBase
		from email import encoders
		msg = MIMEMultipart()
		msg.attach(MIMEText(body, 'plain', 'utf-8'))
		try:
			import os as _os
			with open(attachment_path, 'rb') as f:
				part = MIMEBase('application', 'octet-stream')
				part.set_payload(f.read())
			encoders.encode_base64(part)
			part.add_header('Content-Disposition', f'attachment; filename="{_os.path.basename(attachment_path)}"')
			msg.attach(part)
		except Exception as e:
			return False, f'Ek dosya okunamadı: {e}'
	else:
		msg = MIMEText(body, 'plain', 'utf-8')

	msg['Subject'] = subject
	msg['From'] = from_value
	msg['To'] = to_value

	try:
		if account.use_ssl:
			server = smtplib.SMTP_SSL(account.smtp_host, account.smtp_port, timeout=15)
		else:
			server = smtplib.SMTP(account.smtp_host, account.smtp_port, timeout=15)
			if account.use_tls:
				server.starttls()

		with server:
			server.login(account.smtp_username, password)
			server.sendmail(account.email, [to_value], msg.as_string())
		return True, 'mail_ok'
	except Exception as e:
		return False, str(e)


def _generate_row_report_xlsx(row_results, output_path=None):
	"""row_results listesinden xlsx rapor dosyası üretir. Dosya yolu döndürür."""
	import tempfile
	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill, Alignment

	wb = Workbook()
	ws = wb.active
	ws.title = 'Rapor'

	headers = ['#', 'Cari / Değer', 'Excel Satır No', 'Durum', 'Neden / Notlar', 'Zaman']
	header_font = Font(bold=True, color='FFFFFF')
	header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
	for col_idx, h in enumerate(headers, 1):
		cell = ws.cell(row=1, column=col_idx, value=h)
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

	status_colors = {'ok': 'C8E6C9', 'skip': 'FFF9C4', 'error': 'FFCDD2'}
	for row_idx, r in enumerate(row_results or [], start=2):
		status = str(r.get('status', '') or '').lower()
		fill_color = status_colors.get(status, 'FFFFFF')
		row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
		values = [
			row_idx - 1,
			str(r.get('cari', '') or ''),
			r.get('excel_row', ''),
			str(r.get('status', '') or '').upper(),
			str(r.get('reason', '') or ''),
			str(r.get('timestamp', '') or ''),
		]
		for col_idx, val in enumerate(values, 1):
			cell = ws.cell(row=row_idx, column=col_idx, value=val)
			cell.fill = row_fill

	col_widths = [6, 30, 14, 10, 50, 20]
	for col_idx, w in enumerate(col_widths, 1):
		ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

	# ── 2. Sheet: Doldurma Detayları ─────────────────────────────────────
	detail_rows = []
	for r in (row_results or []):
		for fd in (r.get('filled_data') or []):
			detail_rows.append({
				'excel_row': r.get('excel_row', ''),
				'cari': str(r.get('cari', '') or ''),
				'alan': str(fd.get('alan', '') or ''),
				'kaynak': str(fd.get('kaynak', '') or ''),
				'deger': str(fd.get('deger', '') or ''),
			})

	if detail_rows:
		ws2 = wb.create_sheet(title='Doldurma Detayları')
		d_headers = ['Excel Satır No', 'Cari / Değer', 'SAP Alan ID', 'Kaynak', 'Doldurma Değeri']
		for col_idx, h in enumerate(d_headers, 1):
			cell = ws2.cell(row=1, column=col_idx, value=h)
			cell.font = header_font
			cell.fill = header_fill
			cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
		for row_idx, d in enumerate(detail_rows, start=2):
			ws2.cell(row=row_idx, column=1, value=d['excel_row'])
			ws2.cell(row=row_idx, column=2, value=d['cari'])
			ws2.cell(row=row_idx, column=3, value=d['alan'])
			ws2.cell(row=row_idx, column=4, value=d['kaynak'])
			ws2.cell(row=row_idx, column=5, value=d['deger'])
		for col_idx, w in enumerate([14, 28, 55, 18, 40], 1):
			ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = w

	if not output_path:
		tmp = tempfile.NamedTemporaryFile(delete=False, suffix='_rapor.xlsx', prefix='saggio_rpa_')
		output_path = tmp.name
		tmp.close()

	wb.save(output_path)
	return output_path


def _notify_sap_event(notification, phase, result_payload=None):
	notification = notification or {}
	notes = []
	stamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

	tg_bot_id = notification.get('telegram_bot_id')
	tg_group_id = notification.get('telegram_group_id')
	tg_voice_enabled = bool(notification.get('telegram_voice_enabled'))
	mail_id = notification.get('mail_account_id')

	bot = TelegramBot.objects.filter(pk=tg_bot_id, is_active=True).first() if tg_bot_id else None
	group = TelegramGroup.objects.filter(pk=tg_group_id, is_active=True).first() if tg_group_id else None
	mail_account = MailAccount.objects.filter(pk=mail_id, is_active=True).first() if mail_id else None

	if bot and group:
		default_tg = f'SAP sureci {phase} ({stamp})'
		if phase == 'start':
			text = str(notification.get('telegram_start_message') or '').strip() or default_tg
		else:
			total = len(result_payload or [])
			ok_count = sum(1 for r in (result_payload or []) if r.get('ok'))
			err_count = max(0, total - ok_count)
			end_default = f'SAP süreci tamamlandı ({stamp}) | Başarılı: {ok_count}, Hatalı: {err_count}'
			text = str(notification.get('telegram_end_message') or '').strip() or end_default
		ok, msg = _send_telegram_message(bot, group.chat_id, text)
		notes.append({'channel': 'telegram', 'ok': ok, 'msg': msg})
		if tg_voice_enabled:
			v_ok, v_msg = _send_telegram_voice_message(bot, group.chat_id, text)
			if (not v_ok) and ok:
				notes.append({'channel': 'telegram_voice', 'ok': True, 'msg': f'Voice gonderilemedi ama metin gonderildi: {v_msg}'})
			else:
				notes.append({'channel': 'telegram_voice', 'ok': v_ok, 'msg': v_msg})

	if mail_account:
		to_value = str(notification.get('mail_to') or '').strip() or mail_account.email
		subject = str(notification.get('mail_subject') or '').strip() or 'Saggio RPA SAP Bildirimi'
		if phase == 'start':
			body = str(notification.get('mail_start_message') or '').strip() or f'SAP sureci basladi. Zaman: {stamp}'
			ok, msg = _send_mail_message(mail_account, to_value, subject, body)
		else:
			total = len(result_payload or [])
			ok_count = sum(1 for r in (result_payload or []) if r.get('ok'))
			err_count = max(0, total - ok_count)
			body = str(notification.get('mail_end_message') or '').strip() or f'SAP süreci tamamlandı. Başarılı: {ok_count}, Hatalı: {err_count}. Zaman: {stamp}'
			report_path = None
			if result_payload:
				try:
					report_path = _generate_row_report_xlsx(result_payload)
				except Exception:
					report_path = None
			ok, msg = _send_mail_message(mail_account, to_value, subject, body, attachment_path=report_path)
			if report_path:
				try:
					import os as _os2
					_os2.unlink(report_path)
				except Exception:
					pass
		notes.append({'channel': 'mail', 'ok': ok, 'msg': msg})

	return notes




def _safe_send_popup_mail(cfg, mail_enabled=True, runtime_state=None, notification_cfg=None, popup_title='', popup_text=''):
	if not mail_enabled:
		return 'Mail gönderimi süreç ayarında kapalı.'
	if not bool(cfg.get('send_mail_on_match')):
		return None
	rt = runtime_state if isinstance(runtime_state, dict) else {}
	notify = notification_cfg if isinstance(notification_cfg, dict) else {}

	mail_to = str(cfg.get('mail_to', '') or '').strip() or str(notify.get('mail_to', '') or '').strip()
	account_id = cfg.get('mail_account_id') or notify.get('mail_account_id')
	if not account_id:
		return 'Mail atlandi: mail hesabı seçili değil.'
	account = MailAccount.objects.filter(pk=account_id, is_active=True).first()
	if not account:
		return f'Mail atlandi: hesap bulunamadi (id={account_id}).'
	if not mail_to:
		mail_to = str(account.email or '').strip()
	if not mail_to:
		return 'Mail atlandi: alıcı bulunamadı.'

	subject = str(cfg.get('mail_subject', '') or '').strip() or f'SAP popup uyarısı: {popup_title or "Popup"}'
	body_lines = []
	custom_body = str(cfg.get('mail_body', '') or '').strip()
	if custom_body:
		body_lines.append(custom_body)
	body_lines.append(f'Popup Başlığı: {popup_title or "(boş)"}')
	body_lines.append(f'Popup Metni: {popup_text or "(boş)"}')

	memory_items = []
	for k in sorted(rt.keys(), key=lambda x: str(x)):
		ks = str(k)
		if ks.startswith('sutun_') or ks.startswith('loop_'):
			memory_items.append((ks, rt.get(k)))
	if memory_items:
		body_lines.append('')
		body_lines.append('Hafıza Değerleri:')
		for mk, mv in memory_items:
			body_lines.append(f'- {mk}: {mv}')
	body = '\n'.join(body_lines)

	try:
		msg = MIMEText(body, _charset='utf-8')
		msg['Subject'] = subject
		msg['From'] = account.email
		msg['To'] = mail_to
		if account.use_ssl:
			server = smtplib.SMTP_SSL(account.smtp_host, int(account.smtp_port or 465), timeout=20)
		else:
			server = smtplib.SMTP(account.smtp_host, int(account.smtp_port or 587), timeout=20)
		try:
			server.ehlo()
			if account.use_tls and not account.use_ssl:
				server.starttls()
				server.ehlo()
			server.login(account.smtp_username, account.get_smtp_password())
			server.sendmail(account.email, [mail_to], msg.as_string())
		finally:
			server.quit()
		return f'Mail gonderildi: {mail_to}'
	except Exception as ex:
		return f'Mail gonderim hatasi: {ex}'

