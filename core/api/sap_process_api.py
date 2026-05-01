"""SAP process admin + excel JSON endpoints + builder/list page renders.

Originally lived in ``core/views.py``.
"""
from __future__ import annotations

import json
import os
import re
from datetime import datetime

from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_POST

from ..firebase_service import SAPTemplateService
from ..models import (
	FTPAccount,
	MailAccount,
	SapProcess,
	SapProcessStep,
)
from ..utils.parsing import _as_bool
from ..utils.runtime_state import (
	_runtime_get,
	_runtime_set_controls,
	_runtime_push_log,
	_runtime_finish,
)

try:
	import tkinter as tk
except Exception:
	tk = None

try:
	from openpyxl import load_workbook
except Exception:
	load_workbook = None


def sap_process_list(request):
	"""SAP süreçlerini listele; yeni süreç oluştur (POST)."""
	if request.method == 'POST':
		name = str(request.POST.get('name', '') or '').strip()
		desc = str(request.POST.get('description', '') or '').strip()
		if not name:
			return JsonResponse({'ok': False, 'error': 'Süreç adı boş olamaz.'}, status=400)
		if SapProcess.objects.filter(name=name).exists():
			return JsonResponse({'ok': False, 'error': 'Bu isimde bir süreç zaten var.'}, status=400)
		proc = SapProcess.objects.create(name=name, description=desc)
		return JsonResponse({'ok': True, 'id': proc.pk, 'name': proc.name})
	# GET: hem JSON hem HTML desteği
	if request.headers.get('Accept') == 'application/json' or request.GET.get('fmt') == 'json':
		processes = list(SapProcess.objects.values('id', 'name', 'description', 'updated_at').order_by('name'))
		return JsonResponse({'ok': True, 'processes': processes})
	return render(request, 'core/sap_process_list.html', {
		'current': 'sap_process',
		'page_title': 'SAP Süreç Tanımları',
		'page_subtitle': 'Otomasyon akışı tanımlayın',
	})


@ensure_csrf_cookie
def sap_process_builder(request, process_id):
	"""Belirli bir sürecin adım builder sayfasını göster."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	steps = list(proc.steps.values('id', 'order', 'step_type', 'label', 'config').order_by('order'))
	template_names = SAPTemplateService.list_template_names()
	ftp_accounts = list(FTPAccount.objects.filter(is_active=True).values('id', 'name', 'protocol', 'host', 'port').order_by('name'))
	other_processes = list(SapProcess.objects.exclude(pk=proc.pk).values('id', 'name').order_by('name'))
	mail_accounts = list(MailAccount.objects.filter(is_active=True).values('id', 'name', 'email').order_by('name'))
	return render(request, 'core/sap_process_builder.html', {
		'current': 'sap_process',
		'page_title': f'Süreç: {proc.name}',
		'page_subtitle': 'Adım adım otomasyon akışı',
		'process': proc,
		'steps_json': json.dumps(steps, default=str),
		'template_names': template_names,
		'ftp_accounts_json': json.dumps(ftp_accounts, default=str),
		'processes_json': json.dumps(other_processes, default=str),
		'mail_accounts_json': json.dumps(mail_accounts, default=str),
	})


def sap_process_backup(request, process_id):
	"""Süreç tanımını ve adımlarını JSON dosyası olarak indir."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	steps = list(proc.steps.values('id', 'order', 'step_type', 'label', 'config').order_by('order'))
	now = datetime.now()
	payload = {
		'version': 1,
		'exported_at': now.isoformat(),
		'process': {
			'id': proc.pk,
			'name': proc.name,
			'description': proc.description,
			'ghost_overlay_enabled': bool(proc.ghost_overlay_enabled),
			'office_express_auto_close': bool(proc.office_express_auto_close),
			'telegram_notifications_enabled': bool(proc.telegram_notifications_enabled),
			'telegram_voice_enabled': bool(proc.telegram_voice_enabled),
			'mail_notifications_enabled': bool(proc.mail_notifications_enabled),
		},
		'steps': steps,
	}
	safe_name = re.sub(r'[^0-9A-Za-z_-]+', '_', str(proc.name or f'process_{proc.pk}')).strip('_') or f'process_{proc.pk}'
	filename = f'sap_process_{proc.pk}_{safe_name}_{now.strftime("%Y%m%d_%H%M%S")}.json'
	body = json.dumps(payload, ensure_ascii=False, indent=2, default=str)
	resp = HttpResponse(body, content_type='application/json; charset=utf-8')
	resp['Content-Disposition'] = f'attachment; filename="{filename}"'
	return resp


def sap_process_delete(request, process_id):
	"""Süreci sil."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	proc.delete()
	return JsonResponse({'ok': True})


@require_POST
def sap_process_step_save(request, process_id):
	"""Adımları toplu kaydet (tam liste — mevcut adımları sil, yenileri ekle)."""
	try:
		proc = get_object_or_404(SapProcess, pk=process_id)
		try:
			body = json.loads(request.body)
		except (json.JSONDecodeError, TypeError):
			return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

		steps_data = body.get('steps', [])
		if not isinstance(steps_data, list):
			return JsonResponse({'ok': False, 'error': 'steps bir liste olmalı.'}, status=400)

		valid_types = {c[0] for c in SapProcessStep.STEP_TYPE_CHOICES}
		proc.steps.all().delete()
		saved = []
		for i, s in enumerate(steps_data):
			step_type = str(s.get('step_type', '') or '').strip()
			if step_type not in valid_types:
				continue
			step = SapProcessStep.objects.create(
				process=proc,
				order=i,
				step_type=step_type,
				label=str(s.get('label', '') or '').strip()[:300],
				config=s.get('config', {}) if isinstance(s.get('config'), dict) else {},
			)
			saved.append({'id': step.pk, 'order': step.order, 'step_type': step.step_type})

		return JsonResponse({'ok': True, 'saved': len(saved), 'steps': saved})
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Adımlar kaydedilemedi: {ex}'}, status=500)


@require_POST
def sap_process_rename(request, process_id):
	"""Süreç adını / açıklamasını güncelle."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)
	name = str(body.get('name', '') or '').strip()
	desc = str(body.get('description', '') or '').strip()
	if not name:
		return JsonResponse({'ok': False, 'error': 'Süreç adı boş olamaz.'}, status=400)
	if SapProcess.objects.exclude(pk=process_id).filter(name=name).exists():
		return JsonResponse({'ok': False, 'error': 'Bu isimde başka bir süreç var.'}, status=400)
	proc.name = name
	proc.description = desc
	proc.save(update_fields=['name', 'description', 'updated_at'])
	return JsonResponse({'ok': True, 'name': proc.name})


@require_POST
def sap_process_runtime_settings_save(request, process_id):
	"""Süreç çalışma ayarlarını (overlay + popup + bildirim) güncelle."""
	proc = get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	proc.ghost_overlay_enabled = _as_bool(body.get('ghost_overlay_enabled', proc.ghost_overlay_enabled), proc.ghost_overlay_enabled)
	proc.office_express_auto_close = _as_bool(body.get('office_express_auto_close', proc.office_express_auto_close), proc.office_express_auto_close)
	proc.telegram_notifications_enabled = _as_bool(body.get('telegram_notifications_enabled', proc.telegram_notifications_enabled), proc.telegram_notifications_enabled)
	proc.telegram_voice_enabled = _as_bool(body.get('telegram_voice_enabled', proc.telegram_voice_enabled), proc.telegram_voice_enabled)
	proc.mail_notifications_enabled = _as_bool(body.get('mail_notifications_enabled', proc.mail_notifications_enabled), proc.mail_notifications_enabled)
	proc.sap_retry_enabled = _as_bool(body.get('sap_retry_enabled', proc.sap_retry_enabled), proc.sap_retry_enabled)
	try:
		_ri = int(body.get('sap_retry_interval_minutes', proc.sap_retry_interval_minutes))
		if _ri >= 1:
			proc.sap_retry_interval_minutes = _ri
	except (TypeError, ValueError):
		pass
	try:
		_md = int(body.get('sap_retry_max_duration_minutes', proc.sap_retry_max_duration_minutes))
		if _md >= 0:
			proc.sap_retry_max_duration_minutes = _md
	except (TypeError, ValueError):
		pass

	proc.save(update_fields=[
		'ghost_overlay_enabled',
		'office_express_auto_close',
		'telegram_notifications_enabled',
		'telegram_voice_enabled',
		'mail_notifications_enabled',
		'sap_retry_enabled',
		'sap_retry_interval_minutes',
		'sap_retry_max_duration_minutes',
		'updated_at',
	])

	return JsonResponse({
		'ok': True,
		'ghost_overlay_enabled': proc.ghost_overlay_enabled,
		'office_express_auto_close': proc.office_express_auto_close,
		'telegram_notifications_enabled': proc.telegram_notifications_enabled,
		'telegram_voice_enabled': proc.telegram_voice_enabled,
		'mail_notifications_enabled': proc.mail_notifications_enabled,
		'sap_retry_enabled': proc.sap_retry_enabled,
		'sap_retry_interval_minutes': proc.sap_retry_interval_minutes,
		'sap_retry_max_duration_minutes': proc.sap_retry_max_duration_minutes,
	})


@require_POST
def sap_process_runtime_control(request, process_id):
	"""Canlı süreç kontrolü: duraklat/devam et/durdur."""
	get_object_or_404(SapProcess, pk=process_id)
	try:
		body = json.loads(request.body)
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	action = str(body.get('action', '') or '').strip().casefold()
	state = _runtime_get(process_id)
	if state is None:
		if action == 'force_stop':
			# Kayıt yoksa zaten temiz; başarı dön.
			return JsonResponse({'ok': True, 'state': {'running': False, 'paused': False, 'stop_requested': False}})
		return JsonResponse({'ok': False, 'error': 'Bu süreç için aktif runtime bulunamadı.'}, status=404)

	if action == 'pause_toggle':
		_runtime_set_controls(process_id, paused=not bool(state.get('paused')), stop_requested=bool(state.get('stop_requested')))
	elif action == 'pause':
		_runtime_set_controls(process_id, paused=True)
	elif action == 'resume':
		_runtime_set_controls(process_id, paused=False)
	elif action == 'stop':
		_runtime_set_controls(process_id, paused=False, stop_requested=True)
	elif action == 'force_stop':
		# Sıkışmış (orphan) runtime kaydını anında temizle.
		_runtime_set_controls(process_id, paused=False, stop_requested=True)
		_runtime_push_log(process_id, 'Kullanıcı zorla sıfırlama yaptı; runtime durumu temizlendi.')
		_runtime_finish(process_id)
	else:
		return JsonResponse({'ok': False, 'error': f'Desteklenmeyen aksiyon: {action}'}, status=400)

	return JsonResponse({'ok': True, 'state': _runtime_get(process_id)})


def sap_process_runtime_status(request, process_id):
	"""Canlı süreç durumunu ve hayalet log akışını döndürür."""
	get_object_or_404(SapProcess, pk=process_id)
	state = _runtime_get(process_id)
	if state is None:
		return JsonResponse({'ok': True, 'state': {'running': False, 'paused': False, 'stop_requested': False, 'current_step': 0, 'total_steps': 0, 'step_name': '', 'logs': []}})
	return JsonResponse({'ok': True, 'state': state})

@require_POST
def sap_process_excel_browse(request, process_id):
	"""Sunucu makinesinde Excel dosyası seçtirir."""
	get_object_or_404(SapProcess, pk=process_id)
	if tk is None:
		return JsonResponse({'ok': False, 'error': 'Dosya seçici için tkinter kullanılamıyor.'}, status=500)

	try:
		body = json.loads(request.body) if request.body else {}
	except (json.JSONDecodeError, TypeError):
		body = {}

	initial_path = str((body or {}).get('initial_path', '') or '').strip()
	initial_dir = ''
	if initial_path:
		initial_path = os.path.expandvars(os.path.expanduser(initial_path))
		if os.path.isdir(initial_path):
			initial_dir = initial_path
		elif os.path.isfile(initial_path):
			initial_dir = os.path.dirname(initial_path)

	root = None
	try:
		from tkinter import filedialog
		root = tk.Tk()
		root.withdraw()
		try:
			root.attributes('-topmost', True)
		except Exception:
			pass
		picked = filedialog.askopenfilename(
			title='Excel dosyası seç',
			filetypes=[('Excel Dosyalari', '*.xlsx *.xlsm *.xls'), ('Tum Dosyalar', '*.*')],
			initialdir=initial_dir or None,
		)
	finally:
		if root is not None:
			try:
				root.destroy()
			except Exception:
				pass

	if not picked:
		return JsonResponse({'ok': True, 'cancelled': True, 'path': ''})
	return JsonResponse({'ok': True, 'cancelled': False, 'path': str(picked)})


@require_POST
def sap_process_excel_sheets(request, process_id):
	"""Verilen Excel dosyasındaki sayfa adlarını döndürür."""
	get_object_or_404(SapProcess, pk=process_id)
	if load_workbook is None:
		return JsonResponse({'ok': False, 'error': 'Excel okuma için openpyxl kurulu değil.'}, status=500)

	try:
		body = json.loads(request.body) if request.body else {}
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	excel_path = str((body or {}).get('excel_file_path', '') or '').strip()
	if not excel_path:
		return JsonResponse({'ok': False, 'error': 'Excel dosya yolu boş.'}, status=400)
	excel_path = os.path.expandvars(os.path.expanduser(excel_path))
	if not os.path.isfile(excel_path):
		return JsonResponse({'ok': False, 'error': f'Excel dosyası bulunamadı: {excel_path}'}, status=404)

	try:
		wb = load_workbook(excel_path, data_only=True, read_only=True)
		sheets = list(wb.sheetnames or [])
		wb.close()
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Excel okunamadı: {ex}'}, status=500)

	return JsonResponse({'ok': True, 'sheets': sheets, 'count': len(sheets)})


@require_POST
def sap_process_excel_columns(request, process_id):
	"""Verilen Excel dosyasında seçili sayfanın başlık satırına göre sütun listesini döndürür."""
	get_object_or_404(SapProcess, pk=process_id)
	if load_workbook is None:
		return JsonResponse({'ok': False, 'error': 'Excel okuma için openpyxl kurulu değil.'}, status=500)

	try:
		body = json.loads(request.body) if request.body else {}
	except (json.JSONDecodeError, TypeError):
		return JsonResponse({'ok': False, 'error': 'Geçersiz JSON.'}, status=400)

	excel_path = str((body or {}).get('excel_file_path', '') or '').strip()
	if not excel_path:
		return JsonResponse({'ok': False, 'error': 'Excel dosya yolu boş.'}, status=400)
	excel_path = os.path.expandvars(os.path.expanduser(excel_path))
	if not os.path.isfile(excel_path):
		return JsonResponse({'ok': False, 'error': f'Excel dosyası bulunamadı: {excel_path}'}, status=404)

	sheet_name = str((body or {}).get('excel_sheet_name', '') or '').strip()
	try:
		header_row = max(1, int((body or {}).get('excel_header_row') or 1))
	except (TypeError, ValueError):
		header_row = 1

	def _column_letter(col_index):
		s = ''
		n = int(col_index or 0)
		while n > 0:
			n, r = divmod(n - 1, 26)
			s = chr(65 + r) + s
		return s or 'A'

	try:
		wb = load_workbook(excel_path, data_only=True, read_only=True)
		if sheet_name:
			if sheet_name not in wb.sheetnames:
				wb.close()
				return JsonResponse({'ok': False, 'error': f'Sayfa bulunamadı: {sheet_name}'}, status=404)
			ws = wb[sheet_name]
		else:
			ws = wb[wb.sheetnames[0]]

		max_col = int(getattr(ws, 'max_column', 0) or 0)
		if max_col <= 0:
			try:
				header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), ())
				max_col = len(tuple(header_values or ()))
			except Exception:
				max_col = 0
		max_col = max(1, min(max_col, 300))
		columns = []
		for idx in range(1, max_col + 1):
			letter = _column_letter(idx)
			head_val = ws.cell(row=header_row, column=idx).value
			head = str(head_val or '').strip()
			label = f'{letter} - {head}' if head else letter
			columns.append({'value': letter, 'label': label, 'header': head})
		wb.close()
	except Exception as ex:
		return JsonResponse({'ok': False, 'error': f'Sütunlar okunamadı: {ex}'}, status=500)

	return JsonResponse({'ok': True, 'columns': columns, 'count': len(columns)})


